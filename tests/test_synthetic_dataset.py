from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from village_insight.question_benchmark import load_question_benchmark
from village_insight.synthetic_dataset import (
    PARTY_SPEC,
    POPULATION_SPEC,
    TEMPLATE_SPECS,
    generate_dataset,
)


def test_generated_dataset_is_deterministic_and_template_shaped(tmp_path: Path) -> None:
    first = tmp_path / "first"
    second = tmp_path / "second"

    first_manifest = generate_dataset(first)
    second_manifest = generate_dataset(second)

    assert first_manifest == second_manifest
    assert first_manifest["record_count"] == 300
    assert first_manifest["question_count"] >= 150
    assert all(row["size_bytes"] < 100 * 1024 * 1024 for row in first_manifest["files"])

    for spec in TEMPLATE_SPECS:
        first_path = first / "data" / spec.filename
        second_path = second / "data" / spec.filename
        assert first_path.read_bytes() == second_path.read_bytes()
        workbook = load_workbook(first_path, read_only=True, data_only=False)
        assert workbook.sheetnames == [spec.sheet_name]
        sheet = workbook[spec.sheet_name]
        values = list(sheet.iter_rows(values_only=True))
        assert values[0] == spec.headers
        assert len(values) - 1 == spec.record_count
        assert not workbook._external_links
        assert all(
            not (isinstance(value, str) and value.startswith("="))
            for row in values[1:]
            for value in row
        )
        workbook.close()
        with ZipFile(first_path) as archive:
            assert not any(name.startswith("xl/externalLinks/") for name in archive.namelist())
            assert b"<f" not in archive.read("xl/worksheets/sheet1.xml")


def test_generated_identifiers_are_explicitly_synthetic(tmp_path: Path) -> None:
    generate_dataset(tmp_path)

    population = load_workbook(
        tmp_path / "data" / POPULATION_SPEC.filename,
        read_only=True,
        data_only=True,
    )[POPULATION_SPEC.sheet_name]
    population_rows = list(population.iter_rows(min_row=2, values_only=True))
    assert {row[1] for row in population_rows} == {"演示一村"}
    assert all(str(row[2]).startswith("DEMO-HH-") for row in population_rows)
    assert all(str(row[4]).startswith("演示居民") for row in population_rows)
    assert all(str(row[7]).startswith("TEST-ID-") for row in population_rows)

    party = load_workbook(
        tmp_path / "data" / PARTY_SPEC.filename,
        read_only=True,
        data_only=True,
    )[PARTY_SPEC.sheet_name]
    party_rows = list(party.iter_rows(min_row=2, values_only=True))
    assert all(str(row[1]).startswith("演示居民") for row in party_rows)
    assert all(str(row[8]).startswith("TEST-ID-") for row in party_rows)
    assert all(str(row[12]).startswith("TEST-PHONE-") for row in party_rows)


def test_generated_question_workbook_uses_benchmark_contract(tmp_path: Path) -> None:
    manifest = generate_dataset(tmp_path)
    corpus = load_question_benchmark(tmp_path / "questions.xlsx")
    payload = json.loads((tmp_path / "questions.json").read_text(encoding="utf-8"))

    assert corpus.question_row_count == manifest["question_count"]
    assert corpus.unique_question_count == manifest["question_count"]
    assert payload["case_count"] == manifest["question_count"]
    assert sum(
        case["comparison"] == "policy" for case in payload["cases"]
    ) == manifest["question_category_counts"]["sensitive_permission_blocked"]
