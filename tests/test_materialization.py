import hashlib
import uuid
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from village_insight.db.base import Base
from village_insight.db.models import (
    ApprovedImportPlan,
    DatasetRecord,
    DocumentProfile,
    DocumentTemplate,
    FormalImportStatus,
    ImportExecution,
    IngestionBatch,
    IngestionItem,
    ItemStatus,
    Job,
    MetricDefinition,
    QualityIssue,
    RecordIndexValue,
    RecordValueLineage,
    SemanticField,
    SemanticFieldVersion,
    TemplateMatch,
    TemplateStatus,
    TemplateVersion,
)
from village_insight.materialization import (
    _include_selected_raw_evidence,
    _is_unfilled_form_row,
    _normalized_value,
    _resolve_column,
    materialize_plan,
)
from village_insight.parsing.candidates import select_header_candidates
from village_insight.parsing.router import ParserRouter
from village_insight.questions import (
    MetricFilter,
    MetricQuery,
    MetricQueryScope,
    execute_metric_query,
)
from village_insight.reimport import reset_item_for_reimport
from village_insight.templates.import_plans import approve_plan
from village_insight.templates.matching import layout_fingerprint


def test_selected_semantic_cell_is_always_present_in_raw_projection() -> None:
    cell = SimpleNamespace(
        id="cell-2",
        column=2,
        coordinate="B2",
        raw_value="value",
        display_value="value",
    )
    raw_columns = {
        "column-1": {
            "header_path": ["已有列"],
            "source_cell": {"id": "cell-1", "coordinate": "A2"},
        }
    }

    _include_selected_raw_evidence(
        raw_columns,
        [
            (
                {"source_column_id": "column-2", "header_path": ["映射列"]},
                cell,
            )
        ],
    )
    _include_selected_raw_evidence(
        raw_columns,
        [
            (
                {"source_column_id": "column-2", "header_path": ["映射列"]},
                cell,
            )
        ],
    )

    assert raw_columns["column-2"]["source_cell"] == {
        "id": "cell-2",
        "coordinate": "B2",
        "raw_value": "value",
        "display_value": "value",
    }
    assert len(raw_columns) == 2


def test_column_resolution_is_pinned_to_the_approved_header(tmp_path: Path) -> None:
    source = tmp_path / "competing-headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["标题", None])
    sheet.append(["姓名", "人数"])
    sheet.append(["张三", 1])
    workbook.save(source)
    workbook.close()

    profile = ParserRouter().profile(source)
    candidates = profile.sheets[0].header_candidates
    approved_header = next(candidate for candidate in candidates if candidate.header_rows == [1, 2])
    repeated_column = next(
        column for column in approved_header.columns if column.header_path == ["人数"]
    )

    _, resolved_header, _, _ = _resolve_column(
        profile,
        repeated_column.source_column_id,
        repeated_column.header_path,
        expected_sheet_id=profile.sheets[0].id,
        expected_region_id=approved_header.region_id,
        expected_header_id=approved_header.id,
    )

    assert resolved_header.id == approved_header.id


def test_unfilled_form_prompts_are_not_business_rows() -> None:
    assert _is_unfilled_form_row(
        {
            "name": {"source_cell": {"display_value": "姓名"}},
            "village": {"source_cell": {"display_value": "村"}},
        }
    )
    assert _is_unfilled_form_row(
        {
            "promise": {
                "source_cell": {
                    "display_value": "（二）酒后不驾车。",
                }
            }
        }
    )
    assert _is_unfilled_form_row(
        {
            "signature": {
                "source_cell": {
                    "display_value": "当事人\n签名：                                ",
                }
            }
        }
    )
    assert _is_unfilled_form_row(
        {
            "driver": {"source_cell": {"display_value": "驾驶员姓名"}},
            "student": {"source_cell": {"display_value": "学生姓名"}},
        }
    )
    assert not _is_unfilled_form_row(
        {
            "name": {"source_cell": {"display_value": "张三"}},
            "village": {"source_cell": {"display_value": "法乐村"}},
        }
    )


def test_decimal_normalization_matches_typed_index_storage() -> None:
    normalizer, value = _normalized_value("decimal", "22.75068493150685")
    assert normalizer == "decimal/v1"
    assert value == Decimal("22.7506849315")

    _, positive_tie = _normalized_value("decimal", "1.00000000005")
    _, negative_tie = _normalized_value("decimal", "-1.00000000005")
    assert positive_tie == Decimal("1.0000000001")
    assert negative_tie == Decimal("-1.0000000001")


def test_materialization_is_typed_lineaged_and_idempotent(tmp_path: Path) -> None:
    source = tmp_path / "synthetic.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "人数", "负责人", "备注"])
    sheet.append(["张三", 2, "王五", "低保户"])
    sheet.append(["李四", 3, "赵六", ""])
    sheet.append(["钱七", "无法识别", "孙八", "待核验"])
    sheet.append(["合计：", 5, None, None])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        name_field = SemanticField(code="person.name", published_version=1)
        name_field.versions.append(
            SemanticFieldVersion(
                version=1,
                name="姓名",
                layer="base",
                data_type="text",
                status=TemplateStatus.PUBLISHED,
            )
        )
        count_field = SemanticField(code="person.count", published_version=1)
        count_field.versions.append(
            SemanticFieldVersion(
                version=1,
                name="人数",
                layer="domain",
                data_type="integer",
                status=TemplateStatus.PUBLISHED,
            )
        )
        template = DocumentTemplate(code="people", published_version=1)
        template.versions.append(
            TemplateVersion(
                version=1,
                name="人员表",
                status=TemplateStatus.PUBLISHED,
                layout_fingerprint=layout_fingerprint(profile),
                definition={
                    "contract_version": "document-template/v1",
                    "domain": "population",
                    "region_kind": "table",
                    "record_type": "person",
                    "record_grain": "one_row_per_person",
                    "field_bindings": [
                        {
                            "source_column_id": "old-file-column-1",
                            "header_path": ["姓名"],
                            "semantic_field_code": "person.name",
                            "semantic_field_version": 1,
                            "required": True,
                        },
                        {
                            "source_column_id": "old-file-column-3",
                            "header_path": ["负责人"],
                            "semantic_field_code": "person.name",
                            "semantic_field_version": 1,
                            "role": "responsible_person",
                            "required": True,
                        },
                        {
                            "source_column_id": "old-file-column-2",
                            "header_path": ["人数"],
                            "semantic_field_code": "person.count",
                            "semantic_field_version": 1,
                            "required": True,
                        },
                    ],
                    "data_row_rules": [],
                    "exclusion_rules": [],
                    "metric_codes": [],
                },
                source="bootstrap",
            )
        )
        database.add_all([name_field, count_field, template])
        batch = IngestionBatch(name="synthetic", total_files=1)
        database.add(batch)
        database.flush()
        payload = source.read_bytes()
        item = IngestionItem(
            id=uuid.uuid4(),
            batch_id=batch.id,
            original_name=source.name,
            source_path=str(source),
            source_sha256=hashlib.sha256(payload).hexdigest(),
            size_bytes=len(payload),
        )
        database.add(item)
        database.flush()
        database.add(
            DocumentProfile(
                item_id=item.id,
                contract_version=profile.contract_version,
                source_sha256=profile.source_sha256,
                parser_name=profile.parser_name,
                parser_version=profile.parser_version,
                profile=profile.model_dump(mode="json"),
            )
        )
        database.add(
            TemplateMatch(
                item_id=item.id,
                source_sha256=profile.source_sha256,
                profile_contract_version=profile.contract_version,
                layout_fingerprint=layout_fingerprint(profile),
                match_type="exact",
                score_basis_points=10_000,
                template_id=template.id,
                template_version=1,
                differences={"new_headers": [], "missing_headers": []},
                requires_hermes=False,
                matcher_version="layout-matcher/v1",
            )
        )
        database.flush()
        selected_header = select_header_candidates(profile.sheets[0].header_candidates)[0]
        selected_region = profile.sheets[0].region_candidates[0]
        plan = approve_plan(
            database,
            item=item,
            template_id=template.id,
            template_version=1,
            layout_plan={
                "contract_version": "approved-layout-plan/v1",
                "decisions": [
                    {
                        "region_candidate_id": selected_region.id,
                        "header_candidate_id": selected_header.id,
                        "data_start_row": 2,
                        "data_end_row": 4,
                        "excluded_rows": [1],
                        "classification": "table",
                    }
                ],
            },
            field_mappings=[],
            actor="tester",
            comment="synthetic",
        )
        assert plan.layout_plan["decisions"][0]["excluded_rows"] == []
        first = materialize_plan(database, plan.id)
        second = materialize_plan(database, plan.id)

        assert first.id == second.id
        assert first.status == "partial"
        assert first.record_count == 3
        assert first.value_count == 8
        assert database.query(DatasetRecord).count() == 3
        records = database.query(DatasetRecord).order_by(DatasetRecord.source_row)
        assert records[0].raw_data["columns"]
        assert len(records[0].raw_data["columns"]) == 4
        assert records[0].mapping_status == "partial"
        assert records[0].semantic_data["fields"]["person.count"]["$value"]["value"] == 2
        assert records[2].quality_status == "failed"
        assert any(
            column["source_cell"]["display_value"] == "无法识别"
            for column in records[2].raw_data["columns"].values()
        )
        counts = database.query(RecordIndexValue).filter(
            RecordIndexValue.semantic_field_code == "person.count"
        )
        assert [value.integer_value for value in counts] == [2, 3]
        lineage = database.query(RecordValueLineage).all()
        assert len(lineage) == 8
        assert {entry.coordinate for entry in lineage} == {
            "A2",
            "B2",
            "C2",
            "A3",
            "B3",
            "C3",
            "A4",
            "C4",
        }
        issues = database.query(QualityIssue).all()
        assert len(issues) == 1
        assert issues[0].evidence["coordinate"] == "B4"
        roles = database.query(RecordIndexValue).filter(
            RecordIndexValue.semantic_field_code == "person.name"
        )
        assert {(value.text_value, value.role) for value in roles} == {
            ("张三", ""),
            ("李四", ""),
            ("钱七", ""),
            ("王五", "responsible_person"),
            ("赵六", "responsible_person"),
            ("孙八", "responsible_person"),
        }

        database.add(
            MetricDefinition(
                code="population.total",
                name="总人数",
                semantic_field_code="person.count",
                semantic_field_version=1,
                aggregation="sum",
                unit="人",
                allowed_filter_fields=["person.name"],
            )
        )
        database.flush()
        total = execute_metric_query(
            database,
            MetricQuery(metric_code="population.total"),
            MetricQueryScope(
                tenant_id=batch.tenant_id,
                administrative_unit_ids=frozenset({batch.administrative_unit_id}),
            ),
        )
        filtered = execute_metric_query(
            database,
            MetricQuery(
                metric_code="population.total",
                filters=[
                    MetricFilter(
                        field_code="person.name",
                        value="李四",
                    )
                ],
            ),
            MetricQueryScope(
                tenant_id=batch.tenant_id,
                administrative_unit_ids=frozenset({batch.administrative_unit_id}),
            ),
        )
        assert total.value == 5
        assert total.record_count == 2
        assert filtered.value == 3
        assert filtered.record_count == 1

        reset = reset_item_for_reimport(database, item.id)

        assert reset.id == item.id
        assert reset.status == ItemStatus.PENDING
        assert reset.formal_import_status == FormalImportStatus.PENDING
        assert database.query(DatasetRecord).count() == 0
        assert database.query(RecordIndexValue).count() == 0
        assert database.query(RecordValueLineage).count() == 0
        assert database.query(QualityIssue).count() == 0
        assert database.query(ImportExecution).count() == 0
        assert database.query(ApprovedImportPlan).count() == 0
        assert database.query(DocumentProfile).count() == 0
        assert database.query(TemplateMatch).count() == 0
        jobs = database.query(Job).all()
        assert len(jobs) == 1
        assert jobs[0].kind == "PROFILE_FILE"
        assert jobs[0].payload == {"reimport": True, "item_id": str(item.id)}
        assert database.query(DocumentTemplate).count() == 1


def test_form_and_headerless_table_keep_cell_lineage(tmp_path: Path) -> None:
    source = tmp_path / "form-and-headerless.xlsx"
    workbook = Workbook()
    form_sheet = workbook.active
    form_sheet.title = "申请表"
    form_sheet.append(["姓名", "张三", "性别", "男"])
    form_sheet.append(["电话", "13800000000", "住址", "一组"])
    list_sheet = workbook.create_sheet("任务数")
    list_sheet.append(["王五", 3])
    list_sheet.append(["赵六", 5])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        fields = [
            ("person.name", "姓名", "text"),
            ("person.sex", "性别", "text"),
            ("person.phone", "电话", "text"),
            ("household.address", "住址", "text"),
            ("governance.task_count", "任务数", "integer"),
        ]
        for code, name, data_type in fields:
            field = SemanticField(code=code, published_version=1)
            field.versions.append(
                SemanticFieldVersion(
                    version=1,
                    name=name,
                    layer="domain",
                    data_type=data_type,
                    status=TemplateStatus.PUBLISHED,
                )
            )
            database.add(field)
        template = DocumentTemplate(code="mixed-layout", published_version=1)
        template.versions.append(
            TemplateVersion(
                version=1,
                name="混合布局",
                status=TemplateStatus.PUBLISHED,
                layout_fingerprint=layout_fingerprint(profile),
                definition={
                    "contract_version": "document-template/v1",
                    "domain": "test",
                    "region_kind": "form",
                    "record_type": "mixed_record",
                    "record_grain": "approved_layout",
                    "field_bindings": [],
                },
                source="test",
            )
        )
        database.add(template)
        batch = IngestionBatch(name="mixed-layout", total_files=1)
        database.add(batch)
        database.flush()
        payload = source.read_bytes()
        item = IngestionItem(
            id=uuid.uuid4(),
            batch_id=batch.id,
            original_name=source.name,
            source_path=str(source),
            source_sha256=hashlib.sha256(payload).hexdigest(),
            size_bytes=len(payload),
        )
        database.add(item)
        database.flush()
        database.add(
            DocumentProfile(
                item_id=item.id,
                contract_version=profile.contract_version,
                source_sha256=profile.source_sha256,
                parser_name=profile.parser_name,
                parser_version=profile.parser_version,
                profile=profile.model_dump(mode="json"),
            )
        )
        database.add(
            TemplateMatch(
                item_id=item.id,
                source_sha256=profile.source_sha256,
                profile_contract_version=profile.contract_version,
                layout_fingerprint=layout_fingerprint(profile),
                match_type="exact",
                score_basis_points=10_000,
                template_id=template.id,
                template_version=1,
                differences={},
                requires_hermes=False,
                matcher_version="test",
            )
        )
        database.flush()
        form = profile.sheets[0]
        form_region = form.region_candidates[0]
        form_header = form.header_candidates[0]
        headerless = profile.sheets[1]
        headerless_region = headerless.region_candidates[0]
        headerless_header = headerless.header_candidates[0]

        def mapping(
            *,
            sheet_id: str,
            region_id: str,
            source_column_id: str,
            header: str,
            field_code: str,
            selector: dict[str, object],
        ) -> dict[str, object]:
            return {
                "sheet_id": sheet_id,
                "region_id": region_id,
                "source_column_id": source_column_id,
                "header_path": [header],
                "semantic_field_code": field_code,
                "semantic_field_version": 1,
                "source_selector": selector,
                "required": True,
            }

        plan = approve_plan(
            database,
            item=item,
            template_id=template.id,
            template_version=1,
            layout_plan={
                "contract_version": "approved-layout-plan/v1",
                "decisions": [
                    {
                        "region_candidate_id": form_region.id,
                        "header_candidate_id": form_header.id,
                        "data_start_row": 1,
                        "data_end_row": 2,
                        "layout_mode": "form",
                        "classification": "form",
                    },
                    {
                        "region_candidate_id": headerless_region.id,
                        "header_candidate_id": headerless_header.id,
                        "data_start_row": 1,
                        "data_end_row": 2,
                        "layout_mode": "headerless_table",
                        "classification": "table",
                    },
                ],
            },
            field_mappings=[
                mapping(
                    sheet_id=form.id,
                    region_id=form_region.id,
                    source_column_id="form:name",
                    header="姓名",
                    field_code="person.name",
                    selector={"kind": "cell", "row": 1, "column": 2},
                ),
                mapping(
                    sheet_id=form.id,
                    region_id=form_region.id,
                    source_column_id="form:sex",
                    header="性别",
                    field_code="person.sex",
                    selector={"kind": "cell", "row": 1, "column": 4},
                ),
                mapping(
                    sheet_id=form.id,
                    region_id=form_region.id,
                    source_column_id="form:phone",
                    header="电话",
                    field_code="person.phone",
                    selector={"kind": "cell", "row": 2, "column": 2},
                ),
                mapping(
                    sheet_id=form.id,
                    region_id=form_region.id,
                    source_column_id="form:address",
                    header="住址",
                    field_code="household.address",
                    selector={"kind": "cell", "row": 2, "column": 4},
                ),
                mapping(
                    sheet_id=headerless.id,
                    region_id=headerless_region.id,
                    source_column_id="headerless:name",
                    header="村干部姓名",
                    field_code="person.name",
                    selector={"kind": "physical_column", "column": 1},
                ),
                mapping(
                    sheet_id=headerless.id,
                    region_id=headerless_region.id,
                    source_column_id="headerless:task-count",
                    header="任务数",
                    field_code="governance.task_count",
                    selector={"kind": "physical_column", "column": 2},
                ),
            ],
            actor="tester",
            comment="验证表单和无表头小表契约",
        )
        execution = materialize_plan(database, plan.id)

        records = database.query(DatasetRecord).order_by(
            DatasetRecord.sheet_id,
            DatasetRecord.source_row,
        )
        assert execution.status == "completed"
        assert execution.record_count == 3
        assert execution.value_count == 8
        assert {record.raw_data["layout_mode"] for record in records} == {
            "form",
            "headerless_table",
        }
        form_record = next(record for record in records if record.raw_data["layout_mode"] == "form")
        assert len(form_record.raw_data["columns"]) == 8
        assert form_record.semantic_data["fields"]["person.name"]["$value"]["coordinate"] == "B1"
        lineage = database.query(RecordValueLineage).all()
        assert len(lineage) == 8
        assert {entry.coordinate for entry in lineage} == {
            "A1",
            "A2",
            "B1",
            "B2",
            "D1",
            "D2",
        }
