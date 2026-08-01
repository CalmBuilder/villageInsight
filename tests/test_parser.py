import io
import zipfile
from copy import copy
from pathlib import Path

import pytest
import xlwt
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from village_insight.parsing.candidates import select_header_candidates
from village_insight.parsing.detection import DocumentDetectionError, detect_document
from village_insight.parsing.ooxml_repair import repair_safe_ooxml_packaging
from village_insight.parsing.router import ParserRouter


def make_structured_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "基本情况"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "人口概况"
    sheet["A2"] = "姓名"
    sheet["B2"] = "户数"
    sheet["A3"] = "张三"
    sheet["B3"] = 18
    sheet["C3"] = "=B3*2"
    bold_font = copy(sheet["A2"].font)
    bold_font.bold = True
    sheet["A2"].font = bold_font
    sheet.row_dimensions[3].hidden = True
    sheet["E8"] = "第二块"
    sheet["E9"] = "值"
    workbook.save(path)
    workbook.close()


def test_xlsx_profile_preserves_physical_evidence_and_builds_candidates(
    tmp_path: Path,
) -> None:
    source = tmp_path / "村情.xlsx"
    make_structured_workbook(source)

    profile = ParserRouter().profile(source)

    assert profile.contract_version == "workbook-profile/v2"
    assert profile.detection.format == "xlsx"
    assert profile.workbook_id == f"workbook:{profile.source_sha256}"
    sheet = profile.sheets[0]
    assert sheet.merges[0].range == "A1:B1"
    cells = {cell.coordinate: cell for cell in sheet.cells}
    assert cells["C3"].formula == "=B3*2"
    assert cells["A3"].hidden is True
    assert cells["A2"].style_ref is not None
    assert len(sheet.region_candidates) == 2
    first_region_headers = [
        candidate
        for candidate in sheet.header_candidates
        if candidate.region_id == sheet.region_candidates[0].id
    ]
    assert any(
        column.header_path == ["人口概况", "姓名"]
        for candidate in first_region_headers
        for column in candidate.columns
    )


def test_header_selection_excludes_numeric_total_and_signature_region(
    tmp_path: Path,
) -> None:
    source = tmp_path / "带合计行.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "人数", "金额"])
    sheet.append(["张三", 1, 580])
    for _ in range(8):
        sheet.append([None, None, None])
    sheet.append(["合计", 1, 580])
    sheet.append(["领导签字：", None, None])
    workbook.save(source)
    workbook.close()

    profile = ParserRouter().profile(source)
    selected = select_header_candidates(profile.sheets[0].header_candidates)

    assert [
        [" / ".join(column.header_path) for column in candidate.columns] for candidate in selected
    ] == [["姓名", "人数", "金额"]]


def test_region_candidate_absorbs_disconnected_cells_inside_table_bounds(
    tmp_path: Path,
) -> None:
    source = tmp_path / "稀疏扩展列.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "年龄", "备注"])
    sheet.append(["测试人员甲", 50, None])
    sheet.append(["测试人员乙", 42, "外出"])
    sheet.append(["测试人员丙", 38, None])
    sheet.append(["测试人员丁", 61, "返乡"])
    workbook.save(source)
    workbook.close()

    profile = ParserRouter().profile(source)
    regions = profile.sheets[0].region_candidates

    assert len(regions) == 1
    assert regions[0].bounds.range == "A1:C5"
    assert {
        cell_id.rsplit(":", maxsplit=1)[-1]
        for cell_id in regions[0].nonempty_cell_ids
    }
    assert regions[0].source == "connected-nonempty-cells-contained/v2"


def test_header_selection_does_not_absorb_text_heavy_first_data_row(
    tmp_path: Path,
) -> None:
    source = tmp_path / "文本型明细.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "家庭住址", "与户主关系"])
    sheet.append(["测试人员甲", "一组", "户主"])
    sheet.append(["测试人员乙", "二组", "配偶"])
    sheet.append(["测试人员丙", "三组", "子女"])
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    body_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.fill = body_fill
    workbook.save(source)
    workbook.close()

    profile = ParserRouter().profile(source)
    selected = select_header_candidates(profile.sheets[0].header_candidates)

    assert len(selected) == 1
    assert selected[0].header_rows == [1]
    assert [column.header_path for column in selected[0].columns] == [
        ["姓名"],
        ["家庭住址"],
        ["与户主关系"],
    ]


def test_header_selection_preserves_true_merged_two_row_header(tmp_path: Path) -> None:
    source = tmp_path / "两级表头.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "人员信息"
    sheet["A2"] = "姓名"
    sheet["B2"] = "年龄"
    sheet.append(["测试人员甲", 50])
    sheet.append(["测试人员乙", 42])
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    body_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for row in sheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.fill = header_fill
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.fill = body_fill
    workbook.save(source)
    workbook.close()

    profile = ParserRouter().profile(source)
    selected = select_header_candidates(profile.sheets[0].header_candidates)

    assert len(selected) == 1
    assert selected[0].header_rows == [1, 2]
    assert [column.header_path for column in selected[0].columns] == [
        ["人员信息", "姓名"],
        ["人员信息", "年龄"],
    ]


def test_profile_ids_are_stable_for_identical_content(tmp_path: Path) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "renamed.xlsx"
    make_structured_workbook(first)
    second.write_bytes(first.read_bytes())

    first_profile = ParserRouter().profile(first)
    second_profile = ParserRouter().profile(second)

    assert first_profile.workbook_id == second_profile.workbook_id
    assert first_profile.sheets[0].id == second_profile.sheets[0].id
    assert first_profile.sheets[0].cells[0].id == second_profile.sheets[0].cells[0].id


def test_content_signature_wins_over_extension(tmp_path: Path) -> None:
    disguised = tmp_path / "actually-excel.csv"
    make_structured_workbook(disguised)

    detection = detect_document(disguised)
    profile = ParserRouter().profile(disguised)

    assert detection.format == "xlsx"
    assert detection.extension_matches is False
    assert detection.warnings[0].startswith("EXTENSION_MISMATCH")
    assert profile.parser_name == "openpyxl-native"


def test_csv_profile_uses_same_evidence_contract(tmp_path: Path) -> None:
    source = tmp_path / "村情.csv"
    source.write_text("姓名,人数\n张三,2\n李四,3\n", encoding="utf-8")

    profile = ParserRouter().profile(source)

    assert profile.detection.format == "csv"
    assert profile.sheets[0].cells[0].coordinate == "A1"
    assert profile.sheets[0].region_candidates[0].bounds.range == "A1:B3"
    assert any(
        column.header_path == ["姓名"]
        for candidate in profile.sheets[0].header_candidates
        for column in candidate.columns
    )


def test_ole_signature_routes_as_xls_even_when_extension_is_wrong(tmp_path: Path) -> None:
    legacy = tmp_path / "legacy.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("人员")
    sheet.write(0, 0, "姓名")
    sheet.write(1, 0, "张三")
    workbook.save(str(legacy))
    source = tmp_path / "legacy.xlsx"
    source.write_bytes(legacy.read_bytes())

    detection = detect_document(source)
    profile = ParserRouter().profile(source)

    assert detection.format == "xls"
    assert detection.extension_matches is False
    assert profile.parser_name == "python-calamine"
    assert profile.sheets[0].cells[0].raw_value == "姓名"


def test_excel_html_export_uses_safe_table_evidence_adapter(tmp_path: Path) -> None:
    source = tmp_path / "人员.xls"
    source.write_text(
        """<html xmlns:x="urn:schemas-microsoft-com:office:excel">
        <x:Name>人员台账</x:Name>
        <script>throw new Error('must not execute')</script>
        <table><tr><td colspan="2">人员信息</td></tr>
        <tr><td>姓名</td><td>人数</td></tr>
        <tr><td>张三</td><td x:num>2</td></tr></table></html>""",
        encoding="utf-8",
    )

    profile = ParserRouter().profile(source)

    assert profile.detection.format == "excel_html"
    assert profile.detection.extension_matches is True
    assert profile.parser_name == "stdlib-excel-html"
    assert profile.sheets[0].name == "人员台账"
    assert [cell.raw_value for cell in profile.sheets[0].cells] == [
        "人员信息",
        "姓名",
        "人数",
        "张三",
        "2",
    ]
    assert profile.sheets[0].merges[0].range == "A1:B1"
    assert profile.sheets[0].region_candidates[0].bounds.range == "A1:B3"


def test_legacy_xls_uses_calamine_evidence_adapter(tmp_path: Path) -> None:
    source = tmp_path / "legacy.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("人员")
    sheet.write(0, 0, "姓名")
    sheet.write(0, 1, "人数")
    sheet.write(1, 0, "张三")
    sheet.write(1, 1, 2)
    workbook.save(str(source))

    profile = ParserRouter().profile(source)

    assert profile.parser_name == "python-calamine"
    assert profile.sheets[0].name == "人员"
    assert profile.sheets[0].cells[0].coordinate == "A1"
    assert profile.sheets[0].warnings[0].startswith("XLS_LIMITATION")


def test_unknown_binary_has_stable_error_code(tmp_path: Path) -> None:
    source = tmp_path / "unknown.xlsx"
    source.write_bytes(b"\x00\x01\x02not-a-workbook")

    with pytest.raises(DocumentDetectionError) as caught:
        detect_document(source)

    assert caught.value.code == "UNSUPPORTED_DOCUMENT_FORMAT"


def test_ooxml_repair_removes_only_unused_missing_shared_strings_reference() -> None:
    source = io.BytesIO()
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr(
            "[Content_Types].xml",
            (
                '<Types><Override PartName="/xl/sharedStrings.xml" '
                'ContentType="application/xml"/></Types>'
            ),
        )
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            (
                '<Relationships><Relationship Id="rId1" '
                'Type="http://schemas.openxmlformats.org/officeDocument/'
                '2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
                "</Relationships>"
            ),
        )
        archive.writestr(
            "xl/worksheets/sheet1.xml",
            '<worksheet><sheetData><c r="A1" t="inlineStr"/></sheetData></worksheet>',
        )

    repaired, warnings = repair_safe_ooxml_packaging(source.getvalue())

    assert warnings == ["OOXML_REPAIR: removed unused missing sharedStrings manifest references"]
    with zipfile.ZipFile(io.BytesIO(repaired)) as archive:
        assert b"sharedStrings" not in archive.read("xl/_rels/workbook.xml.rels")
