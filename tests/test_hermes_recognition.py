import asyncio
import hashlib
import uuid
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from village_insight.db.base import Base
from village_insight.db.models import (
    DatasetRecord,
    DocumentProfile,
    DocumentTemplate,
    FieldMatch,
    HermesRecognitionCache,
    HermesRecognitionRecord,
    IngestionBatch,
    IngestionItem,
    MatchType,
    SemanticField,
    SemanticFieldVersion,
    TemplateMatch,
)
from village_insight.hermes.recognition import (
    FieldDecision,
    RecognitionValidationError,
    SemanticCandidateSummary,
    SemanticFieldSummary,
    SheetRangeRequest,
    TemplateDiffResult,
    WorkbookStructureDecision,
    _chunk_recognition_request,
    _provisional_record_type,
    _recognize_sheet_structure,
    accept_recognition_proposal,
    apply_structure_decision,
    build_diff_request,
    complete_omitted_structure_regions,
    create_provisional_template,
    fulfill_range_requests,
    governance_reasons,
    normalize_field_catalog_references,
    normalize_ignored_structure_ranges,
    normalize_structure_data_ranges,
    normalize_structure_header_boundaries,
    normalize_structure_merge_references,
    normalize_structure_row_roles,
    publish_unambiguous_new_fields,
    recognition_cache_key,
    recognize_differences,
    validate_result,
    validate_structure_decision,
)
from village_insight.hermes.runtime import (
    HermesInvalidResponseError,
    HermesUnavailableError,
)
from village_insight.materialization import materialize_plan
from village_insight.parsing.router import ParserRouter
from village_insight.templates.import_plans import approve_plan


def test_provisional_record_type_uses_specific_record_grain() -> None:
    result = TemplateDiffResult(
        record_grain={
            "value": "person",
            "confidence": 0.95,
            "evidence_ids": [],
        }
    )

    assert (
        _provisional_record_type(result=result, base_definition=None)
        == "person"
    )


class FakeRuntime:
    def __init__(self, *, fail_on_call: int | None = None) -> None:
        self.calls = 0
        self.fail_on_call = fail_on_call

    async def run_json(self, **kwargs):
        import json

        from openpyxl.utils import range_boundaries

        from village_insight.hermes.recognition import WorkbookStructureDecision

        self.calls += 1
        if self.calls == self.fail_on_call:
            raise HermesUnavailableError("simulated reasoning timeout")
        output_model = kwargs["output_model"]
        request = kwargs["user_prompt"]
        assert '"raw_value"' not in request
        request_data = json.loads(request)
        if output_model is WorkbookStructureDecision:
            region = request_data["regions"][0]
            _, min_row, _, max_row = range_boundaries(region["range"])
            header_candidate_id = request_data["headers"][0]["header_candidate_id"]
            sheet_id = region["sheet_id"]
            return output_model(
                row_role_segments=[
                    {
                        "sheet_id": sheet_id,
                        "start_row": min_row,
                        "end_row": min_row,
                        "role": "header_leaf",
                    },
                    {
                        "sheet_id": sheet_id,
                        "start_row": min_row + 1,
                        "end_row": max_row,
                        "role": "data",
                    },
                ],
                layout_decisions=[
                    {
                        "region_candidate_id": region["candidate_id"],
                        "header_candidate_id": header_candidate_id,
                        "data_start_row": min_row + 1,
                        "data_end_row": max_row,
                        "classification": "table",
                        "confidence": 0.95,
                    }
                ],
                confidence=0.95,
            )
        source_column_id = request_data["headers"][0]["source_column_id"]
        header_candidate_id = request_data["headers"][0]["header_candidate_id"]
        region_candidate_id = request_data["regions"][0]["candidate_id"]
        return output_model(
            template_suggestion={
                "template_code": "test.synthetic_record",
                "template_name": "合成记录",
                "domain": "test",
                "record_type": "synthetic_record",
                "confidence": 0.8,
                "evidence_ids": [source_column_id],
            },
            layout_decisions=[
                {
                    "region_candidate_id": region_candidate_id,
                    "header_candidate_id": header_candidate_id,
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "table",
                    "confidence": 0.8,
                }
            ],
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "PROPOSE_NEW_FIELD",
                    "proposed_field_code": "person.synthetic",
                    "layer": "domain",
                    "data_type": "text",
                    "confidence": 0.8,
                    "evidence_ids": [source_column_id],
                }
            ],
        )


class SemanticOutageRuntime(FakeRuntime):
    def __init__(self) -> None:
        super().__init__()
        self.fail_semantics = True
        self.structure_calls = 0

    async def run_json(self, **kwargs):
        if kwargs["output_model"] is WorkbookStructureDecision:
            self.structure_calls += 1
            return await super().run_json(**kwargs)
        if self.fail_semantics:
            self.calls += 1
            raise HermesUnavailableError("simulated semantic outage")
        return await super().run_json(**kwargs)


class SecondSemanticChunkOutageRuntime(FakeRuntime):
    def __init__(self) -> None:
        super().__init__()
        self.fail_after_first_semantic = True
        self.structure_calls = 0
        self.successful_semantic_columns: list[str] = []

    async def run_json(self, **kwargs):
        import json

        if kwargs["output_model"] is WorkbookStructureDecision:
            self.structure_calls += 1
            return await super().run_json(**kwargs)
        request = json.loads(kwargs["user_prompt"])
        source_column_id = request["headers"][0]["source_column_id"]
        if self.fail_after_first_semantic and self.successful_semantic_columns:
            self.calls += 1
            raise HermesUnavailableError("simulated second chunk outage")
        self.calls += 1
        self.successful_semantic_columns.append(source_column_id)
        return kwargs["output_model"](
            template_suggestion={
                "template_code": "test.synthetic_record",
                "template_name": "合成记录",
                "domain": "test",
                "record_type": "synthetic_record",
                "confidence": 0.9,
            },
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "PROPOSE_NEW_FIELD",
                    "proposed_field_code": (
                        f"person.synthetic_{len(self.successful_semantic_columns)}"
                    ),
                    "layer": "domain",
                    "data_type": "text",
                    "confidence": 0.9,
                }
            ],
        )


class PublishedCodeProposalRuntime(FakeRuntime):
    async def run_json(self, **kwargs):
        import json

        self.calls += 1
        request = json.loads(kwargs["user_prompt"])
        source_column_id = request["headers"][0]["source_column_id"]
        region = request["regions"][0]
        header_candidate_id = request["headers"][0]["header_candidate_id"]
        return kwargs["output_model"](
            template_suggestion={
                "template_code": "test.synthetic_record",
                "template_name": "合成记录",
                "domain": "test",
                "record_type": "synthetic_record",
                "confidence": 0.9,
            },
            layout_decisions=[
                {
                    "region_candidate_id": region["candidate_id"],
                    "header_candidate_id": header_candidate_id,
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "table",
                    "confidence": 0.9,
                }
            ],
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "PROPOSE_NEW_FIELD",
                    "proposed_field_code": "person.already_published",
                    "layer": "domain",
                    "data_type": "text",
                    "confidence": 0.9,
                }
            ],
        )


class InvalidFirstStructureRuntime(FakeRuntime):
    async def run_json(self, **kwargs):
        if self.calls == 0:
            self.calls += 1
            raise HermesInvalidResponseError("simulated truncated structure response")
        return await super().run_json(**kwargs)


class InvalidStructureEvidenceReviewRuntime(FakeRuntime):
    async def run_json(self, **kwargs):
        if self.calls == 0:
            result = await super().run_json(**kwargs)
            import json

            request = json.loads(kwargs["user_prompt"])
            sheet_id = request["sheets"][0]["sheet_id"]
            return result.model_copy(
                update={
                    "evidence_requests": [
                        SheetRangeRequest(
                            sheet_id=sheet_id,
                            start_row=2,
                            end_row=2,
                            start_column=1,
                            end_column=1,
                            reason="confirm boundary",
                        )
                    ]
                }
            )
        self.calls += 1
        raise HermesInvalidResponseError("simulated invalid evidence review")


class StructureContractRepairRuntime:
    def __init__(self) -> None:
        self.calls = 0

    async def run_json(self, **kwargs):
        import json

        self.calls += 1
        request = json.loads(kwargs["user_prompt"])
        region = request["regions"][0]
        header = request["headers"][0]
        sheet = request["sheets"][0]
        if self.calls == 1:
            return WorkbookStructureDecision(
                row_role_segments=[
                    {
                        "sheet_id": sheet["sheet_id"],
                        "start_row": 1,
                        "end_row": 2,
                        "role": "header_leaf",
                    },
                    {
                        "sheet_id": sheet["sheet_id"],
                        "start_row": 2,
                        "end_row": 2,
                        "role": "data",
                    },
                ],
                layout_decisions=[
                    {
                        "region_candidate_id": region["candidate_id"],
                        "header_candidate_id": header["header_candidate_id"],
                        "data_start_row": 2,
                        "data_end_row": 2,
                        "classification": "table",
                        "confidence": 0.8,
                    }
                ],
                confidence=0.8,
            )
        assert request["deterministic_validation_error"] == "row role segments overlap"
        return WorkbookStructureDecision(
            row_role_segments=[
                {
                    "sheet_id": sheet["sheet_id"],
                    "start_row": 1,
                    "end_row": 1,
                    "role": "header_leaf",
                },
                {
                    "sheet_id": sheet["sheet_id"],
                    "start_row": 2,
                    "end_row": 2,
                    "role": "data",
                },
            ],
            layout_decisions=[
                {
                    "region_candidate_id": region["candidate_id"],
                    "header_candidate_id": header["header_candidate_id"],
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "table",
                    "confidence": 0.9,
                }
            ],
            confidence=0.9,
        )


class FieldContractRepairRuntime:
    def __init__(self) -> None:
        self.calls = 0

    async def run_json(self, **kwargs):
        import json

        self.calls += 1
        request = json.loads(kwargs["user_prompt"])
        output_model = kwargs["output_model"]
        if output_model is WorkbookStructureDecision:
            region = request["regions"][0]
            header = request["headers"][0]
            sheet = request["sheets"][0]
            return output_model(
                row_role_segments=[
                    {
                        "sheet_id": sheet["sheet_id"],
                        "start_row": 1,
                        "end_row": 1,
                        "role": "header_leaf",
                    },
                    {
                        "sheet_id": sheet["sheet_id"],
                        "start_row": 2,
                        "end_row": 2,
                        "role": "data",
                    },
                ],
                layout_decisions=[
                    {
                        "region_candidate_id": region["candidate_id"],
                        "header_candidate_id": header["header_candidate_id"],
                        "data_start_row": 2,
                        "data_end_row": 2,
                        "classification": "table",
                        "confidence": 0.95,
                    }
                ],
                confidence=0.95,
            )
        source_column_id = request["headers"][0]["source_column_id"]
        if "deterministic_validation_error" not in request:
            source_column_id = "foreign-region:column:99"
        else:
            assert (
                "field decisions must cover changed columns exactly"
                in request["deterministic_validation_error"]
            )
        return output_model(
            template_suggestion={
                "template_code": "test.synthetic_record",
                "template_name": "合成记录",
                "domain": "test",
                "record_type": "synthetic_record",
                "confidence": 0.95,
            },
            layout_decisions=[],
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "PROPOSE_NEW_FIELD",
                    "proposed_field_code": "person.synthetic",
                    "layer": "domain",
                    "data_type": "text",
                    "confidence": 0.95,
                }
            ],
        )


class MissingTemplateRepairInvalidRuntime:
    def __init__(self) -> None:
        self.calls = 0
        self.repair_payload: dict[str, object] | None = None

    async def run_json(self, **kwargs):
        import json

        self.calls += 1
        if str(kwargs.get("task_id", "")).startswith(
            "template-diff-contract-repair-"
        ):
            self.repair_payload = json.loads(kwargs["user_prompt"])
            raise HermesInvalidResponseError("simulated invalid repair response")
        request = json.loads(kwargs["user_prompt"])
        output_model = kwargs["output_model"]
        source_column_id = request["headers"][0]["source_column_id"]
        header_candidate_id = request["headers"][0]["header_candidate_id"]
        region_candidate_id = request["regions"][0]["candidate_id"]
        return output_model(
            template_suggestion=None,
            layout_decisions=[
                {
                    "region_candidate_id": region_candidate_id,
                    "header_candidate_id": header_candidate_id,
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "table",
                    "confidence": 0.95,
                }
            ],
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "PROPOSE_NEW_FIELD",
                    "proposed_field_code": "person.synthetic",
                    "layer": "domain",
                    "data_type": "text",
                    "confidence": 0.95,
                }
            ],
        )


def synthetic_profile(tmp_path: Path):
    source = tmp_path / "synthetic.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["新增字段"])
    sheet.append(["脱敏样例"])
    workbook.save(source)
    workbook.close()
    return source, ParserRouter().profile(source)


def add_item(
    database: Session,
    source: Path,
    *,
    administrative_unit_id: uuid.UUID | None = None,
) -> IngestionItem:
    batch = IngestionBatch(
        name=uuid.uuid4().hex,
        total_files=1,
        administrative_unit_id=administrative_unit_id or uuid.uuid4(),
    )
    database.add(batch)
    database.flush()
    payload = source.read_bytes()
    item = IngestionItem(
        tenant_id=batch.tenant_id,
        administrative_unit_id=batch.administrative_unit_id,
        created_by_user_id=batch.created_by_user_id,
        batch_id=batch.id,
        original_name=source.name,
        source_path=str(source),
        source_sha256=hashlib.sha256(payload).hexdigest(),
        size_bytes=len(payload),
    )
    database.add(item)
    database.flush()
    return item


def add_match(database: Session, item: IngestionItem, profile) -> TemplateMatch:
    database.add(
        DocumentProfile(
            item_id=item.id,
            contract_version=profile.contract_version,
            source_sha256=profile.source_sha256,
            parser_name=profile.parser_name,
            parser_version=profile.parser_version,
            profile=profile.model_dump(mode="json"),
        )
    )
    header = profile.sheets[0].header_candidates[0].columns[0]
    match = TemplateMatch(
        item_id=item.id,
        source_sha256=profile.source_sha256,
        profile_contract_version=profile.contract_version,
        layout_fingerprint=hashlib.sha256(b"layout").hexdigest(),
        match_type=MatchType.NONE,
        score_basis_points=0,
        differences={
            "new_headers": [" / ".join(header.header_path)],
            "missing_headers": [],
        },
        requires_hermes=True,
        matcher_version="layout-matcher/v1",
    )
    database.add(match)
    database.flush()
    return match


def test_diff_request_contains_bounded_sheet_preview_and_range_reader(
    tmp_path: Path,
) -> None:
    source = tmp_path / "preview.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "人数"])
    for index in range(1, 51):
        sheet.append([f"人员{index}", index])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(
            profile,
            add_match(database, item, profile),
        )

        assert request.sheets[0].rows == 51
        assert request.range_evidence[0].purpose == "initial_preview"
        assert request.range_evidence[0].range == "A1:B10"
        assert {evidence.range for evidence in request.range_evidence[1:]} == {
            "A26:B26",
            "A51:B51",
        }
        requested = fulfill_range_requests(
            profile,
            [
                SheetRangeRequest(
                    sheet_id=request.sheets[0].sheet_id,
                    start_row=31,
                    end_row=35,
                    start_column=1,
                    end_column=2,
                    reason="确认数据行连续性",
                )
            ],
        )
        assert requested[0].purpose == "requested"
        assert requested[0].range == "A31:B35"
        assert all(
            cell[1] != f"人员{row.row - 1}"
            for row in requested[0].rows
            for cell in row.cells
            if cell[0] == 1
        )
        clamped = fulfill_range_requests(
            profile,
            [
                SheetRangeRequest(
                    sheet_id=request.sheets[0].sheet_id,
                    start_row=48,
                    end_row=55,
                    start_column=1,
                    end_column=2,
                    reason="检查尾部边界",
                )
            ],
        )
        assert clamped[0].range == "A48:B51"

        with pytest.raises(
            RecognitionValidationError,
            match="exceeds row limit",
        ):
            fulfill_range_requests(
                profile,
                [
                    SheetRangeRequest(
                        sheet_id=request.sheets[0].sheet_id,
                        start_row=1,
                        end_row=31,
                        start_column=1,
                        end_column=2,
                        reason="越界请求",
                    )
                ],
            )
    Base.metadata.drop_all(engine)


def test_merge_evidence_limit_is_applied_per_physical_sheet(
    tmp_path: Path,
) -> None:
    source = tmp_path / "多Sheet合并表头.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "大量合并"
    for row in range(1, 21):
        first.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=2,
        )
        first.cell(row=row, column=1, value=f"字段{row}")
    second = workbook.create_sheet("第二张表")
    second.merge_cells("A1:B1")
    second["A1"] = "人员信息"
    second.append(["姓名", "年龄"])
    second.append(["测试人员", 40])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))

        first_merge_ids = [
            merge_id
            for merge_id in request.merge_ids
            if merge_id.startswith(f"{profile.sheets[0].id}:")
        ]
        second_merge_ids = [
            merge_id
            for merge_id in request.merge_ids
            if merge_id.startswith(f"{profile.sheets[1].id}:")
        ]
        assert len(first_merge_ids) == 16
        assert second_merge_ids == [profile.sheets[1].merges[0].id]
    Base.metadata.drop_all(engine)


def test_recognition_runs_sheet_structure_interpretation_before_field_mapping(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = FakeRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=request,
                profile=profile,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-pro",
                hermes_version="structure-test",
            )
        )

        assert runtime.calls == 3
        structure = proposal.proposal["structure_decision"]
        assert structure["layout_decisions"][0]["materialize"] is True
        assert structure["row_role_segments"] == [
            {
                "sheet_id": profile.sheets[0].id,
                "start_row": 1,
                "end_row": 1,
                "role": "header_leaf",
                "evidence_ids": [],
            },
            {
                "sheet_id": profile.sheets[0].id,
                "start_row": 2,
                "end_row": 2,
                "role": "data",
                "evidence_ids": [],
            },
        ]
        records = list(database.query(HermesRecognitionRecord))
        assert [record.model for record in records] == [
            "fake-fast",
            "fake-fast",
            "fake-pro",
        ]
    Base.metadata.drop_all(engine)


def test_semantic_retry_resumes_from_validated_structure_checkpoint(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = SemanticOutageRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))

        with pytest.raises(HermesUnavailableError, match="semantic outage"):
            asyncio.run(
                recognize_differences(
                    database,
                    item_id=item.id,
                    request=request,
                    profile=profile,
                    runtime=runtime,
                    provider="fake",
                    model="fake-fast",
                    reasoning_model="fake-pro",
                    hermes_version="structure-checkpoint-test",
                )
            )

        checkpoints = list(database.query(HermesRecognitionCache))
        assert len(checkpoints) == 1
        assert checkpoints[0].schema_version == "workbook-structure-checkpoint/v2"
        assert runtime.structure_calls == 1

        runtime.fail_semantics = False
        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=request,
                profile=profile,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-pro",
                hermes_version="structure-checkpoint-test",
            )
        )

        assert proposal.proposal["structure_decision"] is not None
        assert runtime.structure_calls == 1
    Base.metadata.drop_all(engine)


def test_semantic_retry_resumes_after_last_validated_chunk(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "village_insight.hermes.recognition.MAX_FIELDS_PER_HERMES_CALL",
        1,
    )
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = SecondSemanticChunkOutageRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        original_header = request.headers[0]
        second_header = original_header.model_copy(
            update={
                "source_column_id": "synthetic-second-column",
                "header_path": ["第二字段"],
            }
        )
        expanded = request.model_copy(
            update={
                "headers": [original_header, second_header],
                "new_headers": ["新增字段", "第二字段"],
                "unresolved_source_column_ids": [
                    original_header.source_column_id,
                    second_header.source_column_id,
                ],
            }
        )

        with pytest.raises(HermesUnavailableError, match="second chunk outage"):
            asyncio.run(
                recognize_differences(
                    database,
                    item_id=item.id,
                    request=expanded,
                    profile=profile,
                    runtime=runtime,
                    provider="fake",
                    model="fake-fast",
                    reasoning_model="fake-pro",
                    hermes_version="semantic-checkpoint-test",
                )
            )

        first_column = original_header.source_column_id
        assert runtime.successful_semantic_columns == [first_column]
        assert runtime.structure_calls == 1
        assert {
            row.schema_version for row in database.query(HermesRecognitionCache)
        } == {
            "workbook-structure-checkpoint/v2",
            "semantic-chunk-checkpoint/v2",
        }

        runtime.fail_after_first_semantic = False
        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=expanded,
                profile=profile,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-pro",
                hermes_version="semantic-checkpoint-test",
            )
        )

        assert runtime.structure_calls == 1
        assert runtime.successful_semantic_columns == [
            first_column,
            second_header.source_column_id,
        ]
        assert len(proposal.proposal["field_decisions"]) == 2
    Base.metadata.drop_all(engine)


def test_invalid_structure_checkpoint_is_a_cache_miss_and_is_rebuilt(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = SemanticOutageRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        with pytest.raises(HermesUnavailableError, match="semantic outage"):
            asyncio.run(
                recognize_differences(
                    database,
                    item_id=item.id,
                    request=request,
                    profile=profile,
                    runtime=runtime,
                    provider="fake",
                    model="fake-fast",
                    reasoning_model="fake-pro",
                    hermes_version="invalid-structure-checkpoint-test",
                )
            )
        checkpoint = database.query(HermesRecognitionCache).one()
        checkpoint.response_payload = {"damaged": True}
        database.flush()

        runtime.fail_semantics = False
        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=request,
                profile=profile,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-pro",
                hermes_version="invalid-structure-checkpoint-test",
            )
        )

        assert runtime.structure_calls == 2
        assert proposal.proposal["structure_decision"] is not None
        assert "structure_decision" in checkpoint.response_payload
    Base.metadata.drop_all(engine)


def test_invalid_semantic_checkpoint_is_a_cache_miss_and_is_rebuilt(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "village_insight.hermes.recognition.MAX_FIELDS_PER_HERMES_CALL",
        1,
    )
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = SecondSemanticChunkOutageRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        first_header = request.headers[0]
        second_header = first_header.model_copy(
            update={
                "source_column_id": "synthetic-second-column",
                "header_path": ["第二字段"],
            }
        )
        expanded = request.model_copy(
            update={
                "headers": [first_header, second_header],
                "new_headers": ["新增字段", "第二字段"],
                "unresolved_source_column_ids": [
                    first_header.source_column_id,
                    second_header.source_column_id,
                ],
            }
        )
        with pytest.raises(HermesUnavailableError, match="second chunk outage"):
            asyncio.run(
                recognize_differences(
                    database,
                    item_id=item.id,
                    request=expanded,
                    profile=profile,
                    runtime=runtime,
                    provider="fake",
                    model="fake-fast",
                    reasoning_model="fake-pro",
                    hermes_version="invalid-semantic-checkpoint-test",
                )
            )
        semantic_checkpoint = database.scalar(
            select(HermesRecognitionCache).where(
                HermesRecognitionCache.schema_version
                == "semantic-chunk-checkpoint/v2"
            )
        )
        assert semantic_checkpoint is not None
        semantic_checkpoint.response_payload = {"damaged": True}
        database.flush()

        runtime.fail_after_first_semantic = False
        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=expanded,
                profile=profile,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-pro",
                hermes_version="invalid-semantic-checkpoint-test",
            )
        )

        assert runtime.structure_calls == 1
        assert runtime.successful_semantic_columns == [
            first_header.source_column_id,
            first_header.source_column_id,
            second_header.source_column_id,
        ]
        assert len(proposal.proposal["field_decisions"]) == 2
        assert "result" in semantic_checkpoint.response_payload
    Base.metadata.drop_all(engine)


def test_unresolved_field_request_keeps_known_sibling_values_as_context(
    tmp_path: Path,
) -> None:
    source = tmp_path / "field-context.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["组别", "新增完成数"])
    sheet.append(["第一组", 12])
    sheet.append(["第二组", 9])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)
    candidate = profile.sheets[0].header_candidates[0]
    known_column, unresolved_column = candidate.columns
    matches = [
        FieldMatch(
            item_id=uuid.uuid4(),
            sheet_id=profile.sheets[0].id,
            region_id=candidate.region_id,
            header_id=candidate.id,
            source_column_id=known_column.source_column_id,
            header_path=known_column.header_path,
            observed_data_type="text",
            match_type=MatchType.EXACT,
            score_basis_points=10_000,
            context={},
            differences={},
            requires_hermes=False,
            matcher_version="test",
        ),
        FieldMatch(
            item_id=uuid.uuid4(),
            sheet_id=profile.sheets[0].id,
            region_id=candidate.region_id,
            header_id=candidate.id,
            source_column_id=unresolved_column.source_column_id,
            header_path=unresolved_column.header_path,
            observed_data_type="integer",
            match_type=MatchType.NONE,
            score_basis_points=0,
            context={},
            differences={},
            requires_hermes=True,
            matcher_version="test",
        ),
    ]
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type=MatchType.EXACT,
            score_basis_points=9_000,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
        field_matches=matches,
    )

    sampled_ids = {
        cell.source_column_id
        for sample in request.source_samples
        for cell in sample.cells
    }
    supplied_ids = {header.source_column_id for header in request.headers}
    assert request.unresolved_source_column_ids == [unresolved_column.source_column_id]
    assert unresolved_column.source_column_id in sampled_ids
    assert known_column.source_column_id in sampled_ids
    assert sampled_ids <= supplied_ids


def test_governance_free_new_field_is_published_for_current_import(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type=MatchType.NONE,
            score_basis_points=0,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
    )
    decision = FieldDecision(
        source_column_id=request.headers[0].source_column_id,
        action="PROPOSE_NEW_FIELD",
        proposed_field_code="test.confirmed_count",
        layer="domain",
        data_type="integer",
        confidence=0.92,
        requires_review=True,
    )
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        publish_unambiguous_new_fields(
            database,
            request=request,
            result=TemplateDiffResult(
                field_decisions=[decision],
                requires_governance=False,
            ),
        )
        database.commit()
        field = database.scalar(
            select(SemanticField).where(SemanticField.code == "test.confirmed_count")
        )
        assert field is not None
        assert field.published_version == 1
        version = database.scalar(
            select(SemanticFieldVersion).where(
                SemanticFieldVersion.field_id == field.id,
                SemanticFieldVersion.version == 1,
            )
        )
        assert version is not None
        assert version.status == "published"
        assert version.source == "auto_governance"
        assert version.source_metadata["source_contract"] == (
            "four-layer-template-source/v1"
        )
        assert version.variants[0].source == "auto_governance"
        assert version.variants[0].header_path == request.headers[0].header_path
    Base.metadata.drop_all(engine)


def test_structure_row_roles_can_be_sparse_when_layout_bounds_are_complete(
    tmp_path: Path,
) -> None:
    _, profile = synthetic_profile(tmp_path)
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type="none",
            score_basis_points=0,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
    )
    region = request.regions[0]
    header = request.headers[0]
    decision = WorkbookStructureDecision(
        row_role_segments=[],
        layout_decisions=[
            {
                "region_candidate_id": region.candidate_id,
                "header_candidate_id": header.header_candidate_id,
                "data_start_row": 2,
                "data_end_row": 2,
                "classification": "table",
                "confidence": 0.95,
            }
        ],
        confidence=0.95,
    )

    validate_structure_decision(profile, request, decision)


def test_materialized_structure_cannot_include_selected_header_rows(
    tmp_path: Path,
) -> None:
    _, profile = synthetic_profile(tmp_path)
    profile.sheets[0].header_candidates[0].header_rows = [1, 2]
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type="none",
            score_basis_points=0,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
    )
    region = request.regions[0]
    header = request.headers[0]
    decision = WorkbookStructureDecision(
        row_role_segments=[],
        layout_decisions=[
            {
                "region_candidate_id": region.candidate_id,
                "header_candidate_id": header.header_candidate_id,
                "data_start_row": 2,
                "data_end_row": 2,
                "classification": "table",
                "confidence": 0.95,
            }
        ],
        confidence=0.95,
    )

    with pytest.raises(
        RecognitionValidationError,
        match="includes the selected header row",
    ):
        validate_structure_decision(profile, request, decision)


def test_ignored_structure_coordinates_are_bounded_without_relaxing_evidence(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        region = request.regions[0]
        header = request.headers[0]
        decision = WorkbookStructureDecision(
            row_role_segments=[
                {
                    "sheet_id": region.sheet_id,
                    "start_row": 1,
                    "end_row": 2,
                    "role": "note",
                }
            ],
            layout_decisions=[
                {
                    "region_candidate_id": region.candidate_id,
                    "header_candidate_id": header.header_candidate_id,
                    "data_start_row": 99,
                    "data_end_row": 100,
                    "classification": "table",
                    "materialize": False,
                    "confidence": 0.8,
                }
            ],
            confidence=0.8,
        )

        normalized, changed = normalize_ignored_structure_ranges(
            request,
            decision,
        )

        assert changed is True
        assert normalized.layout_decisions[0].data_start_row == 2
        assert normalized.layout_decisions[0].data_end_row == 2
        assert normalized.layout_decisions[0].classification == "noise"
        assert normalized.layout_decisions[0].materialize is False
    Base.metadata.drop_all(engine)


def test_retained_structure_coordinates_are_clamped_to_region_bounds(
    tmp_path: Path,
) -> None:
    _, profile = synthetic_profile(tmp_path)
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type=MatchType.NONE,
            score_basis_points=0,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
    )
    region = request.regions[0]
    header = request.headers[0]
    decision = WorkbookStructureDecision(
        layout_decisions=[
            {
                "region_candidate_id": region.candidate_id,
                "header_candidate_id": header.header_candidate_id,
                "data_start_row": 2,
                "data_end_row": 99,
                "excluded_rows": [3, 99],
                "classification": "table",
                "materialize": True,
                "confidence": 0.9,
            }
        ],
        confidence=0.9,
    )

    normalized, changed = normalize_structure_data_ranges(request, decision)

    assert changed is True
    assert normalized.layout_decisions[0].data_end_row == 2
    assert normalized.layout_decisions[0].excluded_rows == []


def test_reversed_structure_range_becomes_reviewable_noise(
    tmp_path: Path,
) -> None:
    _, profile = synthetic_profile(tmp_path)
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type=MatchType.NONE,
            score_basis_points=0,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
    )
    region = request.regions[0]
    header = request.headers[0]
    decision = WorkbookStructureDecision(
        layout_decisions=[
            {
                "region_candidate_id": region.candidate_id,
                "header_candidate_id": header.header_candidate_id,
                "data_start_row": 99,
                "data_end_row": 1,
                "classification": "table",
                "materialize": True,
                "confidence": 0.9,
            }
        ],
        confidence=0.9,
    )

    normalized, changed = normalize_structure_data_ranges(request, decision)
    layout = normalized.layout_decisions[0]

    assert changed is True
    assert layout.data_start_row == 1
    assert layout.data_end_row == 2
    assert layout.classification == "noise"
    assert layout.materialize is False
    assert layout.confidence == 0.0


def test_overlapping_optional_row_roles_are_dropped_without_model_repair(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = StructureContractRepairRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))

        decision, _, models, uncertain = asyncio.run(
            _recognize_sheet_structure(
                profile=profile,
                request=request,
                runtime=runtime,
                item_id=item.id,
                model="fake-fast",
                reasoning_model="fake-reasoning",
            )
        )

        assert runtime.calls == 1
        assert models == ["fake-fast"]
        assert uncertain is False
        assert decision.row_role_segments == []
    Base.metadata.drop_all(engine)


def test_invalid_fast_structure_response_uses_reasoning_fallback(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = InvalidFirstStructureRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))

        decision, _, models, uncertain = asyncio.run(
            _recognize_sheet_structure(
                profile=profile,
                request=request,
                runtime=runtime,
                item_id=item.id,
                model="fake-fast",
                reasoning_model="fake-reasoning",
            )
        )

        assert runtime.calls == 2
        assert models == ["fake-reasoning"]
        assert uncertain is False
        assert decision.layout_decisions[0].materialize is True
    Base.metadata.drop_all(engine)


def test_invalid_structure_evidence_review_keeps_conservative_decision(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = InvalidStructureEvidenceReviewRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))

        decision, enriched, models, uncertain = asyncio.run(
            _recognize_sheet_structure(
                profile=profile,
                request=request,
                runtime=runtime,
                item_id=item.id,
                model="fake-fast",
                reasoning_model="fake-reasoning",
            )
        )

        assert runtime.calls == 2
        assert models == ["fake-fast"]
        assert uncertain is True
        assert decision.layout_decisions[0].materialize is True
        assert any(evidence.purpose == "requested" for evidence in enriched.range_evidence)
    Base.metadata.drop_all(engine)


def test_omitted_structure_region_becomes_explicit_reviewable_noise(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))

        completed, changed = complete_omitted_structure_regions(
            request,
            WorkbookStructureDecision(confidence=0.8),
        )

        assert changed is True
        assert len(completed.layout_decisions) == len(request.regions)
        assert all(
            decision.classification == "noise"
            and decision.materialize is False
            and decision.confidence == 0.0
            for decision in completed.layout_decisions
        )
        validate_structure_decision(profile, request, completed)
    Base.metadata.drop_all(engine)


def test_structure_data_start_is_clamped_after_selected_header(
    tmp_path: Path,
) -> None:
    _, profile = synthetic_profile(tmp_path)
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type="none",
            score_basis_points=0,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
    )
    region = request.regions[0]
    header = request.headers[0]
    decision = WorkbookStructureDecision(
        layout_decisions=[
            {
                "region_candidate_id": region.candidate_id,
                "header_candidate_id": header.header_candidate_id,
                "data_start_row": 1,
                "data_end_row": 2,
                "classification": "table",
                "confidence": 0.9,
            }
        ],
        confidence=0.9,
    )

    normalized, changed = normalize_structure_header_boundaries(profile, decision)

    assert changed is True
    assert normalized.layout_decisions[0].data_start_row == 2
    validate_structure_decision(profile, request, normalized)


def test_out_of_bounds_row_roles_are_clamped_and_unknown_evidence_removed(
    tmp_path: Path,
) -> None:
    _, profile = synthetic_profile(tmp_path)
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type="none",
            score_basis_points=0,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
    )
    decision = WorkbookStructureDecision(
        row_role_segments=[
            {
                "sheet_id": request.sheets[0].sheet_id,
                "start_row": 1,
                "end_row": 99,
                "role": "note",
                "evidence_ids": ["invented"],
            }
        ],
        confidence=0.8,
    )

    normalized, changed = normalize_structure_row_roles(profile, request, decision)

    assert changed is True
    assert normalized.row_role_segments[0].end_row == 2
    assert normalized.row_role_segments[0].evidence_ids == []


def test_field_chunk_contract_conflict_is_repaired_before_merge(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = FieldContractRepairRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))

        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=request,
                profile=profile,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-reasoning",
                hermes_version="field-contract-repair-test",
            )
        )

        assert runtime.calls == 3
        assert proposal.proposal["field_decisions"][0]["source_column_id"] == (
            request.headers[0].source_column_id
        )
        assert "HERMES_CONTRACT_REPAIRED" not in proposal.proposal[
            "governance_reason_codes"
        ]
    Base.metadata.drop_all(engine)


def test_missing_template_suggestion_falls_back_after_invalid_repair(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = MissingTemplateRepairInvalidRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))

        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-reasoning",
                hermes_version="missing-template-fallback-test",
            )
        )

        assert runtime.calls == 2
        assert runtime.repair_payload is not None
        assert runtime.repair_payload["validation_error"] == {
            "code": "TEMPLATE_SUGGESTION_REQUIRED",
            "message": "template suggestion is required when no template matches",
        }
        repair_target = runtime.repair_payload["repair_target"]
        assert isinstance(repair_target, dict)
        assert repair_target["required_path"] == "template_suggestion"
        assert proposal.proposal["template_suggestion"] is None
        assert proposal.proposal["requires_governance"] is True
        assert set(proposal.proposal["governance_reason_codes"]) == {
            "HERMES_TEMPLATE_SUGGESTION_MISSING",
            "HERMES_CONTRACT_REPAIR_INVALID_RESPONSE",
        }

        template, version = create_provisional_template(database, proposal=proposal)
        assert template.code.startswith("hermes.provisional.")
        assert version.definition["domain"] == "unclassified"
        assert version.definition["record_type"] == "unclassified_record"
    Base.metadata.drop_all(engine)


def test_structure_merge_range_is_resolved_only_to_supplied_evidence(
    tmp_path: Path,
) -> None:
    source = tmp_path / "合并表头.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "人员信息"
    sheet.append(["姓名", "年龄"])
    sheet.append(["测试人员", 40])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        region = request.regions[0]
        header = request.headers[0]
        decision = WorkbookStructureDecision(
            row_role_segments=[
                {
                    "sheet_id": region.sheet_id,
                    "start_row": 1,
                    "end_row": 2,
                    "role": "header_leaf",
                },
                {
                    "sheet_id": region.sheet_id,
                    "start_row": 3,
                    "end_row": 3,
                    "role": "data",
                },
            ],
            layout_decisions=[
                {
                    "region_candidate_id": region.candidate_id,
                    "header_candidate_id": header.header_candidate_id,
                    "data_start_row": 3,
                    "data_end_row": 3,
                    "classification": "table",
                    "confidence": 0.9,
                    "merge_decisions": [
                        {
                            "merge_id": "A1:B1",
                            "action": "STRUCTURAL_GROUP",
                        },
                        {
                            "merge_id": "A9:B9",
                            "action": "IGNORE",
                        },
                        {
                            "merge_id": profile.sheets[0].merges[0].id,
                            "action": "PROPAGATE",
                            "target_source_column_ids": ["unknown-column"],
                        },
                    ],
                }
            ],
            confidence=0.9,
        )

        normalized, changed = normalize_structure_merge_references(
            profile,
            request,
            decision,
        )

        assert changed is True
        assert normalized.layout_decisions[0].merge_decisions[0].merge_id == (
            profile.sheets[0].merges[0].id
        )
        assert len(normalized.layout_decisions[0].merge_decisions) == 1
    Base.metadata.drop_all(engine)


def test_unknown_semantic_reuse_becomes_new_field_candidate(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        source_column_id = request.headers[0].source_column_id
        result = TemplateDiffResult(
            template_suggestion={
                "template_code": "test.synthetic",
                "template_name": "合成模板",
                "domain": "test",
                "record_type": "synthetic",
                "confidence": 0.9,
            },
            layout_decisions=[
                {
                    "region_candidate_id": request.regions[0].candidate_id,
                    "header_candidate_id": request.headers[0].header_candidate_id,
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "table",
                    "confidence": 0.9,
                }
            ],
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "REUSE_FIELD",
                    "semantic_field_code": "person.not_published",
                    "confidence": 0.9,
                    "evidence_ids": [source_column_id],
                }
            ],
        )

        normalized, changed = normalize_field_catalog_references(request, result)

        assert changed is True
        assert normalized.field_decisions[0].action == "PROPOSE_NEW_FIELD"
        assert normalized.field_decisions[0].semantic_field_code is None
        assert normalized.field_decisions[0].proposed_field_code == "person.not_published"
        assert normalized.field_decisions[0].confidence == 0.75
        assert governance_reasons(normalized)[0] == ["HERMES_LOW_CONFIDENCE"]
    Base.metadata.drop_all(engine)


def test_model_cannot_invent_free_form_field_role(tmp_path: Path) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        source_column_id = request.headers[0].source_column_id
        result = TemplateDiffResult(
            template_suggestion={
                "template_code": "test.synthetic",
                "template_name": "合成模板",
                "domain": "test",
                "record_type": "synthetic",
                "confidence": 0.9,
            },
            layout_decisions=[],
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "PROPOSE_NEW_FIELD",
                    "proposed_field_code": "test.identifier",
                    "layer": "domain",
                    "data_type": "text",
                    "role": "recordidentifier",
                    "confidence": 0.9,
                    "evidence_ids": [source_column_id],
                }
            ],
        )

        normalized, changed = normalize_field_catalog_references(request, result)

        assert changed is True
        assert normalized.field_decisions[0].role is None
        assert normalized.field_decisions[0].requires_review is True
    Base.metadata.drop_all(engine)


def test_date_scoped_proposals_collapse_to_one_reusable_field(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    request = build_diff_request(
        profile,
        TemplateMatch(
            source_sha256=profile.source_sha256,
            profile_contract_version=profile.contract_version,
            layout_fingerprint="test",
            match_type=MatchType.NONE,
            score_basis_points=0,
            differences={"new_headers": [], "missing_headers": []},
            requires_hermes=True,
            matcher_version="test",
        ),
    )
    first = request.headers[0]
    second = first.model_copy(
        update={
            "source_column_id": f"{first.source_column_id}:second",
            "context": {"role": "date_2025_12_02"},
        }
    )
    request = request.model_copy(
        update={
            "headers": [
                first.model_copy(update={"context": {"role": "date_2025_12_01"}}),
                second,
            ]
        }
    )
    result = TemplateDiffResult(
        field_decisions=[
            FieldDecision(
                source_column_id=first.source_column_id,
                action="PROPOSE_NEW_FIELD",
                proposed_field_code="social_security.daily_completion_2025_12_01",
                layer="domain",
                data_type="integer",
                role="date_2025_12_01",
                confidence=0.9,
            ),
            FieldDecision(
                source_column_id=second.source_column_id,
                action="PROPOSE_NEW_FIELD",
                proposed_field_code="social_security.daily_completion_2025_12_02",
                layer="domain",
                data_type="integer",
                role="date_2025_12_02",
                confidence=0.9,
            ),
        ]
    )

    normalized, changed = normalize_field_catalog_references(request, result)

    assert changed is True
    assert normalized.field_decisions[0].proposed_field_code == (
        "social_security.daily_completion"
    )
    assert normalized.field_decisions[1].action == "ROLE_VARIANT"
    assert normalized.field_decisions[1].semantic_field_code == (
        "social_security.daily_completion"
    )
    assert normalized.field_decisions[1].role == "date_2025_12_02"


def test_legacy_candidate_role_does_not_whitelist_free_form_role(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        source_column_id = request.headers[0].source_column_id
        request.headers[0].semantic_candidates = [
            SemanticCandidateSummary(
                code="person.identity_number",
                version=1,
                name="身份证号",
                data_type="text",
                score_basis_points=8000,
                reasons=["legacy role"],
                compatible_roles=["recordidentifier"],
            )
        ]
        result = TemplateDiffResult(
            template_suggestion={
                "template_code": "test.synthetic",
                "template_name": "合成模板",
                "domain": "test",
                "record_type": "synthetic",
                "confidence": 0.9,
            },
            layout_decisions=[],
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "REUSE_FIELD",
                    "semantic_field_code": "person.identity_number",
                    "data_type": "text",
                    "role": "recordidentifier",
                    "confidence": 0.9,
                    "evidence_ids": [source_column_id],
                }
            ],
        )

        normalized, changed = normalize_field_catalog_references(request, result)

        assert changed is True
        assert normalized.field_decisions[0].role is None
        assert normalized.field_decisions[0].requires_review is True
    Base.metadata.drop_all(engine)


def test_structure_decision_removes_ignored_region_columns_from_semantic_work(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        retained_header = request.headers[0]
        ignored_header = retained_header.model_copy(
            update={
                "header_candidate_id": "ignored-header",
                "region_candidate_id": "ignored-region",
                "source_column_id": "ignored-column",
                "header_path": ["误识别字段"],
            }
        )
        ignored_region = request.regions[0].model_copy(update={"candidate_id": "ignored-region"})
        expanded = request.model_copy(
            update={
                "new_headers": [
                    *request.new_headers,
                    "误识别字段",
                ],
                "unresolved_source_column_ids": [
                    *request.unresolved_source_column_ids,
                    "ignored-column",
                ],
                "headers": [*request.headers, ignored_header],
                "regions": [*request.regions, ignored_region],
                "source_samples": [
                    request.source_samples[0].model_copy(
                        update={
                            "cells": [
                                *request.source_samples[0].cells,
                                request.source_samples[0]
                                .cells[0]
                                .model_copy(update={"source_column_id": "ignored-column"}),
                            ]
                        }
                    )
                ],
            }
        )
        structure = WorkbookStructureDecision(
            row_role_segments=[],
            layout_decisions=[
                {
                    "region_candidate_id": request.regions[0].candidate_id,
                    "header_candidate_id": retained_header.header_candidate_id,
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "table",
                    "materialize": True,
                    "confidence": 0.9,
                },
                {
                    "region_candidate_id": "ignored-region",
                    "header_candidate_id": "ignored-header",
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "noise",
                    "materialize": False,
                    "confidence": 0.9,
                },
            ],
            confidence=0.9,
        )

        narrowed = apply_structure_decision(expanded, structure)

        assert narrowed.unresolved_source_column_ids == [retained_header.source_column_id]
        assert narrowed.new_headers == ["新增字段"]
        assert {header.source_column_id for header in narrowed.headers} == {
            retained_header.source_column_id
        }
        assert {
            cell.source_column_id for sample in narrowed.source_samples for cell in sample.cells
        } == {retained_header.source_column_id}
    Base.metadata.drop_all(engine)


def test_all_noise_structure_is_a_successful_zero_record_plan(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        structure = WorkbookStructureDecision(
            row_role_segments=[],
            layout_decisions=[
                {
                    "region_candidate_id": request.regions[0].candidate_id,
                    "header_candidate_id": request.headers[0].header_candidate_id,
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "noise",
                    "materialize": False,
                    "confidence": 0.55,
                }
            ],
            confidence=0.55,
        )

        narrowed = apply_structure_decision(request, structure)
        result = TemplateDiffResult(
            layout_decisions=list(structure.layout_decisions),
            structure_decision=structure,
        )

        assert narrowed.new_headers == []
        assert narrowed.unresolved_source_column_ids == []
        assert narrowed.headers == request.headers
        assert narrowed.regions == request.regions
        assert narrowed.source_samples == []
        assert _chunk_recognition_request(narrowed) == []
        validate_result(narrowed, result)
        assert governance_reasons(result) == ([], 0.55)
    Base.metadata.drop_all(engine)


def test_published_field_not_supplied_for_column_becomes_ambiguous(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        source_column_id = request.headers[0].source_column_id
        request = request.model_copy(
            update={
                "semantic_catalog": [
                    SemanticFieldSummary(
                        code="person.name",
                        version=1,
                        name="姓名",
                        layer="base",
                        data_type="text",
                    )
                ],
                "headers": [
                    request.headers[0].model_copy(
                        update={
                            "semantic_candidates": [
                                SemanticCandidateSummary(
                                    code="household.address",
                                    version=1,
                                    name="家庭住址",
                                    data_type="text",
                                    score_basis_points=4500,
                                )
                            ]
                        }
                    )
                ],
            }
        )
        result = TemplateDiffResult(
            template_suggestion={
                "template_code": "test.synthetic",
                "template_name": "合成模板",
                "domain": "test",
                "record_type": "synthetic",
                "confidence": 0.9,
            },
            layout_decisions=[
                {
                    "region_candidate_id": request.regions[0].candidate_id,
                    "header_candidate_id": request.headers[0].header_candidate_id,
                    "data_start_row": 2,
                    "data_end_row": 2,
                    "classification": "table",
                    "confidence": 0.9,
                }
            ],
            field_decisions=[
                {
                    "source_column_id": source_column_id,
                    "action": "REUSE_FIELD",
                    "semantic_field_code": "person.name",
                    "confidence": 0.9,
                    "evidence_ids": [source_column_id],
                }
            ],
        )

        normalized, changed = normalize_field_catalog_references(request, result)

        assert changed is True
        assert normalized.field_decisions[0].action == "AMBIGUOUS"
        assert normalized.field_decisions[0].semantic_field_code is None
        assert normalized.field_decisions[0].requires_review is True
        validate_result(request, normalized)
    Base.metadata.drop_all(engine)


def test_substring_only_field_reuse_becomes_ambiguous(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        header = request.headers[0].model_copy(
            update={
                "header_path": ["土地信息", "确权面积"],
                "observed_data_type": "decimal",
                "semantic_candidates": [
                    SemanticCandidateSummary(
                        code="agriculture.area",
                        version=1,
                        name="面积",
                        aliases=["面积"],
                        data_type="decimal",
                        score_basis_points=4500,
                        reasons=["semantic_label_overlap", "data_type"],
                    )
                ],
            }
        )
        request = request.model_copy(
            update={
                "headers": [header],
                "semantic_catalog": [
                    SemanticFieldSummary(
                        code="agriculture.area",
                        version=1,
                        name="面积",
                        layer="domain",
                        data_type="decimal",
                    )
                ],
            }
        )
        result = TemplateDiffResult(
            field_decisions=[
                FieldDecision(
                    source_column_id=header.source_column_id,
                    action="REUSE_FIELD",
                    semantic_field_code="agriculture.area",
                    confidence=0.95,
                )
            ]
        )

        normalized, changed = normalize_field_catalog_references(request, result)

        assert changed is True
        assert normalized.field_decisions[0].action == "AMBIGUOUS"
        assert normalized.field_decisions[0].semantic_field_code is None
        assert normalized.field_decisions[0].requires_review is True
    Base.metadata.drop_all(engine)


def test_large_region_request_is_split_into_bounded_field_groups(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        original = request.headers[0]
        headers = [
            original.model_copy(
                update={
                    "source_column_id": f"column-{index}",
                    "header_path": [f"字段{index}"],
                    "semantic_candidates": [
                        SemanticCandidateSummary(
                            code=f"test.field_{index}",
                            version=1,
                            name=f"字段{index}",
                            data_type="text",
                            score_basis_points=9000,
                        )
                    ],
                }
            )
            for index in range(17)
        ]
        expanded = request.model_copy(
            update={
                "headers": headers,
                "new_headers": [f"字段{index}" for index in range(17)],
                "semantic_catalog": [
                    *[
                        SemanticFieldSummary(
                            code=f"test.field_{index}",
                            name=f"字段{index}",
                            layer="domain",
                            data_type="text",
                        )
                        for index in range(17)
                    ],
                    SemanticFieldSummary(
                        code="test.unused",
                        name="无关字段",
                        layer="domain",
                        data_type="text",
                    ),
                ],
            }
        )

        chunks = _chunk_recognition_request(expanded)

        assert len(chunks) == 4
        assert [len(chunk.new_headers) for chunk in chunks] == [5, 5, 5, 2]
        assert all(len(chunk.regions) == 1 for chunk in chunks)
        assert all(
            {field.code for field in chunk.semantic_catalog}
            == {
                candidate.code
                for header in chunk.headers
                for candidate in header.semantic_candidates
            }
            for chunk in chunks
        )
        assert all(
            "test.unused" not in {field.code for field in chunk.semantic_catalog}
            for chunk in chunks
        )
    Base.metadata.drop_all(engine)


def test_chunk_merge_reconciles_proposals_against_complete_published_catalog(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "village_insight.hermes.recognition.MAX_FIELDS_PER_HERMES_CALL",
        1,
    )
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = PublishedCodeProposalRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        first_header = request.headers[0]
        second_header = first_header.model_copy(
            update={
                "source_column_id": "synthetic-second-column",
                "header_path": ["第二字段"],
            }
        )
        expanded = request.model_copy(
            update={
                "headers": [first_header, second_header],
                "new_headers": ["新增字段", "第二字段"],
                "unresolved_source_column_ids": [
                    first_header.source_column_id,
                    second_header.source_column_id,
                ],
                "semantic_catalog": [
                    *request.semantic_catalog,
                    SemanticFieldSummary(
                        code="person.already_published",
                        name="已发布字段",
                        layer="domain",
                        data_type="text",
                    ),
                ],
            }
        )

        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=expanded,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                hermes_version="full-catalog-reconciliation-test",
            )
        )

        assert runtime.calls == 2
        assert [
            decision["action"] for decision in proposal.proposal["field_decisions"]
        ] == ["AMBIGUOUS", "AMBIGUOUS"]
        assert all(
            decision["proposed_field_code"] is None
            for decision in proposal.proposal["field_decisions"]
        )
        assert proposal.proposal["requires_governance"] is True
    Base.metadata.drop_all(engine)


def test_recognition_cache_avoids_duplicate_model_call(tmp_path: Path) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = FakeRuntime()
    with Session(engine) as database:
        first_item = add_item(database, source, administrative_unit_id=uuid.uuid4())
        first_request = build_diff_request(
            profile,
            add_match(database, first_item, profile),
        )
        second_item = add_item(database, source, administrative_unit_id=uuid.uuid4())
        second_request = build_diff_request(
            profile,
            add_match(database, second_item, profile),
        )
        first = asyncio.run(
            recognize_differences(
                database,
                item_id=first_item.id,
                request=first_request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                hermes_version="test",
            )
        )
        second = asyncio.run(
            recognize_differences(
                database,
                item_id=second_item.id,
                request=second_request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                hermes_version="test",
            )
        )
        assert runtime.calls == 2
        assert first.proposal == second.proposal
        assert first.proposal["recognition_passes"] == 2
        assert first.proposal["requires_governance"] is True
        assert first_request.source_samples
        assert first_request.source_samples[0].cells[0].redacted_value != "脱敏样例"
        asyncio.run(
            recognize_differences(
                database,
                item_id=first_item.id,
                request=first_request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                hermes_version="test",
            )
        )
        assert runtime.calls == 2
        records = database.query(HermesRecognitionRecord).order_by(
            HermesRecognitionRecord.created_at
        )
        assert [record.call_performed for record in records] == [
            True,
            True,
            False,
            False,
        ]
        assert all(record.input_field_count == 1 for record in records)

        provisional_template, provisional_version = create_provisional_template(
            database,
            proposal=first,
        )
        provisional_plan = approve_plan(
            database,
            item=first_item,
            template_id=provisional_template.id,
            template_version=provisional_version.version,
            layout_plan={
                "contract_version": "approved-layout-plan/v1",
                "decisions": first.proposal["layout_decisions"],
            },
            field_mappings=[],
            actor="system:hermes",
            comment="先完成原始 JSONB 入库",
            plan_source="hermes_provisional",
            proposal_id=first.id,
        )
        provisional_execution = materialize_plan(database, provisional_plan.id)
        provisional_record = database.query(DatasetRecord).one()
        assert provisional_execution.status == "partial"
        assert provisional_record.mapping_status == "partial"
        assert provisional_record.raw_data["columns"]
        assert provisional_record.semantic_data["fields"] == {}
        provisional_execution.status = "completed"
        database.flush()

        template = accept_recognition_proposal(
            database,
            proposal=first,
            actor="user",
            comment="确认合成建议",
            template_code="synthetic_sheet",
            template_name="合成表",
            domain="test",
            record_type="synthetic_record",
            record_grain="one_row_per_record",
        )
        assert isinstance(template, DocumentTemplate)
        assert template.versions[0].status == "deprecated"
        assert template.versions[1].status == "user_confirmed"
        assert first.status == "accepted"
        confirmed_version = template.versions[1]
        confirmed_version.source_metadata = {
            **confirmed_version.source_metadata,
            "approved_layout_plan": [],
        }
        plan = approve_plan(
            database,
            item=first_item,
            template_id=template.id,
            template_version=2,
            layout_plan={},
            field_mappings=[],
            actor="user",
            comment="用户确认后正式入库",
            plan_source="hermes",
            proposal_id=first.id,
            supersedes_plan_id=provisional_plan.id,
        )
        execution = materialize_plan(database, plan.id)
        record = database.query(DatasetRecord).one()
        assert plan.plan_source == "hermes"
        assert plan.proposal_id == first.id
        assert plan.layout_plan == provisional_plan.layout_plan
        assert execution.status == "completed"
        assert record.plan_source == "hermes"
        assert record.raw_data["columns"]


def test_invalid_final_cache_is_preserved_but_recomputed_under_recovery_key(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = FakeRuntime()
    with Session(engine) as database:
        first_item = add_item(database, source, administrative_unit_id=uuid.uuid4())
        first_request = build_diff_request(
            profile,
            add_match(database, first_item, profile),
        )
        asyncio.run(
            recognize_differences(
                database,
                item_id=first_item.id,
                request=first_request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                hermes_version="invalid-final-cache-test",
            )
        )
        final_record = database.query(HermesRecognitionRecord).first()
        assert final_record is not None
        damaged_cache = database.get(
            HermesRecognitionCache,
            final_record.cache_key,
        )
        assert damaged_cache is not None
        damaged_cache.response_payload = {"damaged": True}
        database.flush()

        second_item = add_item(database, source, administrative_unit_id=uuid.uuid4())
        second_request = build_diff_request(
            profile,
            add_match(database, second_item, profile),
        )
        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=second_item.id,
                request=second_request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                hermes_version="invalid-final-cache-test",
            )
        )

        assert runtime.calls == 4
        assert damaged_cache.response_payload == {"damaged": True}
        second_record = (
            database.query(HermesRecognitionRecord)
            .filter(HermesRecognitionRecord.item_id == second_item.id)
            .first()
        )
        assert second_record is not None
        assert second_record.cache_key != final_record.cache_key
        assert proposal.proposal["field_decisions"]


def test_invalid_fast_cache_is_recomputed_in_place(tmp_path: Path) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = FakeRuntime()
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        cache_key = recognition_cache_key(
            request,
            hermes_version="invalid-fast-cache-test",
            provider="fake",
            model="fake-fast",
            reasoning_model=None,
        )
        fast_cache_key = hashlib.sha256(f"{cache_key}:fast".encode()).hexdigest()
        fast_cache = HermesRecognitionCache(
            cache_key=fast_cache_key,
            hermes_version="invalid-fast-cache-test",
            prompt_version="template-diff/v18",
            schema_version="template-diff-result/v9",
            provider="fake",
            model="fake-fast",
            request_payload=request.model_dump(mode="json"),
            response_payload={"damaged": True},
        )
        database.add(fast_cache)
        database.flush()

        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                hermes_version="invalid-fast-cache-test",
            )
        )

        assert runtime.calls == 2
        assert proposal.proposal["field_decisions"]
        assert "field_decisions" in fast_cache.response_payload


def test_recognition_cache_isolated_by_model_configuration(tmp_path: Path) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = FakeRuntime()
    with Session(engine) as database:
        first_item = add_item(database, source, administrative_unit_id=uuid.uuid4())
        first_request = build_diff_request(
            profile,
            add_match(database, first_item, profile),
        )
        second_item = add_item(database, source, administrative_unit_id=uuid.uuid4())
        second_request = build_diff_request(
            profile,
            add_match(database, second_item, profile),
        )
        asyncio.run(
            recognize_differences(
                database,
                item_id=first_item.id,
                request=first_request,
                runtime=runtime,
                provider="fake",
                model="fake-fast-v1",
                hermes_version="model-isolation-test",
            )
        )
        asyncio.run(
            recognize_differences(
                database,
                item_id=second_item.id,
                request=second_request,
                runtime=runtime,
                provider="fake",
                model="fake-fast-v2",
                hermes_version="model-isolation-test",
            )
        )

        assert runtime.calls == 4
        assert len(
            {
                record.cache_key
                for record in database.query(HermesRecognitionRecord)
            }
        ) == 2


def test_reasoning_timeout_keeps_valid_fast_result_and_marks_governance(
    tmp_path: Path,
) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    runtime = FakeRuntime(fail_on_call=2)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-pro",
                hermes_version="checkpoint-test",
            )
        )
        assert runtime.calls == 2
        assert proposal.proposal["recognition_passes"] == 2
        assert "HERMES_REVIEW_INVALID_RESPONSE" in proposal.proposal[
            "governance_reason_codes"
        ]

        runtime.fail_on_call = None
        cached_proposal = asyncio.run(
            recognize_differences(
                database,
                item_id=item.id,
                request=request,
                runtime=runtime,
                provider="fake",
                model="fake-fast",
                reasoning_model="fake-pro",
                hermes_version="checkpoint-test",
            )
        )
        assert runtime.calls == 2
        assert cached_proposal.proposal == proposal.proposal


def test_recognition_rejects_invented_evidence(tmp_path: Path) -> None:
    source, profile = synthetic_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source)
        request = build_diff_request(profile, add_match(database, item, profile))
        result = TemplateDiffResult(
            template_suggestion={
                "template_code": "test.synthetic_record",
                "template_name": "合成记录",
                "domain": "test",
                "record_type": "synthetic_record",
                "confidence": 0.8,
                "evidence_ids": [request.headers[0].source_column_id],
            },
            field_decisions=[
                FieldDecision(
                    source_column_id=request.headers[0].source_column_id,
                    action="AMBIGUOUS",
                    confidence=0.2,
                    evidence_ids=["invented-cell"],
                )
            ],
        )
        with pytest.raises(RecognitionValidationError, match="unknown evidence"):
            validate_result(request, result)
