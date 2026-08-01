from pathlib import Path

from openpyxl import Workbook, load_workbook

from village_insight.templates.recomposition_regression import (
    _copy_sheet,
    _synthetic_value,
)


def test_synthetic_value_preserves_type_families_and_masks_text() -> None:
    assert isinstance(_synthetic_value(10, case_number=2, row=3, column=4), int)
    assert isinstance(_synthetic_value(10.5, case_number=2, row=3, column=4), float)
    assert _synthetic_value("张三", case_number=2, row=3, column=1) == "测试同名人员-3"
    assert _synthetic_value("张三", case_number=2, row=3, column=2) == "样例2-3-2"
    assert _synthetic_value("=SUM(A1:A2)", case_number=2, row=3, column=2) == "=SUM(A1:A2)"


def test_copy_sheet_preserves_header_and_masks_body(tmp_path: Path) -> None:
    source_book = Workbook()
    source = source_book.active
    source.append(["姓名", "人数"])
    source.append(["张三", 2])
    source.append(["李四", 3])
    target_book = Workbook()
    target = target_book.active

    _copy_sheet(
        source=source,
        target=target,
        header_end=1,
        case_number=7,
        max_body_rows=2,
    )

    output = tmp_path / "masked.xlsx"
    target_book.save(output)
    target_book.close()
    source_book.close()
    loaded = load_workbook(output, data_only=True)
    try:
        sheet = loaded.active
        assert sheet["A1"].value == "姓名"
        assert sheet["A2"].value == "测试同名人员-2"
        assert sheet["B2"].value == 700_202
    finally:
        loaded.close()
