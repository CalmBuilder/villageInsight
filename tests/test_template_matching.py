import hashlib
import uuid
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, delete, select
from sqlalchemy.orm import Session

from village_insight.db.base import Base
from village_insight.db.models import (
    DatasetRecord,
    DocumentProfile,
    DocumentTemplate,
    FieldMatch,
    GovernanceFieldResolution,
    GovernanceResolution,
    IngestionBatch,
    IngestionItem,
    ItemStatus,
    MatchType,
    ProposalStatus,
    RegionTemplate,
    RegionTemplateMatch,
    RegionTemplateVersion,
    SemanticField,
    SemanticFieldVariant,
    SemanticFieldVersion,
    SheetComposition,
    SheetCompositionMatch,
    SheetCompositionRegionSlot,
    SheetCompositionVersion,
    TemplateMatch,
    TemplateProposal,
    TemplateStatus,
    TemplateVersion,
    WorkbookRoute,
    WorkbookRouteMatch,
    WorkbookRouteSheetSlot,
    WorkbookRouteVersion,
)
from village_insight.db.schema import GovernanceFieldResolutionInput
from village_insight.hermes.recognition import (
    build_diff_request,
    published_semantic_catalog,
)
from village_insight.materialization import materialize_plan
from village_insight.parsing.candidates import select_header_candidates
from village_insight.parsing.router import ParserRouter
from village_insight.templates.field_semantics import header_paths_equivalent
from village_insight.templates.four_layer import (
    backfill_four_layer_foundation,
)
from village_insight.templates.governance import (
    commit_field_governance,
    publish_governed_regions,
)
from village_insight.templates.import_plans import (
    ImportPlanError,
    _disambiguate_duplicate_field_roles,
    _merge_governance_replacement_mappings,
    approve_hybrid_region_plan,
    approve_matched_region_plan,
    approve_plan,
    build_reused_field_match_mappings,
    ensure_layout_projection_snapshot,
    project_layout_plan,
    project_region_data_rows,
    resolve_reused_region_column,
)
from village_insight.templates.matching import (
    _component_score,
    _deterministic_auxiliary_column_reason,
    _region_template_binding,
    layout_fingerprint,
    match_profile,
    profile_region_candidates,
    profile_regions,
    region_signature,
)


def test_profile_region_candidates_preserve_nondefault_header_evidence(tmp_path: Path) -> None:
    source = tmp_path / "alternate-header-evidence.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "人数"])
    sheet.append(["甲", 1])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)
    selected = profile_regions(profile)[0]
    alternate = selected.header.model_copy(
        update={"id": f"{selected.header.id}:catalog-selected", "confidence": 0.0}
    )
    profile.sheets[0].header_candidates.append(alternate)

    candidates = profile_region_candidates(profile)

    assert (selected.sheet.id, selected.region.id, alternate.id) in {
        (candidate.sheet.id, candidate.region.id, candidate.header.id) for candidate in candidates
    }


def create_profile(tmp_path: Path):
    source = tmp_path / "synthetic.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "人数"])
    sheet.append(["张三", 2])
    workbook.save(source)
    workbook.close()
    return source, ParserRouter().profile(source)


def test_new_field_match_extends_reused_region_plan(tmp_path: Path) -> None:
    source, profile = create_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        source_region = profile_regions(profile)[0]
        new_column = source_region.header.columns[1]
        database.add(
            FieldMatch(
                item_id=item.id,
                sheet_id=source_region.sheet.id,
                region_id=source_region.region.id,
                header_id=source_region.header.id,
                source_column_id=new_column.source_column_id,
                header_path=new_column.header_path,
                observed_data_type=None,
                semantic_field_code="test.new_field",
                semantic_field_version=2,
                match_type=MatchType.EXACT,
                score_basis_points=10_000,
                context={},
                differences={"matched_by": "field_catalog"},
                requires_hermes=False,
                matcher_version="test",
            )
        )
        database.flush()

        additions = build_reused_field_match_mappings(
            database,
            item=item,
            reused_decisions=[
                {
                    "region_candidate_id": source_region.region.id,
                    "layout_mode": "table",
                }
            ],
            reused_mappings=[
                {
                    "source_column_id": source_region.header.columns[0].source_column_id,
                }
            ],
        )

        assert additions == [
            {
                "sheet_id": source_region.sheet.id,
                "region_id": source_region.region.id,
                "source_column_id": new_column.source_column_id,
                "header_path": new_column.header_path,
                "semantic_field_code": "test.new_field",
                "semantic_field_version": 2,
                "role": None,
                "normalizer": None,
                "required": False,
                "field_match_id": str(additions[0]["field_match_id"]),
                "role_source": None,
            }
        ]
    Base.metadata.drop_all(engine)


def test_region_score_ignores_uniform_document_title_parent() -> None:
    score, differences = _component_score(
        {
            "kind": "table",
            "columns": 2,
            "header_depth": 1,
            "headers": [["农作物登记表", "序号"], ["农作物登记表", "姓名"]],
        },
        {
            "kind": "table",
            "columns": 2,
            "header_depth": 1,
            "headers": [["序号"], ["姓名"]],
        },
    )

    assert score == 10_000
    assert differences == {
        "missing_headers": [],
        "new_headers": [],
        "structural_mismatches": [],
    }


def test_governance_replacement_preserves_existing_mappings_and_overrides_decision() -> None:
    assert _merge_governance_replacement_mappings(
        [
            {"source_column_id": "column:a", "semantic_field_code": "person.name"},
            {"source_column_id": "column:b", "semantic_field_code": "old.field"},
        ],
        [
            {"source_column_id": "column:b", "semantic_field_code": "new.field"},
            {"source_column_id": "column:c", "semantic_field_code": "person.birth_date"},
        ],
    ) == [
        {"source_column_id": "column:a", "semantic_field_code": "person.name"},
        {"source_column_id": "column:b", "semantic_field_code": "new.field"},
        {"source_column_id": "column:c", "semantic_field_code": "person.birth_date"},
    ]


def test_field_binding_ignores_only_document_title_prefix() -> None:
    assert header_paths_equivalent(
        ["姓名"],
        ["2024年农作物耕种登记表", "姓名"],
    )
    assert header_paths_equivalent(
        ["碧江区坝黄镇官庄村医保进度倒排期统计表（按照85%计算）29日通报", "序号"],
        ["碧江区坝黄镇官庄村医保进度倒排期统计表（按照85%计算）12月2日通报", "序号"],
    )
    assert not header_paths_equivalent(
        ["姓名"],
        ["监护人", "姓名"],
    )


def test_profile_regions_excludes_formula_only_derived_column(tmp_path: Path) -> None:
    source = tmp_path / "formula-helper.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "人数", None, None])
    sheet.append(["张三", 2, None, "=B2*2"])
    sheet.append(["李四", 3, None, "=B3*2"])
    workbook.save(source)
    workbook.close()

    regions = profile_regions(ParserRouter().profile(source))

    assert len(regions) == 1
    assert regions[0].region.bounds.range == "A1:B3"


@pytest.mark.parametrize(
    "values",
    [
        ["=B2*2", "=B3*2"],
        ["甲组", "乙组"],
    ],
)
def test_unnamed_columns_are_always_excluded_from_semantic_mapping(
    tmp_path: Path,
    values: list[str],
) -> None:
    source = tmp_path / "unnamed-column.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "人数", None])
    sheet.append(["张三", 2, values[0]])
    sheet.append(["李四", 3, values[1]])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        match_profile(database, item_id=item.id, profile=profile)
        unnamed = database.scalar(
            select(FieldMatch).where(
                FieldMatch.item_id == item.id,
                FieldMatch.header_path == [],
            )
        )

        assert unnamed is not None
        assert unnamed.requires_hermes is False
        assert unnamed.differences["ignore_reason"] == "unnamed_column"
        assert unnamed.differences["matched_by"] == "deterministic_auxiliary_column"
        assert unnamed.differences["ignored"] is True
    Base.metadata.drop_all(engine)


def test_observed_value_headers_are_excluded_from_semantic_mapping(
    tmp_path: Path,
) -> None:
    source = tmp_path / "observed-value-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "123456789012345678"])
    sheet.append(["测试甲", "普通值甲"])
    sheet.append(["测试乙", "普通值乙"])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    assert (
        _deterministic_auxiliary_column_reason(
            column=SimpleNamespace(header_path=["123456789012345678"])
        )
        == "observed_value_header"
    )
    signature = region_signature(
        profile.sheets[0].region_candidates[0],
        profile.sheets[0].header_candidates[0],
    )
    assert signature["columns"] == 1
    assert signature["headers"] == [["姓名"]]


def test_layout_fingerprint_ignores_data_row_count_and_footer_regions(
    tmp_path: Path,
) -> None:
    short = tmp_path / "short.xlsx"
    long = tmp_path / "long.xlsx"
    for path, rows, include_footer in (
        (short, [["张三", 2]], False),
        (long, [["张三", 2], ["李四", 3], ["王五", 4]], True),
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["姓名", "人数"])
        for row in rows:
            sheet.append(row)
        if include_footer:
            for _ in range(3):
                sheet.append([None, None])
            sheet.append(["领导签字：", None])
        workbook.save(path)
        workbook.close()

    assert layout_fingerprint(ParserRouter().profile(short)) == layout_fingerprint(
        ParserRouter().profile(long)
    )


def add_item(database: Session, source: Path, profile):
    payload = source.read_bytes()
    batch = IngestionBatch(name="synthetic", total_files=1)
    database.add(batch)
    database.flush()
    item = IngestionItem(
        id=uuid.uuid4(),
        batch_id=batch.id,
        original_name=source.name,
        source_path=str(source),
        source_sha256=hashlib.sha256(payload).hexdigest(),
        size_bytes=len(payload),
        status=ItemStatus.NEEDS_REVIEW,
    )
    database.add(item)
    database.flush()
    database.add(
        DocumentProfile(
            item_id=item.id,
            contract_version=profile.contract_version,
            source_sha256=profile.source_sha256,
            parser_name=profile.parser_name,
            parser_version=profile.parser_version,
            profile=profile.model_dump(mode="json"),
        )
    )
    return item


def add_published_template(database: Session, fingerprint: str):
    template = DocumentTemplate(code="person_roster", published_version=1)
    version = TemplateVersion(
        version=1,
        name="人员名册",
        status=TemplateStatus.PUBLISHED,
        layout_fingerprint=fingerprint,
        definition={
            "contract_version": "document-template/v1",
            "domain": "population",
            "region_kind": "table",
            "record_type": "person",
            "record_grain": "one_row_per_person",
            "field_bindings": [],
            "data_row_rules": [],
            "exclusion_rules": [],
            "metric_codes": [],
        },
        source="bootstrap",
    )
    template.versions.append(version)
    database.add(template)
    database.flush()
    return template


def test_published_field_variant_reuses_field_without_region_template(
    tmp_path: Path,
) -> None:
    source, profile = create_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        field = SemanticField(code="person.name", published_version=1)
        version = SemanticFieldVersion(
            version=1,
            name="人员姓名",
            layer="base",
            data_type="text",
            status=TemplateStatus.PUBLISHED,
        )
        version.variants.append(
            SemanticFieldVariant(
                variant_key=hashlib.sha256(b"name-header").hexdigest(),
                kind="header_path",
                normalized_value="姓名",
                alias=None,
                header_path=["姓名"],
                parent_path=[],
                role=None,
                domain=None,
                record_type=None,
                observed_data_type="text",
                unit_dimension=None,
                source="codex",
                confidence_basis_points=10_000,
                evidence={"source_count": 2},
            )
        )
        field.versions.append(version)
        database.add(field)
        item = add_item(database, source, profile)
        database.flush()

        match_profile(database, item_id=item.id, profile=profile)

        name_match = database.scalar(
            select(FieldMatch).where(
                FieldMatch.item_id == item.id,
                FieldMatch.header_path == ["姓名"],
            )
        )
        assert name_match is not None
        assert name_match.match_type == "exact"
        assert name_match.semantic_field_code == "person.name"
        assert name_match.requires_hermes is False

        baseline = (
            name_match.match_type,
            name_match.score_basis_points,
            name_match.semantic_field_code,
            name_match.requires_hermes,
        )
        version.source = "auto_governance"
        version.source_metadata = {
            "source_contract": "four-layer-template-source/v1",
            "source": "auto_governance",
        }
        database.execute(delete(FieldMatch).where(FieldMatch.item_id == item.id))
        database.flush()
        match_profile(database, item_id=item.id, profile=profile)
        second_match = database.scalar(
            select(FieldMatch).where(
                FieldMatch.item_id == item.id,
                FieldMatch.header_path == ["姓名"],
            )
        )
        assert second_match is not None
        assert (
            second_match.match_type,
            second_match.score_basis_points,
            second_match.semantic_field_code,
            second_match.requires_hermes,
        ) == baseline
    Base.metadata.drop_all(engine)


def add_region_template(
    database: Session,
    *,
    code: str,
    name: str,
    header_paths: list[list[str]],
    record_type: str,
) -> DocumentTemplate:
    field_bindings = []
    for index, header_path in enumerate(header_paths, start=1):
        field = SemanticField(
            code=f"{code}.field_{index}",
            published_version=1,
        )
        field.versions.append(
            SemanticFieldVersion(
                version=1,
                name=header_path[-1],
                layer="domain",
                data_type="text",
                status=TemplateStatus.PUBLISHED,
            )
        )
        database.add(field)
        field_bindings.append(
            {
                "source_column_id": f"{code}:column:{index}",
                "header_path": header_path,
                "semantic_field_code": field.code,
                "semantic_field_version": 1,
                "required": False,
            }
        )
    template = DocumentTemplate(code=code, published_version=1)
    template.versions.append(
        TemplateVersion(
            version=1,
            name=name,
            status=TemplateStatus.PUBLISHED,
            layout_fingerprint=hashlib.sha256(code.encode()).hexdigest(),
            definition={
                "contract_version": "document-template/v1",
                "domain": code,
                "region_kind": "table",
                "record_type": record_type,
                "record_grain": f"one_row_per_{record_type}",
                "field_bindings": field_bindings,
                "data_row_rules": [],
                "exclusion_rules": [],
                "metric_codes": [],
            },
            source="bootstrap",
            source_metadata={
                "layout_projection_snapshot": {
                    "contract_version": "layout-projection-snapshot/v1",
                    "decisions": [
                        {
                            "sheet_index": 0,
                            "header_signature": header_paths,
                            "data_start_offset_from_header_end": 1,
                            "data_end_gap_from_region_end": 0,
                            "excluded_row_offsets": [],
                            "classification": "unknown",
                        }
                    ],
                }
            },
        )
    )
    database.add(template)
    database.flush()
    return template


def test_exact_match_skips_hermes_and_approved_plan_is_idempotent(
    tmp_path: Path,
) -> None:
    source, profile = create_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        template = add_published_template(database, layout_fingerprint(profile))
        match = match_profile(database, item_id=item.id, profile=profile)

        assert match.match_type == "exact"
        assert match.score_basis_points == 10_000
        assert match.requires_hermes is True

        first = approve_plan(
            database,
            item=item,
            template_id=template.id,
            template_version=1,
            layout_plan={"region": "A1:B2"},
            field_mappings=[],
            actor="tester",
            comment="synthetic",
        )
        second = approve_plan(
            database,
            item=item,
            template_id=template.id,
            template_version=1,
            layout_plan={"region": "A1:B2"},
            field_mappings=[],
            actor="tester",
            comment="retry",
        )
        assert first.id == second.id
        assert second.layout_plan == {"region": "A1:B2"}
        corrected = approve_plan(
            database,
            item=item,
            template_id=template.id,
            template_version=1,
            layout_plan={"region": "A1:B3"},
            field_mappings=[],
            actor="tester",
            comment="corrected",
            supersedes_plan_id=first.id,
        )
        assert corrected.id != first.id
        assert corrected.revision == 2
        assert corrected.supersedes_plan_id == first.id
        assert item.status == ItemStatus.MATERIALIZING


def test_approved_plan_resolves_duplicate_header_by_current_source_id(
    tmp_path: Path,
) -> None:
    source = tmp_path / "duplicate-header-binding.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "姓名"])
    sheet.append(["张三", "李四"])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)
    selected_column = profile_regions(profile)[0].header.columns[1]

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        template = add_published_template(database, layout_fingerprint(profile))
        definition = dict(template.versions[0].definition)
        definition["field_bindings"] = [
            {
                "source_column_id": selected_column.source_column_id,
                "header_path": ["姓名"],
                "semantic_field_code": "person.name",
                "semantic_field_version": 1,
                "required": False,
            }
        ]
        template.versions[0].definition = definition
        match_profile(database, item_id=item.id, profile=profile)

        plan = approve_plan(
            database,
            item=item,
            template_id=template.id,
            template_version=1,
            layout_plan={"region": "A1:B2"},
            field_mappings=[],
            actor="tester",
            comment="resolve duplicate header binding",
        )

        assert plan.field_mappings == [
            {
                "source_column_id": selected_column.source_column_id,
                "header_path": ["姓名"],
                "semantic_field_code": "person.name",
                "semantic_field_version": 1,
                "role": None,
                "normalizer": None,
                "required": False,
            }
        ]
    Base.metadata.drop_all(engine)


def test_approved_plan_rejects_template_change(tmp_path: Path) -> None:
    source, profile = create_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        first_template = add_published_template(database, layout_fingerprint(profile))
        second_template = DocumentTemplate(code="other", published_version=1)
        second_template.versions.append(
            TemplateVersion(
                version=1,
                name="其他",
                status=TemplateStatus.PUBLISHED,
                layout_fingerprint=hashlib.sha256(b"other").hexdigest(),
                definition=first_template.versions[0].definition,
                source="bootstrap",
            )
        )
        database.add(second_template)
        database.flush()
        match_profile(database, item_id=item.id, profile=profile)
        approve_plan(
            database,
            item=item,
            template_id=first_template.id,
            template_version=1,
            layout_plan={},
            field_mappings=[],
            actor="tester",
            comment="",
        )
        with pytest.raises(ImportPlanError, match="explicitly supersede"):
            approve_plan(
                database,
                item=item,
                template_id=second_template.id,
                template_version=1,
                layout_plan={},
                field_mappings=[],
                actor="tester",
                comment="",
            )


def test_region_templates_are_reused_across_workbook_compositions(
    tmp_path: Path,
) -> None:
    source = tmp_path / "mixed.xlsx"
    workbook = Workbook()
    people = workbook.active
    people.title = "任意人员页名"
    people.append(["姓名", "人数"])
    people.append(["张三", 1])
    land = workbook.create_sheet("任意土地页名")
    land.append(["地块", "面积"])
    land.append(["一号地", 3])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        people_template = add_region_template(
            database,
            code="people_region",
            name="人员区域",
            header_paths=[["姓名"], ["人数"]],
            record_type="person",
        )
        land_template = add_region_template(
            database,
            code="land_region",
            name="土地区域",
            header_paths=[["地块"], ["面积"]],
            record_type="land",
        )

        summary = match_profile(database, item_id=item.id, profile=profile)
        region_matches = list(
            database.scalars(
                select(RegionTemplateMatch).where(RegionTemplateMatch.item_id == item.id)
            )
        )

        assert summary.match_type == "exact"
        assert summary.requires_hermes is False
        assert summary.total_regions == 2
        assert summary.matched_regions == 2
        assert summary.coverage_basis_points == 10_000
        assert {match.template_id for match in region_matches} == {
            people_template.id,
            land_template.id,
        }
        assert all(match.template_region_component_id for match in region_matches)

        backfill_four_layer_foundation(database)
        match_profile(database, item_id=item.id, profile=profile)
        plan = approve_matched_region_plan(database, item=item)
        assert plan.layout_plan["contract_version"] == "approved-region-import-plan/v2"
        assert len(plan.layout_plan["decisions"]) == 2
        assert {decision["template_id"] for decision in plan.layout_plan["decisions"]} == {
            str(people_template.id),
            str(land_template.id),
        }
        execution = materialize_plan(database, plan.id)
        records = list(
            database.scalars(select(DatasetRecord).where(DatasetRecord.approved_plan_id == plan.id))
        )
        assert execution.record_count == 2
        assert {record.record_type for record in records} == {"person", "land"}
        assert {record.template_id for record in records} == {
            people_template.id,
            land_template.id,
        }


def test_hermes_request_contains_only_unmatched_regions(tmp_path: Path) -> None:
    source = tmp_path / "partially-known.xlsx"
    workbook = Workbook()
    people = workbook.active
    people.append(["姓名", "人数"])
    people.append(["张三", 1])
    land = workbook.create_sheet()
    land.append(["地块", "面积"])
    land.append(["一号地", 3])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        add_region_template(
            database,
            code="known_people_region",
            name="已知人员区域",
            header_paths=[["姓名"], ["人数"]],
            record_type="person",
        )
        summary = match_profile(database, item_id=item.id, profile=profile)
        request = build_diff_request(profile, summary)

        assert summary.match_type == "partial"
        assert summary.matched_regions == 1
        assert summary.total_regions == 2
        assert len(request.regions) == 1
        assert {header.header_path[0] for header in request.headers} == {
            "地块",
            "面积",
        }
        assert {sample.region_candidate_id for sample in request.source_samples} == {
            request.regions[0].candidate_id
        }


def test_field_match_reuses_published_semantics_inside_unmatched_region(
    tmp_path: Path,
) -> None:
    source = tmp_path / "field-reuse.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "新增备注"])
    sheet.append(["张三", "示例"])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        field = SemanticField(code="person.name", published_version=1)
        field.versions.append(
            SemanticFieldVersion(
                version=1,
                name="姓名",
                layer="base",
                data_type="text",
                aliases=["人员姓名"],
                status=TemplateStatus.PUBLISHED,
            )
        )
        database.add(field)
        template = DocumentTemplate(code="person_region", published_version=1)
        template.versions.append(
            TemplateVersion(
                version=1,
                name="人员区域",
                status=TemplateStatus.PUBLISHED,
                layout_fingerprint=hashlib.sha256(b"person-region").hexdigest(),
                definition={
                    "contract_version": "document-template/v1",
                    "domain": "person",
                    "region_kind": "table",
                    "record_type": "person",
                    "record_grain": "one_row_per_person",
                    "field_bindings": [
                        {
                            "source_column_id": "seed:name",
                            "header_path": ["姓名"],
                            "semantic_field_code": "person.name",
                            "semantic_field_version": 1,
                        }
                    ],
                    "data_row_rules": [],
                    "exclusion_rules": [],
                    "metric_codes": [],
                },
                source="bootstrap",
                source_metadata={
                    "layout_projection_snapshot": {
                        "contract_version": "layout-projection-snapshot/v1",
                        "decisions": [
                            {
                                "sheet_index": 0,
                                "header_signature": [["姓名"]],
                                "data_start_offset_from_header_end": 1,
                                "data_end_gap_from_region_end": 0,
                                "excluded_row_offsets": [],
                                "classification": "table",
                            }
                        ],
                    }
                },
            )
        )
        database.add(template)
        database.flush()

        match_profile(database, item_id=item.id, profile=profile)
        field_matches = list(
            database.scalars(select(FieldMatch).where(FieldMatch.item_id == item.id))
        )

        by_header = {entry.header_path[-1]: entry for entry in field_matches}
        assert by_header["姓名"].match_type == "exact"
        assert by_header["姓名"].semantic_field_code == "person.name"
        assert by_header["姓名"].requires_hermes is False
        assert by_header["新增备注"].requires_hermes is True
        request = build_diff_request(
            profile,
            database.get(TemplateMatch, item.id),
            field_matches=field_matches,
        )
        assert request.new_headers == ["新增备注"]
    assert request.unresolved_source_column_ids == [by_header["新增备注"].source_column_id]


def test_governance_publishes_new_field_path_and_region_for_immediate_reuse(
    tmp_path: Path,
) -> None:
    source, profile = create_profile(tmp_path)
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        name_field = SemanticField(code="person.name", published_version=1)
        name_version = SemanticFieldVersion(
            version=1,
            name="姓名",
            layer="base",
            data_type="text",
            status=TemplateStatus.PUBLISHED,
        )
        name_version.variants.append(
            SemanticFieldVariant(
                variant_key=hashlib.sha256(b"person-name-path").hexdigest(),
                kind="header_path",
                normalized_value="姓名",
                header_path=["姓名"],
                parent_path=[],
                source="bootstrap",
                confidence_basis_points=10_000,
            )
        )
        name_field.versions.append(name_version)
        database.add(name_field)
        item = add_item(database, source, profile)
        summary = match_profile(database, item_id=item.id, profile=profile)
        matches = list(database.scalars(select(FieldMatch).where(FieldMatch.item_id == item.id)))
        unresolved = next(match for match in matches if match.requires_hermes)
        region = profile_regions(profile)[0]
        proposal = TemplateProposal(
            source="hermes",
            source_item_id=item.id,
            proposal={
                "contract_version": "template-diff-result/v2",
                "template_suggestion": {
                    "template_code": "population.person",
                    "template_name": "人口信息",
                    "domain": "population",
                    "record_type": "person",
                    "confidence": 0.9,
                    "evidence_ids": [],
                },
                "record_grain": {
                    "value": "one_row_per_person",
                    "confidence": 0.9,
                    "evidence_ids": [],
                },
                "layout_decisions": [
                    {
                        "region_candidate_id": region.region.id,
                        "header_candidate_id": region.header.id,
                        "data_start_row": 2,
                        "data_end_row": 2,
                        "excluded_rows": [],
                        "classification": "table",
                        "materialize": True,
                        "confidence": 0.9,
                        "evidence_ids": [],
                        "merge_decisions": [],
                    }
                ],
                "field_decisions": [
                    {
                        "source_column_id": unresolved.source_column_id,
                        "action": "PROPOSE_NEW_FIELD",
                        "proposed_field_code": "population.person_count",
                        "layer": "domain",
                        "data_type": "integer",
                        "confidence": 0.9,
                        "evidence_ids": [],
                        "requires_review": True,
                    }
                ],
                "requires_governance": True,
            },
            status=ProposalStatus.PENDING,
        )
        database.add(proposal)
        database.flush()

        governance = commit_field_governance(
            database,
            proposal=proposal,
            resolutions=[
                GovernanceFieldResolutionInput(
                    source_column_id=unresolved.source_column_id,
                    mode="create_new",
                    new_field_code="population.person_count",
                    new_field_name="人数",
                    new_field_layer="domain",
                    new_field_data_type="integer",
                    learn_path=True,
                )
            ],
            domain="population",
            record_type="person",
            record_grain="one_row_per_person",
            actor="admin",
            actor_user_id=uuid.uuid4(),
            comment="确认人数字段",
        )
        refs = publish_governed_regions(
            database,
            proposal=proposal,
            governance=governance,
            template_name="人口信息",
            actor="admin",
            actor_user_id=uuid.uuid4(),
        )

        published = database.scalar(
            select(SemanticField).where(SemanticField.code == "population.person_count")
        )
        assert published is not None
        assert published.published_version == 1
        assert refs
        assert database.query(GovernanceResolution).count() == 1
        audit = database.query(GovernanceFieldResolution).one()
        assert audit.sheet_name == profile.sheets[0].name
        assert audit.column_coordinate == "B"
        assert audit.header_path == ["人数"]

        summary = match_profile(database, item_id=item.id, profile=profile)
        rematched = list(database.scalars(select(FieldMatch).where(FieldMatch.item_id == item.id)))
        assert summary.match_type == "exact"
        assert all(not match.requires_hermes for match in rematched)
        assert {match.semantic_field_code for match in rematched} == {
            "person.name",
            "population.person_count",
        }


def test_governance_reuses_one_initial_field_version_across_same_submission(
    tmp_path: Path,
) -> None:
    source = tmp_path / "duplicate-governance-field.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["户别", "户别"])
    sheet.append(["家庭户", "家庭户"])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        field = SemanticField(code="household.type", published_version=1)
        field.versions.append(
            SemanticFieldVersion(
                version=1,
                name="户别",
                layer="base",
                data_type="text",
                status=TemplateStatus.PUBLISHED,
            )
        )
        database.add(field)
        item = add_item(database, source, profile)
        match_profile(database, item_id=item.id, profile=profile)
        unresolved = list(
            database.scalars(
                select(FieldMatch).where(
                    FieldMatch.item_id == item.id,
                    FieldMatch.requires_hermes.is_(True),
                )
            )
        )
        assert len(unresolved) == 2
        proposal = TemplateProposal(
            source="hermes",
            source_item_id=item.id,
            proposal={"field_decisions": []},
            status=ProposalStatus.PENDING,
        )
        database.add(proposal)
        database.flush()

        governance = commit_field_governance(
            database,
            proposal=proposal,
            resolutions=[
                GovernanceFieldResolutionInput(
                    source_column_id=match.source_column_id,
                    mode="reuse_existing",
                    semantic_field_code="household.type",
                    expected_field_version=1,
                    learn_path=True,
                )
                for match in unresolved
            ],
            domain="population",
            record_type="person",
            record_grain="one_row_per_person",
            actor="admin",
            actor_user_id=uuid.uuid4(),
            comment="同批复用",
        )

        assert governance.field_versions == {"household.type": 2}
        assert field.published_version == 2
        assert database.query(GovernanceFieldResolution).count() == 2
    Base.metadata.drop_all(engine)


def test_unresolved_field_request_contains_ranked_published_candidates(
    tmp_path: Path,
) -> None:
    source = tmp_path / "candidate-choice.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["户主姓名", "家庭住址"])
    sheet.append(["张某", "某村一组"])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        for code, name, aliases in (
            ("person.name", "姓名", ["人员姓名"]),
            ("household.address", "家庭住址", ["住址"]),
        ):
            field = SemanticField(code=code, published_version=1)
            field.versions.append(
                SemanticFieldVersion(
                    version=1,
                    name=name,
                    layer="base",
                    data_type="text",
                    aliases=aliases,
                    status=TemplateStatus.PUBLISHED,
                )
            )
            database.add(field)
        database.flush()

        summary = match_profile(database, item_id=item.id, profile=profile)
        field_matches = list(
            database.scalars(select(FieldMatch).where(FieldMatch.item_id == item.id))
        )
        request = build_diff_request(
            profile,
            summary,
            semantic_catalog=published_semantic_catalog(database),
            field_matches=field_matches,
        )

        by_header = {entry.header_path[-1]: entry for entry in request.headers}
        name_candidates = by_header["户主姓名"].semantic_candidates
        address_candidates = by_header["家庭住址"].semantic_candidates
        assert name_candidates[0].code == "person.name"
        assert "normalized_base_alias" in name_candidates[0].reasons
        assert by_header["户主姓名"].context["role"] == "household_head"
        assert by_header["户主姓名"].context["base_label"] == "姓名"
        assert address_candidates[0].code == "household.address"
        assert name_candidates[0].data_type == "text"
        assert by_header["户主姓名"].observed_data_type == "text"


def test_duplicate_semantic_columns_keep_one_value_and_gain_stable_roles() -> None:
    mappings = [
        {
            "region_id": "region-1",
            "source_column_id": "person",
            "header_path": ["姓名"],
            "semantic_field_code": "person.name",
            "role": None,
        },
        {
            "region_id": "region-1",
            "source_column_id": "head-1",
            "header_path": ["户主姓名"],
            "semantic_field_code": "person.name",
            "role": None,
            "mapping_source": "hermes",
        },
        {
            "region_id": "region-1",
            "source_column_id": "head-2",
            "header_path": ["户主姓名"],
            "semantic_field_code": "person.name",
            "role": None,
            "mapping_source": "hermes",
        },
        {
            "region_id": "region-1",
            "source_column_id": "spouse",
            "header_path": ["配偶姓名"],
            "semantic_field_code": "person.name",
            "role": "spouse",
            "mapping_source": "hermes",
        },
    ]

    normalized = _disambiguate_duplicate_field_roles(mappings)

    by_column = {mapping["source_column_id"]: mapping for mapping in normalized}
    assert by_column["person"]["role"] is None
    assert by_column["head-1"]["role"] == "household_head"
    assert by_column["head-2"]["role"] == "household_head_2"
    assert by_column["spouse"]["role"] == "spouse"
    assert by_column["head-1"]["requires_review"] is True
    assert mappings[1]["role"] is None


def test_duplicate_existing_role_is_downgraded_without_blocking_materialization() -> None:
    mappings = [
        {
            "region_id": "region-1",
            "source_column_id": "reported",
            "header_path": ["与户主关系"],
            "semantic_field_code": "household.relationship_to_head",
            "role": "household_head",
        },
        {
            "region_id": "region-1",
            "source_column_id": "registry",
            "header_path": ["与户主关系（与派出所人口比对）"],
            "semantic_field_code": "household.relationship_to_head",
            "role": "household_head",
            "mapping_source": "hermes",
        },
    ]

    normalized = _disambiguate_duplicate_field_roles(mappings)

    assert normalized[0]["role"] == "household_head"
    assert normalized[1]["role"] == "registry_comparison"
    assert normalized[1]["requires_review"] is True
    assert normalized[1]["role_source"] == "backend_conflict_disambiguation"


def test_hybrid_plan_keeps_reused_and_hermes_regions(tmp_path: Path) -> None:
    source = tmp_path / "hybrid.xlsx"
    workbook = Workbook()
    people = workbook.active
    people.append(["姓名", "人数"])
    people.append(["张三", 1])
    land = workbook.create_sheet()
    land.append(["地块", "面积"])
    land.append(["一号地", 3])
    workbook.save(source)
    workbook.close()
    profile = ParserRouter().profile(source)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        known = add_region_template(
            database,
            code="hybrid_people",
            name="人员区域",
            header_paths=[["姓名"], ["人数"]],
            record_type="person",
        )
        summary = match_profile(database, item_id=item.id, profile=profile)
        assert summary.matched_regions == 1
        backfill_four_layer_foundation(database)
        summary = match_profile(database, item_id=item.id, profile=profile)
        assert summary.matched_regions == 1

        proposal = TemplateProposal(
            tenant_id=item.tenant_id,
            administrative_unit_id=item.administrative_unit_id,
            created_by_user_id=item.created_by_user_id,
            idempotency_key="hybrid-proposal",
            source="hermes",
            source_item_id=item.id,
            proposal={},
            status=ProposalStatus.PENDING,
        )
        database.add(proposal)
        database.flush()
        provisional = DocumentTemplate(code="hybrid_land")
        provisional.versions.append(
            TemplateVersion(
                version=1,
                name="土地临时区域",
                status=TemplateStatus.ADMIN_REVIEW,
                layout_fingerprint=hashlib.sha256(b"hybrid-land").hexdigest(),
                definition={
                    "contract_version": "document-template/v1",
                    "domain": "land",
                    "region_kind": "table",
                    "record_type": "land",
                    "record_grain": "one_row_per_land",
                    "field_bindings": [],
                    "data_row_rules": [],
                    "exclusion_rules": [],
                    "metric_codes": [],
                },
                source="hermes_provisional",
                source_metadata={"proposal_id": str(proposal.id)},
            )
        )
        database.add(provisional)
        database.flush()

        unknown_match = database.scalar(
            select(RegionTemplateMatch).where(
                RegionTemplateMatch.item_id == item.id,
                RegionTemplateMatch.requires_hermes.is_(True),
            )
        )
        assert unknown_match is not None
        unknown_sheet = next(
            sheet for sheet in profile.sheets if sheet.id == unknown_match.sheet_id
        )
        unknown_region = next(
            region
            for region in unknown_sheet.region_candidates
            if region.id == unknown_match.region_id
        )
        unknown_header = next(
            header
            for header in unknown_sheet.header_candidates
            if header.id == unknown_match.header_id
        )
        area_column = next(
            column for column in unknown_header.columns if column.header_path[-1] == "面积"
        )
        name_column = next(
            column for column in unknown_header.columns if column.header_path[-1] == "地块"
        )
        name_field = SemanticField(code="land.name", published_version=1)
        name_field.versions.append(
            SemanticFieldVersion(
                version=1,
                name="地块",
                layer="domain",
                data_type="text",
                status=TemplateStatus.PUBLISHED,
            )
        )
        area_field = SemanticField(code="land.area", published_version=1)
        area_field.versions.append(
            SemanticFieldVersion(
                version=1,
                name="面积",
                layer="domain",
                data_type="decimal",
                status=TemplateStatus.PUBLISHED,
            )
        )
        database.add_all([name_field, area_field])
        name_match = database.scalar(
            select(FieldMatch).where(
                FieldMatch.item_id == item.id,
                FieldMatch.source_column_id == name_column.source_column_id,
            )
        )
        assert name_match is not None
        name_match.semantic_field_code = "land.name"
        name_match.semantic_field_version = 1
        name_match.match_type = MatchType.EXACT
        name_match.score_basis_points = 10_000
        name_match.requires_hermes = False
        database.flush()
        plan = approve_hybrid_region_plan(
            database,
            item=item,
            provisional_template_id=provisional.id,
            provisional_template_version=1,
            proposal_id=proposal.id,
            hermes_layout_decisions=[
                {
                    "region_candidate_id": unknown_region.id,
                    "header_candidate_id": unknown_header.id,
                    "data_start_row": max(unknown_header.header_rows) + 1,
                    "data_end_row": unknown_region.bounds.max_row,
                    "excluded_rows": [],
                    "classification": "table",
                    "evidence_ids": [],
                    "merge_decisions": [],
                }
            ],
            hermes_field_decisions=[
                {
                    "source_column_id": area_column.source_column_id,
                    "action": "REUSE_FIELD",
                    "semantic_field_code": "land.area",
                    "role": None,
                }
            ],
        )
        assert any(
            mapping["semantic_field_code"] == "land.area" and mapping["mapping_source"] == "hermes"
            for mapping in plan.field_mappings
        )
        assert any(
            mapping["semantic_field_code"] == "land.name"
            and mapping.get("field_match_id") == str(name_match.id)
            for mapping in plan.field_mappings
        )
        hermes_mapping = next(
            mapping for mapping in plan.field_mappings if mapping.get("mapping_source") == "hermes"
        )
        assert hermes_mapping["header_path"] == area_column.header_path
        land_decision = next(
            decision
            for decision in plan.layout_plan["decisions"]
            if decision["region_candidate_id"] == unknown_region.id
        )
        assert {mapping["semantic_field_code"] for mapping in land_decision["field_mappings"]} == {
            "land.area",
            "land.name",
        }
        execution = materialize_plan(database, plan.id)
        records = list(
            database.scalars(select(DatasetRecord).where(DatasetRecord.approved_plan_id == plan.id))
        )
        assert execution.record_count == 2
        assert execution.status == "completed"
        assert {record.record_type for record in records} == {"person", "land"}
        land_record = next(record for record in records if record.record_type == "land")
        assert land_record.mapping_status == "complete"
        assert set(land_record.semantic_data["fields"]) == {
            "land.area",
            "land.name",
        }
        assert {record.template_id for record in records} == {
            known.id,
            provisional.id,
        }


def test_approved_layout_projects_relative_data_range_to_new_evidence(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    current = tmp_path / "current.xlsx"
    for path, names in (
        (source, ["张三", "李四"]),
        (current, ["张三", "李四", "王五", "赵六"]),
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["姓名", "人数"])
        for name in names:
            sheet.append([name, 1])
        sheet.append(["合计", len(names)])
        workbook.save(path)
        workbook.close()
    source_profile = ParserRouter().profile(source)
    current_profile = ParserRouter().profile(current)
    source_header = select_header_candidates(source_profile.sheets[0].header_candidates)[0]
    source_region = next(
        region
        for region in source_profile.sheets[0].region_candidates
        if region.id == source_header.region_id
    )
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        source_item = add_item(database, source, source_profile)
        template = add_published_template(
            database,
            layout_fingerprint(source_profile),
        )
        version = template.versions[0]
        version.source_metadata = {
            "source_item_id": str(source_item.id),
            "approved_layout_plan": [
                {
                    "region_candidate_id": source_region.id,
                    "header_candidate_id": source_header.id,
                    "data_start_row": 2,
                    "data_end_row": 3,
                    "data_start_column": 1,
                    "data_end_column": 1,
                    "excluded_rows": [],
                    "materialize": False,
                }
            ],
        }
        database.flush()

        projected = project_layout_plan(
            database,
            version=version,
            current_profile=current_profile,
        )
        ensure_layout_projection_snapshot(database, version)
        database.delete(database.get(DocumentProfile, source_item.id))
        database.flush()
        projected_without_source = project_layout_plan(
            database,
            version=version,
            current_profile=current_profile,
        )

    decision = projected["decisions"][0]
    assert projected_without_source == projected
    current_header = select_header_candidates(current_profile.sheets[0].header_candidates)[0]
    assert decision["header_candidate_id"] == current_header.id
    assert decision["region_candidate_id"] == current_header.region_id
    assert decision["data_start_row"] == 2
    assert decision["data_end_row"] == 5
    assert decision["data_start_column"] == 1
    assert decision["data_end_column"] == 1
    assert decision["materialize"] is False
    snapshot_decision = version.source_metadata["layout_projection_snapshot"]["decisions"][0]
    assert [column["offset"] for column in snapshot_decision["source_columns"]] == [0, 1]


def test_native_four_layer_match_and_import_requires_no_legacy_template(
    tmp_path: Path,
) -> None:
    source, profile = create_profile(tmp_path)
    source_region = profile_regions(profile)[0]
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        item = add_item(database, source, profile)
        region = RegionTemplate(
            code="region.native.people",
            published_version=1,
        )
        region_version = RegionTemplateVersion(
            version=1,
            name="人员表",
            status=TemplateStatus.PUBLISHED,
            domain="population",
            record_type="person",
            record_grain="one_row_per_person",
            region_kind="table",
            region_fingerprint=source_region.fingerprint,
            header_signature=[column.header_path for column in source_region.header.columns],
            layout_rules={
                "data_start_offset_from_header_end": 1,
                "data_end_gap_from_region_end": 0,
                "excluded_row_offsets": [],
                "classification": "table",
                "materialize": True,
            },
            field_bindings=[
                {
                    "source_column_id": column.source_column_id,
                    "header_path": column.header_path,
                    "semantic_field_code": f"native.field_{index}",
                    "semantic_field_version": 1,
                    "required": False,
                }
                for index, column in enumerate(
                    source_region.header.columns,
                    start=1,
                )
            ],
            source="manual",
        )
        for index, column in enumerate(source_region.header.columns, start=1):
            field = SemanticField(
                code=f"native.field_{index}",
                published_version=1,
            )
            field.versions.append(
                SemanticFieldVersion(
                    version=1,
                    name=column.header_path[-1],
                    layer="domain",
                    data_type="text",
                    status=TemplateStatus.PUBLISHED,
                )
            )
            database.add(field)
        region.versions.append(region_version)
        database.add(region)
        database.flush()

        composition = SheetComposition(
            code="sheet.native.people",
            published_version=1,
        )
        composition_version = SheetCompositionVersion(
            version=1,
            name="人员 Sheet",
            status=TemplateStatus.PUBLISHED,
            composition_fingerprint=hashlib.sha256(b"sheet-native").hexdigest(),
        )
        composition_version.region_slots.append(
            SheetCompositionRegionSlot(
                slot_key="people",
                region_template_id=region.id,
                region_template_version=1,
                ordinal=0,
            )
        )
        composition.versions.append(composition_version)
        database.add(composition)
        database.flush()

        route = WorkbookRoute(
            code="workbook.native.people",
            published_version=1,
        )
        route_version = WorkbookRouteVersion(
            version=1,
            name="人员工作簿",
            status=TemplateStatus.PUBLISHED,
            route_fingerprint=hashlib.sha256(b"workbook-native").hexdigest(),
        )
        route_version.sheet_slots.append(
            WorkbookRouteSheetSlot(
                slot_key="people_sheet",
                sheet_composition_id=composition.id,
                sheet_composition_version=1,
                ordinal=0,
            )
        )
        route.versions.append(route_version)
        database.add(route)
        database.flush()

        summary = match_profile(database, item_id=item.id, profile=profile)
        region_match = database.scalar(
            select(RegionTemplateMatch).where(RegionTemplateMatch.item_id == item.id)
        )
        sheet_match = database.scalar(
            select(SheetCompositionMatch).where(SheetCompositionMatch.item_id == item.id)
        )
        route_match = database.get(WorkbookRouteMatch, item.id)
        assert summary.match_type == MatchType.EXACT
        assert region_match is not None
        assert region_match.region_template_id == region.id
        assert region_match.template_id is None
        assert sheet_match is not None
        assert sheet_match.match_type == MatchType.EXACT
        assert route_match is not None
        assert route_match.match_type == MatchType.EXACT

        plan = approve_matched_region_plan(database, item=item)
        assert plan.template_id is None
        assert plan.primary_region_template_id == region.id
        execution = materialize_plan(database, plan.id)
        record = database.scalar(
            select(DatasetRecord).where(DatasetRecord.approved_plan_id == plan.id)
        )
        assert execution.record_count == 1
        assert record is not None
        assert record.template_id is None
        assert record.region_template_id == region.id
        assert record.raw_data["columns"]
    Base.metadata.drop_all(engine)


def test_region_start_projection_supports_v4_generated_layout_rules() -> None:
    assert project_region_data_rows(
        region_start=2,
        region_end=15,
        header_end=4,
        projection={
            "data_start_offset_from_region_start": 3,
            "data_end_gap_from_region_end": 4,
        },
    ) == (5, 11)


def test_reused_region_column_prefers_stable_source_id_over_header_text() -> None:
    class Column:
        def __init__(
            self,
            source_column_id: str,
            header_path: list[str],
        ) -> None:
            self.source_column_id = source_column_id
            self.header_path = header_path

    columns = [
        Column("workbook:stable:sheet:2:region:0:column:1", ["汇总表", "填报单位"]),
        Column("workbook:stable:sheet:2:region:0:column:2", ["汇总表"]),
    ]
    bindings = [
        {
            "source_column_id": "workbook:stable:sheet:2:region:0:column:1",
            "header_path": ["序号"],
        },
        {
            "source_column_id": "workbook:stable:sheet:2:region:0:column:2",
            "header_path": ["乡镇"],
        },
    ]

    resolved = resolve_reused_region_column(
        binding=bindings[0],
        binding_index=0,
        bindings=bindings,
        current_columns=columns,
        projected_columns=[],
    )

    assert resolved is columns[0]


def test_reused_region_column_rejects_physical_ordinal_with_different_header() -> None:
    class Column:
        def __init__(
            self,
            column: int,
            source_column_id: str,
            header_path: list[str],
        ) -> None:
            self.column = column
            self.source_column_id = source_column_id
            self.header_path = header_path

    current = Column(
        3,
        "workbook:new-format:sheet:0:region:0:column:3",
        ["工作表标题"],
    )
    binding = {
        "source_column_id": "workbook:template:sheet:0:region:0:column:3",
        "header_path": ["序号"],
    }

    assert (
        resolve_reused_region_column(
            binding=binding,
            binding_index=0,
            bindings=[binding, {"source_column_id": "other"}],
            current_columns=[current],
            projected_columns=[],
        )
        is None
    )


def test_reused_region_column_accepts_same_leaf_under_new_parent_header() -> None:
    class Column:
        def __init__(
            self,
            column: int,
            source_column_id: str,
            header_path: list[str],
        ) -> None:
            self.column = column
            self.source_column_id = source_column_id
            self.header_path = header_path

    current = Column(
        1,
        "workbook:new-format:sheet:0:region:0:column:1",
        ["户信息", "序号"],
    )
    binding = {
        "source_column_id": "workbook:template:sheet:0:region:0:column:1",
        "header_path": ["序号"],
    }

    assert (
        resolve_reused_region_column(
            binding=binding,
            binding_index=0,
            bindings=[binding, {"source_column_id": "unmapped-sibling"}],
            current_columns=[current],
            projected_columns=[],
        )
        is current
    )


def test_verified_source_region_can_reuse_approved_physical_selector(
    tmp_path: Path,
) -> None:
    _, profile = create_profile(tmp_path)
    source_region = profile_regions(profile)[0]
    column = source_region.header.columns[0]
    binding = {
        "source_column_id": "representative-sheet-column",
        "header_path": ["代表工作表表头"],
        "semantic_field_code": "test.person_name",
        "semantic_field_version": 1,
        "source_selector": {
            "kind": "physical_column",
            "column_offset": 0,
            "header_path_sha256": "not-the-current-header-hash",
        },
    }
    version = RegionTemplateVersion(
        version=1,
        name="已批准源证据选择器",
        status=TemplateStatus.PUBLISHED,
        domain="population",
        record_type="person",
        record_grain="one_row_per_person",
        region_kind="table",
        region_fingerprint=source_region.fingerprint,
        field_bindings=[binding],
    )

    assert (
        _region_template_binding(
            version,
            list(column.header_path),
            source_region=source_region,
            column=column,
            allow_ordinal=False,
        )
        is None
    )
    assert (
        _region_template_binding(
            version,
            list(column.header_path),
            source_region=source_region,
            column=column,
            allow_ordinal=False,
            verified_source_region=True,
        )
        is binding
    )


@pytest.mark.parametrize(
    "layout_mode",
    ["headerless_table", "explicit_header_table", "form"],
)
def test_exact_region_uses_selector_bindings_for_exact_field_matches(
    tmp_path: Path,
    layout_mode: str,
) -> None:
    source, profile = create_profile(tmp_path)
    source_region = profile_regions(profile)[0]
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as database:
        for code, name, data_type in (
            ("test.person_name", "人员姓名", "text"),
            ("test.person_count", "人员数量", "integer"),
        ):
            field = SemanticField(code=code, published_version=1)
            field.versions.append(
                SemanticFieldVersion(
                    version=1,
                    name=name,
                    layer="domain",
                    data_type=data_type,
                    status=TemplateStatus.PUBLISHED,
                )
            )
            database.add(field)
        item = add_item(database, source, profile)
        region = RegionTemplate(code=f"region.selector.{layout_mode}", published_version=1)
        if layout_mode == "form":
            bindings = [
                {
                    "source_column_id": "form:r2:c1",
                    "header_path": ["人员姓名"],
                    "semantic_field_code": "test.person_name",
                    "semantic_field_version": 1,
                    "source_selector": {
                        "kind": "cell",
                        "row_offset": 1,
                        "column_offset": 0,
                    },
                },
                {
                    "source_column_id": "form:r2:c2",
                    "header_path": ["人员数量"],
                    "semantic_field_code": "test.person_count",
                    "semantic_field_version": 1,
                    "source_selector": {
                        "kind": "cell",
                        "row_offset": 1,
                        "column_offset": 1,
                    },
                },
            ]
        else:
            bindings = [
                {
                    "source_column_id": "physical-column:1",
                    "header_path": ["不同的姓名表头"],
                    "semantic_field_code": "test.person_name",
                    "semantic_field_version": 1,
                    "source_selector": {
                        "kind": "physical_column",
                        "column_offset": 0,
                    },
                },
                {
                    "source_column_id": "physical-column:2",
                    "header_path": ["不同的数量表头"],
                    "semantic_field_code": "test.person_count",
                    "semantic_field_version": 1,
                    "source_selector": {
                        "kind": "physical_column",
                        "column_offset": 1,
                    },
                },
            ]
        region.versions.append(
            RegionTemplateVersion(
                version=1,
                name="选择器字段匹配",
                status=TemplateStatus.PUBLISHED,
                domain="population",
                record_type="person",
                record_grain="one_row_per_person",
                region_kind="form" if layout_mode == "form" else "table",
                region_fingerprint=source_region.fingerprint,
                header_signature=[column.header_path for column in source_region.header.columns],
                layout_rules={
                    "layout_mode": layout_mode,
                    "data_start_offset_from_region_start": 1,
                    "data_end_gap_from_region_end": 0,
                    "materialize": True,
                },
                field_bindings=bindings,
                source="codex",
            )
        )
        database.add(region)
        database.flush()

        summary = match_profile(database, item_id=item.id, profile=profile)
        matches = list(
            database.scalars(
                select(FieldMatch)
                .where(FieldMatch.item_id == item.id)
                .order_by(FieldMatch.source_column_id)
            )
        )

        assert summary.match_type == MatchType.EXACT
        assert len(matches) == 2
        assert all(match.match_type == MatchType.EXACT for match in matches)
        assert all(not match.requires_hermes for match in matches)
        assert {match.semantic_field_code for match in matches} == {
            "test.person_name",
            "test.person_count",
        }
        if layout_mode == "explicit_header_table":
            plan = approve_matched_region_plan(database, item=item)
            execution = materialize_plan(database, plan.id)
            assert execution.record_count == 1
    Base.metadata.drop_all(engine)
