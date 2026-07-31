from __future__ import annotations

from collections import deque

from openpyxl.utils import get_column_letter, range_boundaries

from village_insight.parsing.contracts import (
    Bounds,
    CellEvidence,
    HeaderCandidate,
    HeaderColumnCandidate,
    MergeEvidence,
    RegionCandidate,
)


def _style_similarity(
    cell_by_position: dict[tuple[int, int], CellEvidence],
    *,
    left_row: int,
    right_rows: range,
    min_column: int,
    max_column: int,
) -> float | None:
    comparisons: list[bool] = []
    for right_row in right_rows:
        for column in range(min_column, max_column + 1):
            left = cell_by_position.get((left_row, column))
            right = cell_by_position.get((right_row, column))
            if (
                left is None
                or right is None
                or left.display_value in {None, ""}
                or right.display_value in {None, ""}
                or left.style_ref is None
                or right.style_ref is None
            ):
                continue
            comparisons.append(left.style_ref == right.style_ref)
    if not comparisons:
        return None
    return sum(comparisons) / len(comparisons)


def make_bounds(min_row: int, min_column: int, max_row: int, max_column: int) -> Bounds:
    start = f"{get_column_letter(min_column)}{min_row}"
    end = f"{get_column_letter(max_column)}{max_row}"
    return Bounds(
        min_row=min_row,
        min_column=min_column,
        max_row=max_row,
        max_column=max_column,
        range=f"{start}:{end}",
    )


def observed_bounds(cells: list[CellEvidence]) -> Bounds | None:
    if not cells:
        return None
    return make_bounds(
        min(cell.row for cell in cells),
        min(cell.column for cell in cells),
        max(cell.row for cell in cells),
        max(cell.column for cell in cells),
    )


def build_region_candidates(
    parent_sheet_id: str,
    cells: list[CellEvidence],
    merges: list[MergeEvidence],
) -> list[RegionCandidate]:
    cell_by_position = {(cell.row, cell.column): cell for cell in cells}
    occupied = set(cell_by_position)
    for merge in merges:
        min_column, min_row, max_column, max_row = range_boundaries(merge.range)
        if (max_row - min_row + 1) * (max_column - min_column + 1) > 10_000:
            continue
        occupied.update(
            (row, column)
            for row in range(min_row, max_row + 1)
            for column in range(min_column, max_column + 1)
        )

    components: list[set[tuple[int, int]]] = []
    remaining = set(occupied)
    while remaining:
        seed = remaining.pop()
        component = {seed}
        queue = deque([seed])
        while queue:
            row, column = queue.popleft()
            for neighbor in (
                (row - 1, column),
                (row + 1, column),
                (row, column - 1),
                (row, column + 1),
            ):
                if neighbor in remaining:
                    remaining.remove(neighbor)
                    component.add(neighbor)
                    queue.append(neighbor)
        components.append(component)

    # A sparse column inside an otherwise continuous table can be disconnected
    # from the table by blank cells even though it is physically contained by
    # the table's rectangular bounds. Treat those components as evidence of the
    # containing region instead of presenting overlapping duplicate regions to
    # downstream matching and Hermes.
    component_bounds = {
        id(component): (
            min(row for row, _ in component),
            min(column for _, column in component),
            max(row for row, _ in component),
            max(column for _, column in component),
        )
        for component in components
    }
    absorbed: set[int] = set()
    consolidated: list[set[tuple[int, int]]] = []
    for component in sorted(
        components,
        key=lambda item: (
            -(
                (component_bounds[id(item)][2] - component_bounds[id(item)][0] + 1)
                * (
                    component_bounds[id(item)][3]
                    - component_bounds[id(item)][1]
                    + 1
                )
            ),
            component_bounds[id(item)][0],
            component_bounds[id(item)][1],
        ),
    ):
        component_key = id(component)
        if component_key in absorbed:
            continue
        min_row, min_column, max_row, max_column = component_bounds[component_key]
        merged_component = set(component)
        for child in components:
            child_key = id(child)
            if child_key == component_key or child_key in absorbed:
                continue
            (
                child_min_row,
                child_min_column,
                child_max_row,
                child_max_column,
            ) = component_bounds[child_key]
            if (
                min_row <= child_min_row <= child_max_row <= max_row
                and min_column <= child_min_column <= child_max_column <= max_column
            ):
                merged_component.update(child)
                absorbed.add(child_key)
        consolidated.append(merged_component)

    regions: list[RegionCandidate] = []
    ordered = sorted(
        consolidated,
        key=lambda component: (
            min(row for row, _ in component),
            min(column for _, column in component),
        ),
    )
    for index, component in enumerate(ordered):
        min_row = min(row for row, _ in component)
        max_row = max(row for row, _ in component)
        min_column = min(column for _, column in component)
        max_column = max(column for _, column in component)
        area = (max_row - min_row + 1) * (max_column - min_column + 1)
        evidence = [
            cell.id for position, cell in sorted(cell_by_position.items()) if position in component
        ]
        density = len(component) / area
        regions.append(
            RegionCandidate(
                id=f"{parent_sheet_id}:region:{index}",
                bounds=make_bounds(min_row, min_column, max_row, max_column),
                nonempty_cell_ids=evidence,
                density=density,
                confidence=min(0.95, 0.45 + density * 0.5),
                source="connected-nonempty-cells-contained/v2",
            )
        )
    return regions


def build_header_candidates(
    parent_sheet_id: str,
    cells: list[CellEvidence],
    merges: list[MergeEvidence],
    regions: list[RegionCandidate],
) -> list[HeaderCandidate]:
    cell_by_position = {(cell.row, cell.column): cell for cell in cells}
    merge_by_position: dict[tuple[int, int], MergeEvidence] = {}
    for merge in merges:
        min_column, min_row, max_column, max_row = range_boundaries(merge.range)
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                merge_by_position[(row, column)] = merge

    candidates: list[HeaderCandidate] = []
    for region in regions:
        bounds = region.bounds
        region_height = bounds.max_row - bounds.min_row + 1
        if region_height < 2:
            continue
        latest_start = min(bounds.max_row - 1, bounds.min_row + 7)
        for start_row in range(bounds.min_row, latest_start + 1):
            max_depth = min(3, bounds.max_row - start_row)
            for depth in range(1, max_depth + 1):
                header_rows = list(range(start_row, start_row + depth))
                columns: list[HeaderColumnCandidate] = []
                populated_paths = 0
                leaf_labels: list[str] = []
                text_labels = 0
                for column in range(bounds.min_column, bounds.max_column + 1):
                    path: list[str] = []
                    evidence_ids: list[str] = []
                    for row in header_rows:
                        cell = cell_by_position.get((row, column))
                        value = cell.display_value if cell is not None else None
                        evidence_id = cell.id if cell is not None else None
                        if cell is None:
                            mapped_merge = merge_by_position.get((row, column))
                            if mapped_merge is not None:
                                value = mapped_merge.anchor_value
                                evidence_id = mapped_merge.anchor_cell_id
                        label = str(value).strip() if value is not None else ""
                        if label and (not path or path[-1] != label):
                            path.append(label)
                        if evidence_id and evidence_id not in evidence_ids:
                            evidence_ids.append(evidence_id)
                    if path:
                        populated_paths += 1
                        leaf_labels.append(path[-1])
                        if any(character.isalpha() for character in path[-1]):
                            text_labels += 1
                    columns.append(
                        HeaderColumnCandidate(
                            column=column,
                            source_column_id=f"{region.id}:column:{column}",
                            header_path=path,
                            evidence_cell_ids=evidence_ids,
                        )
                    )
                width = bounds.max_column - bounds.min_column + 1
                coverage = populated_paths / width
                uniqueness = len(set(leaf_labels)) / len(leaf_labels) if leaf_labels else 0
                text_ratio = text_labels / len(leaf_labels) if leaf_labels else 0
                nonleaf_labels = [label for column in columns for label in column.header_path[:-1]]
                prefix_dominance = (
                    max(nonleaf_labels.count(label) for label in set(nonleaf_labels))
                    / len(leaf_labels)
                    if nonleaf_labels and leaf_labels
                    else 0
                )
                hierarchy_ratio = (
                    sum(max(0, len(column.header_path) - 1) for column in columns)
                    / len(leaf_labels)
                    / max(1, depth - 1)
                    if leaf_labels and depth > 1
                    else 0
                )
                next_row = start_row + depth
                next_values = [
                    cell_by_position[(next_row, column)].display_value
                    for column in range(bounds.min_column, bounds.max_column + 1)
                    if (next_row, column) in cell_by_position
                    and cell_by_position[(next_row, column)].display_value not in {None, ""}
                ]
                numeric_next = 0
                for value in next_values:
                    try:
                        float(str(value).replace(",", ""))
                    except ValueError:
                        continue
                    numeric_next += 1
                data_contrast = numeric_next / len(next_values) if next_values else 0
                body_rows = range(
                    next_row,
                    min(bounds.max_row, next_row + 2) + 1,
                )
                header_body_style_similarity = _style_similarity(
                    cell_by_position,
                    left_row=header_rows[-1],
                    right_rows=body_rows,
                    min_column=bounds.min_column,
                    max_column=bounds.max_column,
                )
                style_contrast = (
                    1 - header_body_style_similarity
                    if header_body_style_similarity is not None
                    else 0
                )
                added_row_body_similarity = 0.0
                if depth > 1:
                    added_similarity = _style_similarity(
                        cell_by_position,
                        left_row=header_rows[-1],
                        right_rows=body_rows,
                        min_column=bounds.min_column,
                        max_column=bounds.max_column,
                    )
                    if added_similarity is not None:
                        added_row_body_similarity = added_similarity
                confidence = min(
                    0.98,
                    0.2
                    + coverage * 0.25
                    + uniqueness * 0.35
                    + text_ratio * 0.18
                    + hierarchy_ratio * (1 - min(1, prefix_dominance)) * 0.18
                    + data_contrast * 0.14
                    + style_contrast * 0.16
                    - (1 - text_ratio) * (1 - data_contrast) * 0.35
                    - (depth - 1) * 0.03
                    - added_row_body_similarity * 0.3,
                )
                candidates.append(
                    HeaderCandidate(
                        id=f"{parent_sheet_id}:header:{len(candidates)}",
                        region_id=region.id,
                        header_rows=header_rows,
                        columns=columns,
                        confidence=max(0.05, confidence),
                        source="sliding-header-paths/v2",
                    )
                )
    return candidates


def select_header_candidates(
    candidates: list[HeaderCandidate],
    *,
    per_region: int = 1,
    min_confidence: float = 0.7,
) -> list[HeaderCandidate]:
    selected: list[HeaderCandidate] = []
    region_ids = dict.fromkeys(candidate.region_id for candidate in candidates)
    for region_id in region_ids:
        region_candidates = [
            candidate
            for candidate in candidates
            if candidate.region_id == region_id and candidate.confidence >= min_confidence
        ]
        ranked = sorted(
            region_candidates,
            key=lambda candidate: (
                -candidate.confidence,
                candidate.header_rows[0],
                len(candidate.header_rows),
            ),
        )
        if per_region == 1 and ranked:
            top = ranked[0]
            top_leaves = tuple(
                column.header_path[-1] if column.header_path else "" for column in top.columns
            )
            richer: list[HeaderCandidate] = [top]
            for candidate in ranked:
                if candidate.id == top.id:
                    continue
                leaves = tuple(
                    column.header_path[-1] if column.header_path else ""
                    for column in candidate.columns
                )
                nonleaf = [
                    label for column in candidate.columns for label in column.header_path[:-1]
                ]
                populated = sum(bool(column.header_path) for column in candidate.columns)
                dominance = (
                    max(nonleaf.count(label) for label in set(nonleaf)) / populated
                    if nonleaf and populated
                    else 0
                )
                if (
                    candidate.confidence >= top.confidence - 0.05
                    and leaves == top_leaves
                    and dominance <= 0.6
                ):
                    richer.append(candidate)
            selected.append(
                max(
                    richer,
                    key=lambda candidate: (
                        sum(len(column.header_path) for column in candidate.columns),
                        candidate.confidence,
                    ),
                    default=top,
                )
            )
        else:
            selected.extend(ranked[:per_region])
    return selected
