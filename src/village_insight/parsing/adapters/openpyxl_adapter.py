from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from itertools import zip_longest
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

import openpyxl
from openpyxl.utils import range_boundaries

from village_insight.parsing.candidates import (
    build_header_candidates,
    build_region_candidates,
    make_bounds,
    observed_bounds,
)
from village_insight.parsing.contracts import (
    CellEvidence,
    ColumnProperties,
    DetectionResult,
    DocumentFormat,
    MergeEvidence,
    RowProperties,
    SheetProfile,
    WorkbookProfile,
)
from village_insight.parsing.identity import cell_id, file_sha256, sheet_id, workbook_id
from village_insight.parsing.ooxml_repair import repair_safe_ooxml_packaging


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


@dataclass
class SheetMetadata:
    declared_bounds: tuple[int, int, int, int] | None
    merge_ranges: list[str]
    row_properties: list[RowProperties]
    column_properties: list[ColumnProperties]
    hidden_rows: set[int]
    hidden_column_ranges: list[tuple[int, int]]
    warnings: list[str]


def _sheet_metadata(payload: bytes, worksheet_path: str) -> SheetMetadata:
    declared_bounds = None
    merge_ranges: list[str] = []
    row_properties: list[RowProperties] = []
    column_properties: list[ColumnProperties] = []
    hidden_rows: set[int] = set()
    hidden_column_ranges: list[tuple[int, int]] = []
    warnings: list[str] = []
    with ZipFile(BytesIO(payload)) as archive:
        with archive.open(worksheet_path) as source:
            for event, element in ElementTree.iterparse(source, events=("start", "end")):
                if event == "end":
                    element.clear()
                    continue
                tag = element.tag.rsplit("}", 1)[-1]
                if tag == "dimension" and element.get("ref"):
                    declared_bounds = range_boundaries(str(element.get("ref")))
                elif tag == "mergeCell" and element.get("ref"):
                    merge_ranges.append(str(element.get("ref")))
                elif tag == "row" and element.get("r"):
                    row = int(str(element.get("r")))
                    hidden = element.get("hidden") in {"1", "true"}
                    height = (
                        float(str(element.get("ht"))) if element.get("ht") is not None else None
                    )
                    if hidden:
                        hidden_rows.add(row)
                    if hidden or height is not None:
                        row_properties.append(RowProperties(row=row, hidden=hidden, height=height))
                elif tag == "col" and element.get("min") and element.get("max"):
                    minimum = int(str(element.get("min")))
                    maximum = int(str(element.get("max")))
                    hidden = element.get("hidden") in {"1", "true"}
                    width = (
                        float(str(element.get("width")))
                        if element.get("width") is not None
                        else None
                    )
                    if hidden:
                        hidden_column_ranges.append((minimum, maximum))
                    if hidden or width is not None:
                        if maximum - minimum > 4096:
                            warnings.append(
                                "COLUMN_PROPERTY_RANGE_OMITTED: an oversized style-only "
                                "column range was excluded from per-column evidence"
                            )
                            continue
                        column_properties.extend(
                            ColumnProperties(column=column, hidden=hidden, width=width)
                            for column in range(minimum, maximum + 1)
                        )
    return SheetMetadata(
        declared_bounds=declared_bounds,
        merge_ranges=merge_ranges,
        row_properties=row_properties,
        column_properties=column_properties,
        hidden_rows=hidden_rows,
        hidden_column_ranges=hidden_column_ranges,
        warnings=list(dict.fromkeys(warnings)),
    )


def _column_hidden(column: int, ranges: list[tuple[int, int]]) -> bool:
    return any(minimum <= column <= maximum for minimum, maximum in ranges)


class OpenPyxlAdapter:
    name = "openpyxl-native"

    def supports(self, document_format: DocumentFormat) -> bool:
        return document_format == "xlsx"

    def profile(self, path: Path, detection: DetectionResult) -> WorkbookProfile:
        source_sha256 = file_sha256(path)
        parent_workbook_id = workbook_id(source_sha256)
        payload, repair_warnings = repair_safe_ooxml_packaging(path.read_bytes())
        formulas = openpyxl.load_workbook(
            BytesIO(payload),
            data_only=False,
            read_only=True,
            keep_links=False,
        )
        values = openpyxl.load_workbook(
            BytesIO(payload),
            data_only=True,
            read_only=True,
            keep_links=False,
        )
        try:
            sheets: list[SheetProfile] = []
            for index, worksheet in enumerate(formulas.worksheets):
                parent_sheet_id = sheet_id(parent_workbook_id, index)
                value_sheet = values[worksheet.title]
                metadata = _sheet_metadata(payload, worksheet._worksheet_path)
                worksheet.reset_dimensions()
                value_sheet.reset_dimensions()
                cells: list[CellEvidence] = []
                for formula_row, value_row in zip_longest(
                    worksheet.iter_rows(),
                    value_sheet.iter_rows(),
                    fillvalue=(),
                ):
                    cached_by_coordinate = {
                        cell.coordinate: cell.value
                        for cell in value_row
                        if hasattr(cell, "coordinate")
                    }
                    for cell in formula_row:
                        if not hasattr(cell, "coordinate"):
                            continue
                        cached = cached_by_coordinate.get(cell.coordinate)
                        if cell.value is None and cached is None:
                            continue
                        is_formula = cell.data_type == "f"
                        raw_value = json_value(cell.value)
                        display_value = json_value(cached if is_formula else cell.value)
                        row_hidden = cell.row in metadata.hidden_rows
                        column_hidden = _column_hidden(
                            cell.column,
                            metadata.hidden_column_ranges,
                        )
                        cells.append(
                            CellEvidence(
                                id=cell_id(parent_sheet_id, cell.row, cell.column),
                                coordinate=cell.coordinate,
                                row=cell.row,
                                column=cell.column,
                                raw_value=raw_value,
                                display_value=display_value,
                                formula=str(cell.value) if is_formula else None,
                                data_type=str(cell.data_type),
                                number_format=cell.number_format,
                                style_ref=f"openpyxl-style:{cell._style_id}",
                                hidden=row_hidden or column_hidden,
                            )
                        )

                cell_by_coordinate = {cell.coordinate: cell for cell in cells}
                merges = [
                    MergeEvidence(
                        id=f"{parent_sheet_id}:merge:{merge_index}",
                        range=merge_range,
                        anchor_cell_id=cell_id(
                            parent_sheet_id,
                            range_boundaries(merge_range)[1],
                            range_boundaries(merge_range)[0],
                        ),
                        anchor=(
                            f"{openpyxl.utils.get_column_letter(range_boundaries(merge_range)[0])}"
                            f"{range_boundaries(merge_range)[1]}"
                        ),
                        anchor_value=(
                            cell_by_coordinate[
                                f"{openpyxl.utils.get_column_letter(range_boundaries(merge_range)[0])}"
                                f"{range_boundaries(merge_range)[1]}"
                            ].display_value
                            if (
                                f"{openpyxl.utils.get_column_letter(range_boundaries(merge_range)[0])}"
                                f"{range_boundaries(merge_range)[1]}"
                            )
                            in cell_by_coordinate
                            else None
                        ),
                    )
                    for merge_index, merge_range in enumerate(sorted(metadata.merge_ranges))
                ]
                regions = build_region_candidates(parent_sheet_id, cells, merges)
                headers = build_header_candidates(
                    parent_sheet_id,
                    cells,
                    merges,
                    regions,
                )
                declared = None
                if metadata.declared_bounds is not None:
                    minimum_column, minimum_row, maximum_column, maximum_row = (
                        metadata.declared_bounds
                    )
                    declared = make_bounds(
                        minimum_row,
                        minimum_column,
                        maximum_row,
                        maximum_column,
                    )
                warnings = list(metadata.warnings)
                measured = observed_bounds(cells)
                if (
                    declared is not None
                    and measured is not None
                    and (
                        declared.max_row > measured.max_row * 10
                        or declared.max_column > measured.max_column * 10
                    )
                ):
                    warnings.append(
                        "DECLARED_BOUNDS_EXCEED_OBSERVED: source declaration was "
                        "preserved and was not used to crop evidence"
                    )
                sheets.append(
                    SheetProfile(
                        id=parent_sheet_id,
                        name=worksheet.title,
                        index=index,
                        hidden=worksheet.sheet_state != "visible",
                        declared_bounds=declared,
                        observed_bounds=measured,
                        cells=cells,
                        merges=merges,
                        row_properties=metadata.row_properties,
                        column_properties=metadata.column_properties,
                        region_candidates=regions,
                        header_candidates=headers,
                        warnings=warnings,
                    )
                )
            return WorkbookProfile(
                workbook_id=parent_workbook_id,
                source_sha256=source_sha256,
                parser_name=self.name,
                parser_version=f"{openpyxl.__version__}+layout-v3",
                file_name=path.name,
                detection=detection,
                sheets=sheets,
                warnings=[*detection.warnings, *repair_warnings],
            )
        finally:
            formulas.close()
            values.close()
