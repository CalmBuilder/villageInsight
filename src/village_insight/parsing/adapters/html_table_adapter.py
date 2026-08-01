from __future__ import annotations

from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path

from openpyxl.utils import get_column_letter

from village_insight.parsing.candidates import (
    build_header_candidates,
    build_region_candidates,
    observed_bounds,
)
from village_insight.parsing.contracts import (
    CellEvidence,
    DetectionResult,
    DocumentFormat,
    MergeEvidence,
    SheetProfile,
    WorkbookProfile,
)
from village_insight.parsing.identity import cell_id, file_sha256, sheet_id, workbook_id

_MAX_SPAN = 10_000


@dataclass
class _HtmlCell:
    value: str
    attributes: dict[str, str]


@dataclass
class _HtmlTable:
    rows: list[list[_HtmlCell]] = field(default_factory=list)


class _ExcelTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[_HtmlTable] = []
        self.sheet_names: list[str] = []
        self._table: _HtmlTable | None = None
        self._row: list[_HtmlCell] | None = None
        self._cell_attributes: dict[str, str] | None = None
        self._cell_text: list[str] = []
        self._capture_sheet_name = False
        self._sheet_name_text: list[str] = []

    def handle_starttag(
        self, tag: str, attrs: list[tuple[str, str | None]]
    ) -> None:
        tag = tag.lower()
        if tag == "x:name":
            self._capture_sheet_name = True
            self._sheet_name_text = []
        elif tag == "table" and self._table is None:
            self._table = _HtmlTable()
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell_attributes = {
                key.lower(): value or "" for key, value in attrs
            }
            self._cell_text = []

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "x:name" and self._capture_sheet_name:
            name = "".join(self._sheet_name_text).strip()
            if name:
                self.sheet_names.append(name)
            self._capture_sheet_name = False
        elif tag in {"td", "th"} and self._cell_attributes is not None:
            if self._row is not None:
                self._row.append(
                    _HtmlCell(
                        value="".join(self._cell_text).strip(),
                        attributes=self._cell_attributes,
                    )
                )
            self._cell_attributes = None
            self._cell_text = []
        elif tag == "tr" and self._row is not None:
            if self._table is not None:
                self._table.rows.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            self.tables.append(self._table)
            self._table = None

    def handle_data(self, data: str) -> None:
        if self._capture_sheet_name:
            self._sheet_name_text.append(data)
        if self._cell_attributes is not None:
            self._cell_text.append(data)


def _positive_span(attributes: dict[str, str], name: str) -> int:
    raw = attributes.get(name, "1")
    try:
        value = int(raw)
    except ValueError:
        return 1
    return value if 1 <= value <= _MAX_SPAN else 1


class ExcelHtmlAdapter:
    name = "stdlib-excel-html"
    version = "1"

    def supports(self, document_format: DocumentFormat) -> bool:
        return document_format == "excel_html"

    def profile(self, path: Path, detection: DetectionResult) -> WorkbookProfile:
        source_sha256 = file_sha256(path)
        parent_workbook_id = workbook_id(source_sha256)
        encoding = detection.signature.rsplit(":", maxsplit=1)[-1]
        parser = _ExcelTableParser()
        parser.feed(path.read_bytes().decode(encoding))
        parser.close()

        sheets: list[SheetProfile] = []
        for index, table in enumerate(parser.tables):
            parent_sheet_id = sheet_id(parent_workbook_id, index)
            cells: list[CellEvidence] = []
            merges: list[MergeEvidence] = []
            occupied: set[tuple[int, int]] = set()
            for row_index, row in enumerate(table.rows, start=1):
                column_index = 1
                for html_cell in row:
                    while (row_index, column_index) in occupied:
                        column_index += 1
                    row_span = _positive_span(html_cell.attributes, "rowspan")
                    column_span = _positive_span(html_cell.attributes, "colspan")
                    anchor_id = cell_id(parent_sheet_id, row_index, column_index)
                    style = html_cell.attributes.get("style", "").lower()
                    hidden = "display:none" in style.replace(" ", "")
                    if html_cell.value:
                        cells.append(
                            CellEvidence(
                                id=anchor_id,
                                coordinate=(
                                    f"{get_column_letter(column_index)}{row_index}"
                                ),
                                row=row_index,
                                column=column_index,
                                raw_value=html_cell.value,
                                display_value=html_cell.value,
                                data_type=(
                                    "number"
                                    if "x:num" in html_cell.attributes
                                    else "string"
                                ),
                                style_ref=html_cell.attributes.get("class"),
                                hidden=hidden,
                            )
                        )
                    if row_span > 1 or column_span > 1:
                        end_row = row_index + row_span - 1
                        end_column = column_index + column_span - 1
                        merge_range = (
                            f"{get_column_letter(column_index)}{row_index}:"
                            f"{get_column_letter(end_column)}{end_row}"
                        )
                        merges.append(
                            MergeEvidence(
                                id=f"{parent_sheet_id}:merge:{merge_range}",
                                range=merge_range,
                                anchor_cell_id=anchor_id,
                                anchor=f"{get_column_letter(column_index)}{row_index}",
                                anchor_value=html_cell.value or None,
                            )
                        )
                    occupied.update(
                        (spanned_row, spanned_column)
                        for spanned_row in range(row_index, row_index + row_span)
                        for spanned_column in range(
                            column_index, column_index + column_span
                        )
                    )
                    column_index += column_span
            regions = build_region_candidates(parent_sheet_id, cells, merges)
            headers = build_header_candidates(parent_sheet_id, cells, merges, regions)
            sheets.append(
                SheetProfile(
                    id=parent_sheet_id,
                    name=(
                        parser.sheet_names[index]
                        if index < len(parser.sheet_names)
                        else f"Table {index + 1}"
                    ),
                    index=index,
                    hidden=False,
                    declared_bounds=observed_bounds(cells),
                    observed_bounds=observed_bounds(cells),
                    cells=cells,
                    merges=merges,
                    row_properties=[],
                    column_properties=[],
                    region_candidates=regions,
                    header_candidates=headers,
                    warnings=[
                        "EXCEL_HTML_LIMITATION: formulas, external links, embedded "
                        "objects and scripts are intentionally not loaded"
                    ],
                )
            )
        return WorkbookProfile(
            workbook_id=parent_workbook_id,
            source_sha256=source_sha256,
            parser_name=self.name,
            parser_version=f"{self.version}+layout-v3",
            file_name=path.name,
            detection=detection,
            sheets=sheets,
            warnings=list(detection.warnings),
        )
