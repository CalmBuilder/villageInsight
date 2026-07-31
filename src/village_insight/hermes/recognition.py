from __future__ import annotations

import hashlib
import json
import re
import uuid
from importlib.metadata import PackageNotFoundError, version
from typing import Any, Literal

from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.datetime import from_excel
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from village_insight.db.models import (
    DocumentProfile,
    DocumentTemplate,
    FieldMatch,
    HermesRecognitionCache,
    HermesRecognitionRecord,
    IngestionItem,
    MatchType,
    ProposalStatus,
    SemanticField,
    SemanticFieldReviewEvent,
    SemanticFieldVersion,
    TemplateMatch,
    TemplateProposal,
    TemplateStatus,
    TemplateVersion,
    utcnow,
)
from village_insight.hermes.runtime import (
    HermesCallPolicy,
    HermesRuntime,
    HermesUnavailableError,
)
from village_insight.parsing.candidates import select_header_candidates
from village_insight.parsing.contracts import WorkbookProfile
from village_insight.parsing.profile_storage import load_workbook_profile
from village_insight.templates.contracts import (
    TemplateDefinition,
    TemplateFieldBinding,
)
from village_insight.templates.field_semantics import normalize_role_code
from village_insight.templates.field_variants import build_field_variant

PROMPT_VERSION = "template-diff/v15"
SCHEMA_VERSION = "template-diff-result/v6"
REVIEW_CONFIDENCE_THRESHOLD = 0.85
MAX_SAMPLE_ROWS_PER_REGION = 3
MAX_SAMPLE_CELLS = 24
MAX_MERGE_EVIDENCE = 16
MAX_FIELDS_PER_HERMES_CALL = 3
INITIAL_PREVIEW_ROWS = 10
INITIAL_PREVIEW_COLUMNS = 40
MAX_INITIAL_PREVIEW_CELLS = 800
MAX_RANGE_REQUESTS = 3
MAX_RANGE_ROWS = 30
MAX_RANGE_COLUMNS = 40
MAX_RANGE_EVIDENCE_CELLS = 800


class SemanticCandidateSummary(BaseModel):
    code: str
    version: int
    name: str
    data_type: str
    unit: str | None = None
    aliases: list[str] = Field(default_factory=list)
    compatible_roles: list[str] = Field(default_factory=list)
    score_basis_points: int = Field(ge=0, le=10_000)
    reasons: list[str] = Field(default_factory=list)


class HeaderEvidenceSummary(BaseModel):
    header_candidate_id: str
    region_candidate_id: str
    source_column_id: str
    header_path: list[str]
    evidence_cell_ids: list[str]
    observed_data_type: str | None = None
    context: dict[str, str | list[str] | None] = Field(default_factory=dict)
    semantic_candidates: list[SemanticCandidateSummary] = Field(default_factory=list)


class RegionEvidenceSummary(BaseModel):
    candidate_id: str
    sheet_id: str
    kind: str
    range: str
    columns: int
    density: float


class SourceSampleCell(BaseModel):
    source_column_id: str
    source_cell_id: str
    value_kind: str
    redacted_value: str | int | float | bool | None


class SourceSampleRow(BaseModel):
    region_candidate_id: str
    sheet_id: str
    source_row: int = Field(ge=1)
    cells: list[SourceSampleCell]


class SheetEvidenceSummary(BaseModel):
    sheet_id: str
    name: str
    observed_range: str | None
    rows: int
    columns: int
    merge_count: int = 0
    merge_ranges: list[str] = Field(default_factory=list)
    hidden_rows: list[int] = Field(default_factory=list)
    hidden_columns: list[int] = Field(default_factory=list)


class RangeEvidenceRow(BaseModel):
    row: int = Field(ge=1)
    cells: list[
        tuple[
            int,
            str | int | float | bool | None,
            str | None,
            str | None,
        ]
    ] = Field(default_factory=list)


class SheetRangeEvidence(BaseModel):
    evidence_id: str
    sheet_id: str
    range: str
    purpose: Literal["initial_preview", "body_sample", "requested"]
    rows: list[RangeEvidenceRow] = Field(default_factory=list)


class SheetRangeRequest(BaseModel):
    sheet_id: str
    start_row: int = Field(ge=1)
    end_row: int = Field(ge=1)
    start_column: int = Field(ge=1)
    end_column: int = Field(ge=1)
    reason: str = Field(min_length=1, max_length=200)


class RowRoleSegment(BaseModel):
    sheet_id: str
    start_row: int = Field(ge=1)
    end_row: int = Field(ge=1)
    role: Literal[
        "title",
        "context",
        "header_group",
        "header_leaf",
        "data",
        "summary",
        "note",
        "footer",
        "separator",
    ]
    evidence_ids: list[str] = Field(default_factory=list)


class SemanticFieldSummary(BaseModel):
    code: str
    name: str
    layer: str
    data_type: str
    unit: str | None = None
    aliases: list[str] = Field(default_factory=list)


class TemplateDiffRequest(BaseModel):
    contract_version: Literal["template-diff-request/v5"] = "template-diff-request/v5"
    profile_contract_version: str
    parser_name: str
    parser_version: str
    layout_fingerprint: str
    template_id: uuid.UUID | None
    template_version: int | None
    match_type: str
    new_headers: list[str]
    missing_headers: list[str]
    headers: list[HeaderEvidenceSummary]
    regions: list[RegionEvidenceSummary]
    merge_ids: list[str]
    source_samples: list[SourceSampleRow] = Field(default_factory=list)
    sheets: list[SheetEvidenceSummary] = Field(default_factory=list)
    range_evidence: list[SheetRangeEvidence] = Field(default_factory=list)
    semantic_catalog: list[SemanticFieldSummary] = Field(default_factory=list)
    unresolved_source_column_ids: list[str] = Field(default_factory=list)


class RecordGrainDecision(BaseModel):
    value: str = Field(min_length=1, max_length=120)
    confidence: float = Field(ge=0, le=1)
    evidence_ids: list[str] = Field(default_factory=list)


class TemplateSuggestion(BaseModel):
    template_code: str = Field(pattern=r"^[a-z][a-z0-9_.]{1,159}$")
    template_name: str = Field(min_length=1, max_length=200)
    domain: str = Field(pattern=r"^[a-z][a-z0-9_]{1,79}$")
    record_type: str = Field(pattern=r"^[a-z][a-z0-9_]{1,119}$")
    confidence: float = Field(ge=0, le=1)
    evidence_ids: list[str] = Field(default_factory=list)


class FieldDecision(BaseModel):
    source_column_id: str
    action: Literal[
        "REUSE_FIELD",
        "ADD_ALIAS",
        "PROPOSE_NEW_FIELD",
        "ROLE_VARIANT",
        "IGNORE_COLUMN",
        "SEMANTIC_CONFLICT",
        "AMBIGUOUS",
    ]
    semantic_field_code: str | None = None
    proposed_field_code: str | None = None
    layer: Literal["base", "domain"] | None = None
    data_type: (
        Literal[
            "text",
            "integer",
            "decimal",
            "boolean",
            "date",
            "datetime",
        ]
        | None
    ) = None
    unit: str | None = None
    role: str | None = None
    confidence: float = Field(ge=0, le=1)
    evidence_ids: list[str] = Field(default_factory=list)
    requires_review: bool = True


class MergeDecision(BaseModel):
    merge_id: str
    action: Literal["PROPAGATE", "TITLE_ONLY", "STRUCTURAL_GROUP", "IGNORE"]
    target_source_column_ids: list[str] = Field(default_factory=list)
    evidence_ids: list[str] = Field(default_factory=list)


class LayoutDecision(BaseModel):
    region_candidate_id: str
    header_candidate_id: str
    data_start_row: int = Field(ge=1)
    data_end_row: int = Field(ge=1)
    excluded_rows: list[int] = Field(default_factory=list)
    classification: Literal["table", "form", "matrix", "noise"]
    materialize: bool = True
    confidence: float = Field(ge=0, le=1)
    evidence_ids: list[str] = Field(default_factory=list)
    merge_decisions: list[MergeDecision] = Field(default_factory=list)


class CompactFieldDecision(BaseModel):
    source_column_id: str
    action: Literal[
        "REUSE_FIELD",
        "ADD_ALIAS",
        "PROPOSE_NEW_FIELD",
        "ROLE_VARIANT",
        "IGNORE_COLUMN",
        "SEMANTIC_CONFLICT",
        "AMBIGUOUS",
    ]
    semantic_field_code: str | None = None
    proposed_field_code: str | None = None
    layer: Literal["base", "domain"] | None = None
    data_type: (
        Literal[
            "text",
            "integer",
            "decimal",
            "boolean",
            "date",
            "datetime",
        ]
        | None
    ) = None
    unit: str | None = None
    role: str | None = None
    confidence: float = Field(ge=0, le=1)
    requires_review: bool = True


class CompactLayoutDecision(BaseModel):
    region_candidate_id: str
    header_candidate_id: str
    data_start_row: int = Field(ge=1)
    data_end_row: int = Field(ge=1)
    excluded_rows: list[int] = Field(default_factory=list)
    classification: Literal["table", "form", "matrix", "noise"]
    confidence: float = Field(ge=0, le=1)


class CompactTemplateSuggestion(BaseModel):
    template_code: str = Field(pattern=r"^[a-z][a-z0-9_.]{1,159}$")
    template_name: str = Field(min_length=1, max_length=200)
    domain: str = Field(pattern=r"^[a-z][a-z0-9_]{1,79}$")
    record_type: str = Field(pattern=r"^[a-z][a-z0-9_]{1,119}$")
    confidence: float = Field(ge=0, le=1)


class CompactRecordGrainDecision(BaseModel):
    value: str = Field(min_length=1, max_length=120)
    confidence: float = Field(ge=0, le=1)


class TemplateDiffChunkResult(BaseModel):
    template_suggestion: CompactTemplateSuggestion | None = None
    record_grain: CompactRecordGrainDecision | None = None
    layout_decisions: list[CompactLayoutDecision] = Field(default_factory=list)
    field_decisions: list[CompactFieldDecision] = Field(default_factory=list)


class WorkbookStructureDecision(BaseModel):
    contract_version: Literal["workbook-structure-decision/v1"] = "workbook-structure-decision/v1"
    row_role_segments: list[RowRoleSegment] = Field(default_factory=list)
    layout_decisions: list[LayoutDecision] = Field(default_factory=list)
    evidence_requests: list[SheetRangeRequest] = Field(default_factory=list)
    confidence: float = Field(ge=0, le=1)


class TemplateDiffResult(BaseModel):
    contract_version: Literal["template-diff-result/v2"] = "template-diff-result/v2"
    template_suggestion: TemplateSuggestion | None = None
    record_grain: RecordGrainDecision | None = None
    layout_decisions: list[LayoutDecision] = Field(default_factory=list)
    field_decisions: list[FieldDecision] = Field(default_factory=list)
    recognition_passes: int = Field(default=1, ge=1, le=2)
    requires_governance: bool = False
    governance_reason_codes: list[str] = Field(default_factory=list)
    minimum_confidence: float | None = Field(default=None, ge=0, le=1)
    evidence_requests: list[SheetRangeRequest] = Field(default_factory=list)
    structure_decision: WorkbookStructureDecision | None = None


class RecognitionValidationError(ValueError):
    pass


class ProposalResolutionError(ValueError):
    pass


_SAFE_CATEGORY_VALUES = {
    "是",
    "否",
    "有",
    "无",
    "男",
    "女",
    "正常",
    "已婚",
    "未婚",
    "党员",
    "群众",
}


def _redacted_sample_value(value: object) -> tuple[str, str | int | float | bool | None]:
    if value is None:
        return "empty", None
    if isinstance(value, bool):
        return "boolean", value
    if isinstance(value, int | float):
        text = str(value)
        if len(re.sub(r"\D", "", text)) >= 7:
            return "long_number", f"<NUMBER:{len(text)}>"
        return "number", value
    text = " ".join(str(value).split())
    if not text:
        return "empty", None
    if text in _SAFE_CATEGORY_VALUES:
        return "category", text
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 7:
        return "identifier_like", f"<DIGITS:{len(digits)}>"
    if re.fullmatch(r"\d{4}[-/.年]\d{1,2}(?:[-/.月]\d{1,2}日?)?", text):
        return "date_like", text
    character_kind = (
        "cn_text"
        if re.search(r"[\u4e00-\u9fff]", text)
        else "latin_text"
        if re.search(r"[A-Za-z]", text)
        else "text"
    )
    digest = hashlib.sha256(text.encode()).hexdigest()[:8]
    return character_kind, f"<{character_kind.upper()}:{len(text)}:{digest}>"


def _representative_rows(rows: list[int]) -> list[int]:
    if len(rows) <= MAX_SAMPLE_ROWS_PER_REGION:
        return rows
    return [rows[0], rows[len(rows) // 2], rows[-1]]


def _range_name(
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
) -> str:
    return f"{get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row}"


def _range_evidence(
    profile: WorkbookProfile,
    *,
    sheet_id: str,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
    purpose: Literal["initial_preview", "body_sample", "requested"],
) -> SheetRangeEvidence:
    sheet = next((entry for entry in profile.sheets if entry.id == sheet_id), None)
    if sheet is None:
        raise RecognitionValidationError(f"unknown sheet_id: {sheet_id}")
    if end_row < start_row or end_column < start_column:
        raise RecognitionValidationError("requested sheet range is inverted")
    if end_row - start_row + 1 > MAX_RANGE_ROWS:
        raise RecognitionValidationError("requested sheet range exceeds row limit")
    if end_column - start_column + 1 > MAX_RANGE_COLUMNS:
        raise RecognitionValidationError("requested sheet range exceeds column limit")
    bounds = sheet.observed_bounds
    if bounds is None:
        raise RecognitionValidationError("requested sheet has no observed cells")
    if (
        start_row < bounds.min_row
        or end_row > bounds.max_row
        or start_column < bounds.min_column
        or end_column > bounds.max_column
    ):
        raise RecognitionValidationError("requested sheet range exceeds observed bounds")
    rows: dict[
        int,
        list[
            tuple[
                int,
                str | int | float | bool | None,
                str | None,
                str | None,
            ]
        ],
    ] = {}
    for cell in sheet.cells:
        if not (start_row <= cell.row <= end_row and start_column <= cell.column <= end_column):
            continue
        _, redacted_value = _redacted_sample_value(cell.display_value)
        rows.setdefault(cell.row, []).append(
            (
                cell.column,
                redacted_value,
                cell.data_type,
                cell.style_ref,
            )
        )
    range_name = _range_name(start_row, end_row, start_column, end_column)
    return SheetRangeEvidence(
        evidence_id=f"{sheet_id}:range:{range_name}",
        sheet_id=sheet_id,
        range=range_name,
        purpose=purpose,
        rows=[RangeEvidenceRow(row=row, cells=cells) for row, cells in sorted(rows.items())],
    )


def _initial_sheet_evidence(
    profile: WorkbookProfile,
) -> tuple[list[SheetEvidenceSummary], list[SheetRangeEvidence]]:
    summaries: list[SheetEvidenceSummary] = []
    evidence: list[SheetRangeEvidence] = []
    for sheet in profile.sheets:
        bounds = sheet.observed_bounds
        summaries.append(
            SheetEvidenceSummary(
                sheet_id=sheet.id,
                name=sheet.name,
                observed_range=bounds.range if bounds is not None else None,
                rows=(bounds.max_row - bounds.min_row + 1 if bounds is not None else 0),
                columns=(bounds.max_column - bounds.min_column + 1 if bounds is not None else 0),
                merge_count=len(sheet.merges),
                merge_ranges=[merge.range for merge in sheet.merges[:40]],
                hidden_rows=[row.row for row in sheet.row_properties if row.hidden],
                hidden_columns=[
                    column.column for column in sheet.column_properties if column.hidden
                ],
            )
        )
        if bounds is None:
            continue
        end_column = min(
            bounds.max_column,
            bounds.min_column + INITIAL_PREVIEW_COLUMNS - 1,
        )
        preview_width = end_column - bounds.min_column + 1
        preview_rows = min(
            INITIAL_PREVIEW_ROWS,
            max(1, MAX_INITIAL_PREVIEW_CELLS // preview_width),
        )
        preview_end = min(
            bounds.max_row,
            bounds.min_row + preview_rows - 1,
        )
        evidence.append(
            _range_evidence(
                profile,
                sheet_id=sheet.id,
                start_row=bounds.min_row,
                end_row=preview_end,
                start_column=bounds.min_column,
                end_column=end_column,
                purpose="initial_preview",
            )
        )
        if bounds.max_row > preview_end:
            sample_rows = sorted(
                {
                    max(preview_end + 1, (bounds.min_row + bounds.max_row) // 2),
                    bounds.max_row,
                }
            )
            for row in sample_rows:
                evidence.append(
                    _range_evidence(
                        profile,
                        sheet_id=sheet.id,
                        start_row=row,
                        end_row=row,
                        start_column=bounds.min_column,
                        end_column=end_column,
                        purpose="body_sample",
                    )
                )
    return summaries, evidence


def fulfill_range_requests(
    profile: WorkbookProfile,
    requests: list[SheetRangeRequest],
) -> list[SheetRangeEvidence]:
    if len(requests) > MAX_RANGE_REQUESTS:
        raise RecognitionValidationError("Hermes requested too many sheet ranges")
    evidence: list[SheetRangeEvidence] = []
    cell_count = 0
    seen: set[tuple[str, int, int, int, int]] = set()
    for request in requests:
        key = (
            request.sheet_id,
            request.start_row,
            request.end_row,
            request.start_column,
            request.end_column,
        )
        if key in seen:
            continue
        seen.add(key)
        item = _range_evidence(
            profile,
            sheet_id=request.sheet_id,
            start_row=request.start_row,
            end_row=request.end_row,
            start_column=request.start_column,
            end_column=request.end_column,
            purpose="requested",
        )
        cell_count += sum(len(row.cells) for row in item.rows)
        if cell_count > MAX_RANGE_EVIDENCE_CELLS:
            raise RecognitionValidationError("Hermes requested too many sheet evidence cells")
        evidence.append(item)
    return evidence


def _request_evidence_ids(request: TemplateDiffRequest) -> set[str]:
    evidence_ids = set(request.merge_ids)
    for sheet in request.sheets:
        evidence_ids.add(sheet.sheet_id)
    for header in request.headers:
        evidence_ids.add(header.source_column_id)
        evidence_ids.add(header.header_candidate_id)
        evidence_ids.update(header.evidence_cell_ids)
    for region in request.regions:
        evidence_ids.add(region.candidate_id)
        evidence_ids.add(region.sheet_id)
    for sample in request.source_samples:
        evidence_ids.add(sample.region_candidate_id)
        evidence_ids.add(sample.sheet_id)
        evidence_ids.update(cell.source_cell_id for cell in sample.cells)
    for evidence in request.range_evidence:
        evidence_ids.add(evidence.evidence_id)
        evidence_ids.add(evidence.sheet_id)
    return evidence_ids


def validate_structure_decision(
    profile: WorkbookProfile,
    request: TemplateDiffRequest,
    result: WorkbookStructureDecision,
) -> None:
    evidence_ids = _request_evidence_ids(request)
    sheet_by_id = {sheet.id: sheet for sheet in profile.sheets}
    region_by_id = {region.candidate_id: region for region in request.regions}
    header_candidate_by_id = {
        candidate.id: candidate
        for sheet in profile.sheets
        for candidate in select_header_candidates(sheet.header_candidates)
    }
    header_region = {
        header.header_candidate_id: header.region_candidate_id for header in request.headers
    }
    layout_region_ids = [layout.region_candidate_id for layout in result.layout_decisions]
    if len(layout_region_ids) != len(set(layout_region_ids)):
        raise RecognitionValidationError("duplicate structure decision for region")
    if set(layout_region_ids) != set(region_by_id):
        raise RecognitionValidationError("structure decisions must cover supplied regions exactly")
    for layout in result.layout_decisions:
        region = region_by_id[layout.region_candidate_id]
        if header_region.get(layout.header_candidate_id) != layout.region_candidate_id:
            raise RecognitionValidationError(
                "structure header candidate does not belong to its region"
            )
        _, min_row, _, max_row = range_boundaries(region.range)
        selected_header = header_candidate_by_id.get(layout.header_candidate_id)
        if selected_header is None:
            raise RecognitionValidationError(
                "structure decision references unknown header candidate"
            )
        first_data_row = max(selected_header.header_rows) + 1
        if (
            layout.data_end_row < layout.data_start_row
            or layout.data_start_row < min_row
            or layout.data_end_row > max_row
        ):
            raise RecognitionValidationError("structure data range exceeds candidate region")
        if layout.materialize and layout.data_start_row < first_data_row:
            raise RecognitionValidationError(
                "structure data range includes the selected header row"
            )
        if layout.materialize and layout.classification == "noise":
            raise RecognitionValidationError("noise structure cannot be materialized")
        if not layout.materialize and layout.classification != "noise":
            raise RecognitionValidationError("ignored structure must be classified as noise")
        unknown = set(layout.evidence_ids) - evidence_ids
        if unknown:
            raise RecognitionValidationError(
                f"unknown structure evidence ids: {', '.join(sorted(unknown))}"
            )

    covered: dict[tuple[str, int], str] = {}
    for segment in result.row_role_segments:
        sheet = sheet_by_id.get(segment.sheet_id)
        if sheet is None or sheet.observed_bounds is None:
            raise RecognitionValidationError("row role references unknown sheet")
        if (
            segment.end_row < segment.start_row
            or segment.start_row < sheet.observed_bounds.min_row
            or segment.end_row > sheet.observed_bounds.max_row
        ):
            raise RecognitionValidationError("row role exceeds observed sheet bounds")
        unknown = set(segment.evidence_ids) - evidence_ids
        if unknown:
            raise RecognitionValidationError(
                f"unknown row role evidence ids: {', '.join(sorted(unknown))}"
            )
        for row in range(segment.start_row, segment.end_row + 1):
            key = (segment.sheet_id, row)
            if key in covered:
                raise RecognitionValidationError("row role segments overlap")
            covered[key] = segment.role

    if len(result.evidence_requests) > MAX_RANGE_REQUESTS:
        raise RecognitionValidationError("Hermes requested too many sheet ranges")


def normalize_ignored_structure_ranges(
    request: TemplateDiffRequest,
    result: WorkbookStructureDecision,
) -> tuple[WorkbookStructureDecision, bool]:
    """Canonicalize unused coordinates without weakening retained-region checks."""
    region_by_id = {region.candidate_id: region for region in request.regions}
    normalized: list[LayoutDecision] = []
    changed = False
    for layout in result.layout_decisions:
        region = region_by_id.get(layout.region_candidate_id)
        if region is None:
            normalized.append(layout)
            continue
        ignored = not layout.materialize or layout.classification == "noise"
        if not ignored:
            normalized.append(layout)
            continue
        _, min_row, _, max_row = range_boundaries(region.range)
        start_row = min(max(layout.data_start_row, min_row), max_row)
        end_row = min(max(layout.data_end_row, min_row), max_row)
        if end_row < start_row:
            end_row = start_row
        changed = changed or (
            layout.materialize
            or layout.classification != "noise"
            or start_row != layout.data_start_row
            or end_row != layout.data_end_row
        )
        normalized.append(
            layout.model_copy(
                update={
                    "materialize": False,
                    "classification": "noise",
                    "data_start_row": start_row,
                    "data_end_row": end_row,
                }
            )
        )
    return result.model_copy(update={"layout_decisions": normalized}), changed


def normalize_structure_data_ranges(
    request: TemplateDiffRequest,
    result: WorkbookStructureDecision,
) -> tuple[WorkbookStructureDecision, bool]:
    """Clamp retained layout rows to immutable Region candidate bounds."""
    region_by_id = {region.candidate_id: region for region in request.regions}
    normalized: list[LayoutDecision] = []
    changed = False
    for layout in result.layout_decisions:
        region = region_by_id.get(layout.region_candidate_id)
        if region is None:
            normalized.append(layout)
            continue
        _, min_row, _, max_row = range_boundaries(region.range)
        start_row = max(min_row, layout.data_start_row)
        end_row = min(max_row, layout.data_end_row)
        excluded_rows = [
            row for row in layout.excluded_rows if start_row <= row <= end_row
        ]
        changed = changed or (
            start_row != layout.data_start_row
            or end_row != layout.data_end_row
            or excluded_rows != layout.excluded_rows
        )
        normalized.append(
            layout.model_copy(
                update={
                    "data_start_row": start_row,
                    "data_end_row": end_row,
                    "excluded_rows": excluded_rows,
                }
            )
        )
    return result.model_copy(update={"layout_decisions": normalized}), changed


def normalize_structure_merge_references(
    profile: WorkbookProfile,
    request: TemplateDiffRequest,
    result: WorkbookStructureDecision,
) -> tuple[WorkbookStructureDecision, bool]:
    """Resolve unambiguous A1 ranges to supplied immutable merge evidence IDs."""
    valid_merge_ids = set(request.merge_ids)
    sheet_by_id = {sheet.id: sheet for sheet in profile.sheets}
    region_sheet = {region.candidate_id: region.sheet_id for region in request.regions}
    changed = False
    normalized_layouts: list[LayoutDecision] = []
    for layout in result.layout_decisions:
        sheet = sheet_by_id.get(region_sheet.get(layout.region_candidate_id, ""))
        merge_ids_by_range: dict[tuple[int, int, int, int], list[str]] = {}
        if sheet is not None:
            for physical_merge in sheet.merges:
                if physical_merge.id not in valid_merge_ids:
                    continue
                merge_ids_by_range.setdefault(
                    range_boundaries(physical_merge.range),
                    [],
                ).append(physical_merge.id)
        normalized_merges: list[MergeDecision] = []
        for decision_merge in layout.merge_decisions:
            if decision_merge.merge_id in valid_merge_ids:
                normalized_merges.append(decision_merge)
                continue
            try:
                coordinate = range_boundaries(decision_merge.merge_id)
            except ValueError:
                normalized_merges.append(decision_merge)
                continue
            candidates = merge_ids_by_range.get(coordinate, [])
            if len(candidates) != 1:
                normalized_merges.append(decision_merge)
                continue
            changed = True
            normalized_merges.append(decision_merge.model_copy(update={"merge_id": candidates[0]}))
        normalized_layouts.append(layout.model_copy(update={"merge_decisions": normalized_merges}))
    return result.model_copy(update={"layout_decisions": normalized_layouts}), changed


def apply_structure_decision(
    request: TemplateDiffRequest,
    result: WorkbookStructureDecision,
) -> TemplateDiffRequest:
    accepted_region_ids = {
        decision.region_candidate_id for decision in result.layout_decisions if decision.materialize
    }
    if not accepted_region_ids:
        # A workbook containing only an unfilled form or explanatory material is a
        # successful zero-record import, not a recognition failure. Preserve the
        # Regions and headers as immutable structure evidence, but schedule no
        # semantic field work.
        return request.model_copy(
            update={
                "new_headers": [],
                "missing_headers": [],
                "unresolved_source_column_ids": [],
                "source_samples": [],
            }
        )
    headers = [
        header for header in request.headers if header.region_candidate_id in accepted_region_ids
    ]
    retained_column_ids = {header.source_column_id for header in headers}
    retained_paths = {" / ".join(header.header_path) for header in headers}
    return request.model_copy(
        update={
            "new_headers": [header for header in request.new_headers if header in retained_paths],
            "unresolved_source_column_ids": [
                source_column_id
                for source_column_id in request.unresolved_source_column_ids
                if source_column_id in retained_column_ids
            ],
            "headers": headers,
            "regions": [
                region for region in request.regions if region.candidate_id in accepted_region_ids
            ],
            "source_samples": [
                sample.model_copy(
                    update={
                        "cells": [
                            cell
                            for cell in sample.cells
                            if cell.source_column_id in retained_column_ids
                        ]
                    }
                )
                for sample in request.source_samples
                if sample.region_candidate_id in accepted_region_ids
                and any(cell.source_column_id in retained_column_ids for cell in sample.cells)
            ],
        }
    )


def build_diff_request(
    profile: WorkbookProfile,
    match: TemplateMatch,
    *,
    semantic_catalog: list[SemanticFieldSummary] | None = None,
    field_matches: list[FieldMatch] | None = None,
) -> TemplateDiffRequest:
    unmatched = match.differences.get("unmatched_regions", [])
    unmatched_entries = (
        [entry for entry in unmatched if isinstance(entry, dict)]
        if isinstance(unmatched, list)
        else []
    )
    target_region_ids = {
        str(entry["region_id"]) for entry in unmatched_entries if entry.get("region_id")
    }
    new_headers = {
        str(header)
        for entry in unmatched_entries
        for header in (
            entry.get("differences", {}).get("new_headers", [])
            if isinstance(entry.get("differences"), dict)
            else []
        )
    }
    missing_headers = {
        str(header)
        for entry in unmatched_entries
        for header in (
            entry.get("differences", {}).get("missing_headers", [])
            if isinstance(entry.get("differences"), dict)
            else []
        )
    }
    if not target_region_ids:
        new_headers.update(str(item) for item in match.differences.get("new_headers", []))
        missing_headers.update(str(item) for item in match.differences.get("missing_headers", []))
    field_matches_by_source = {entry.source_column_id: entry for entry in field_matches or []}
    unresolved_field_matches = [
        entry for entry in field_matches or [] if entry.requires_hermes
    ]
    target_region_ids.update(entry.region_id for entry in unresolved_field_matches)
    new_headers.update(
        " / ".join(entry.header_path)
        for entry in unresolved_field_matches
        if entry.header_path
    )
    catalog_by_code = {entry.code: entry for entry in semantic_catalog or []}
    headers: list[HeaderEvidenceSummary] = []
    regions: list[RegionEvidenceSummary] = []
    merge_ids: list[str] = []
    source_samples: list[SourceSampleRow] = []
    sheet_summaries, range_evidence = _initial_sheet_evidence(profile)
    unresolved_source_column_ids = {
        entry.source_column_id for entry in unresolved_field_matches
    }
    for sheet in profile.sheets:
        cells_by_position = {(cell.row, cell.column): cell for cell in sheet.cells}
        selected_candidates = select_header_candidates(
            sheet.header_candidates,
            per_region=1,
        )
        if target_region_ids:
            selected_candidates = [
                candidate
                for candidate in selected_candidates
                if candidate.region_id in target_region_ids
            ]
        selected_region_ids = {candidate.region_id for candidate in selected_candidates}
        for candidate in selected_candidates:
            for column in candidate.columns:
                header_name = " / ".join(column.header_path)
                if not header_name:
                    continue
                field_match = field_matches_by_source.get(column.source_column_id)
                field_candidates: list[SemanticCandidateSummary] = []
                if field_match is not None:
                    raw_candidates = field_match.differences.get("candidates", [])
                    for raw_candidate in raw_candidates:
                        if not isinstance(raw_candidate, dict):
                            continue
                        code = str(raw_candidate.get("semantic_field_code") or "")
                        catalog_entry = catalog_by_code.get(code)
                        if catalog_entry is None:
                            continue
                        field_candidates.append(
                            SemanticCandidateSummary(
                                code=code,
                                version=int(raw_candidate.get("semantic_field_version") or 1),
                                name=catalog_entry.name,
                                data_type=catalog_entry.data_type,
                                unit=catalog_entry.unit,
                                aliases=catalog_entry.aliases,
                                compatible_roles=[
                                    str(role)
                                    for role in raw_candidate.get(
                                        "compatible_roles",
                                        [],
                                    )
                                ],
                                score_basis_points=int(
                                    raw_candidate.get("score_basis_points") or 0
                                ),
                                reasons=[
                                    str(reason) for reason in raw_candidate.get("reasons", [])
                                ],
                            )
                            )
                context = dict(field_match.context if field_match is not None else {})
                leaf_label = str(column.header_path[-1]).strip() if column.header_path else ""
                if not context.get("role") and re.fullmatch(r"\d{5}", leaf_label):
                    serial = int(leaf_label)
                    if 30_000 <= serial <= 60_000:
                        period_date = from_excel(serial).date().isoformat()
                        context["period_date"] = period_date
                        context["role"] = f"date_{period_date.replace('-', '_')}"
                headers.append(
                    HeaderEvidenceSummary(
                        header_candidate_id=candidate.id,
                        region_candidate_id=candidate.region_id,
                        source_column_id=column.source_column_id,
                        header_path=column.header_path,
                        evidence_cell_ids=column.evidence_cell_ids,
                        observed_data_type=(
                            field_match.observed_data_type if field_match is not None else None
                        ),
                        context=context,
                        semantic_candidates=field_candidates,
                    )
                )
            region = next(
                (item for item in sheet.region_candidates if item.id == candidate.region_id),
                None,
            )
            if region is None:
                continue
            sample_cell_count = 0
            sample_columns = sorted(
                [
                    column
                    for column in candidate.columns
                    if column.header_path
                    and any(str(part).strip() for part in column.header_path)
                ],
                key=lambda column: (
                    column.source_column_id not in unresolved_source_column_ids,
                    column.column,
                ),
            )
            data_start = max(candidate.header_rows) + 1
            populated_rows = [
                row
                for row in range(data_start, region.bounds.max_row + 1)
                if any(
                    (cell := cells_by_position.get((row, column.column))) is not None
                    and cell.display_value not in (None, "")
                    for column in sample_columns
                )
            ]
            for row in _representative_rows(populated_rows):
                sample_cells: list[SourceSampleCell] = []
                for column in sample_columns:
                    if sample_cell_count >= MAX_SAMPLE_CELLS:
                        break
                    cell = cells_by_position.get((row, column.column))
                    if cell is None or cell.display_value in (None, ""):
                        continue
                    value_kind, redacted_value = _redacted_sample_value(cell.display_value)
                    sample_cells.append(
                        SourceSampleCell(
                            source_column_id=column.source_column_id,
                            source_cell_id=cell.id,
                            value_kind=value_kind,
                            redacted_value=redacted_value,
                        )
                    )
                    sample_cell_count += 1
                if sample_cells:
                    source_samples.append(
                        SourceSampleRow(
                            region_candidate_id=candidate.region_id,
                            sheet_id=sheet.id,
                            source_row=row,
                            cells=sample_cells,
                        )
                    )
        for region in sheet.region_candidates:
            if region.id not in selected_region_ids:
                continue
            regions.append(
                RegionEvidenceSummary(
                    candidate_id=region.id,
                    sheet_id=sheet.id,
                    kind=region.kind,
                    range=region.bounds.range,
                    columns=region.bounds.max_column - region.bounds.min_column + 1,
                    density=region.density,
                )
            )
        if selected_region_ids:
            selected_regions = [
                region for region in sheet.region_candidates if region.id in selected_region_ids
            ]
            relevant_merges = []
            for merge in sheet.merges:
                min_column, min_row, max_column, max_row = range_boundaries(merge.range)
                if any(
                    min_column <= region.bounds.max_column
                    and max_column >= region.bounds.min_column
                    and min_row <= region.bounds.max_row
                    and max_row >= region.bounds.min_row
                    for region in selected_regions
                ):
                    relevant_merges.append(merge.id)
            merge_ids.extend(relevant_merges[:MAX_MERGE_EVIDENCE])
    resolved_source_column_ids = {
        entry.source_column_id
        for entry in field_matches or []
        if not entry.requires_hermes and entry.semantic_field_code is not None
    }
    unresolved_headers = [
        header
        for header in headers
        if header.header_path
        and " / ".join(header.header_path) in new_headers
        and header.source_column_id not in resolved_source_column_ids
    ]
    effective_new_headers = {" / ".join(header.header_path) for header in unresolved_headers}
    supplied_match_types = {
        str(entry.get("match_type")) for entry in unmatched_entries if entry.get("match_type")
    }
    request_match_type = (
        MatchType.NONE
        if supplied_match_types == {MatchType.NONE}
        else MatchType.PARTIAL
        if unmatched_entries
        else match.match_type
    )
    candidate_pairs = {
        (entry.get("template_id"), entry.get("template_version"))
        for entry in unmatched_entries
        if entry.get("template_id") and entry.get("template_version") is not None
    }
    candidate_template_id: uuid.UUID | None = None
    candidate_template_version: int | None = None
    if len(candidate_pairs) == 1:
        raw_template_id, raw_template_version = next(iter(candidate_pairs))
        if raw_template_id is not None and raw_template_version is not None:
            candidate_template_id = uuid.UUID(str(raw_template_id))
            candidate_template_version = int(raw_template_version)
    selected_sheet_ids = {region.sheet_id for region in regions}
    if selected_sheet_ids:
        sheet_summaries = [
            sheet for sheet in sheet_summaries if sheet.sheet_id in selected_sheet_ids
        ]
        range_evidence = [
            evidence
            for evidence in range_evidence
            if evidence.sheet_id in selected_sheet_ids
        ]
    return TemplateDiffRequest(
        profile_contract_version=profile.contract_version,
        parser_name=profile.parser_name,
        parser_version=profile.parser_version,
        layout_fingerprint=match.layout_fingerprint,
        template_id=candidate_template_id,
        template_version=candidate_template_version,
        match_type=request_match_type,
        new_headers=sorted(effective_new_headers),
        unresolved_source_column_ids=[header.source_column_id for header in unresolved_headers],
        missing_headers=sorted(missing_headers),
        headers=headers,
        regions=regions,
        merge_ids=sorted(merge_ids),
        source_samples=source_samples,
        sheets=sheet_summaries,
        range_evidence=range_evidence,
        semantic_catalog=semantic_catalog or [],
    )


def published_semantic_catalog(database: Session) -> list[SemanticFieldSummary]:
    versions = database.scalars(
        select(SemanticFieldVersion)
        .join(SemanticField)
        .where(
            SemanticFieldVersion.status == TemplateStatus.PUBLISHED,
            SemanticField.published_version == SemanticFieldVersion.version,
        )
        .order_by(SemanticField.code)
    )
    return [
        SemanticFieldSummary(
            code=version.field.code,
            name=version.name,
            layer=version.layer,
            data_type=version.data_type,
            unit=version.unit_dimension,
            aliases=sorted(
                {
                    *version.aliases,
                    *(variant.alias for variant in version.variants if variant.alias),
                    *(
                        variant.header_path[-1]
                        for variant in version.variants
                        if variant.header_path
                    ),
                }
            ),
        )
        for version in versions
    ]


def publish_unambiguous_new_fields(
    database: Session,
    *,
    request: TemplateDiffRequest,
    result: TemplateDiffResult,
) -> None:
    """Publish only contract-valid new fields from a governance-free recognition result."""
    if result.requires_governance:
        return
    headers_by_column = {header.source_column_id: header for header in request.headers}
    for decision in result.field_decisions:
        if decision.action != "PROPOSE_NEW_FIELD" or not decision.proposed_field_code:
            continue
        if (
            decision.layer is None
            or decision.data_type is None
            or decision.confidence < REVIEW_CONFIDENCE_THRESHOLD
        ):
            raise RecognitionValidationError(
                "governance-free new field lacks publishable metadata"
            )
        code = decision.proposed_field_code
        existing = database.scalar(select(SemanticField).where(SemanticField.code == code))
        if existing is not None:
            published = next(
                (
                    version
                    for version in existing.versions
                    if version.version == existing.published_version
                ),
                None,
            )
            if published is None or published.data_type != decision.data_type:
                raise RecognitionValidationError(
                    f"new field code conflicts with existing catalog: {code}"
                )
            continue
        header = headers_by_column.get(decision.source_column_id)
        header_path = list(header.header_path) if header is not None else [code]
        role = normalize_role_code(decision.role)
        name = (
            "每日完成数"
            if role is not None and role.startswith("date_")
            else header_path[-1]
        )
        field = SemanticField(code=code, published_version=1)
        version = SemanticFieldVersion(
            version=1,
            name=name,
            description="Hermes 二次判定并通过后端契约校验后自动发布",
            layer=decision.layer,
            data_type=decision.data_type,
            unit_dimension=decision.unit,
            aliases=[name],
            validators=[],
            status=TemplateStatus.PUBLISHED,
        )
        version.variants.append(
            build_field_variant(
                {
                    "kind": "header_path",
                    "header_path": header_path,
                    "source": "hermes_verified",
                    "confidence_basis_points": int(decision.confidence * 10_000),
                    "evidence": {
                        "source_column_id": decision.source_column_id,
                        "role": role,
                    },
                }
            )
        )
        field.versions.append(version)
        database.add(field)
        database.add(
            SemanticFieldReviewEvent(
                field_version=version,
                action="hermes_auto_publish",
                from_status=TemplateStatus.DRAFT,
                to_status=TemplateStatus.PUBLISHED,
                actor="system:hermes-verified",
                actor_type="system",
                comment="双模型判定无治理原因，自动沉淀为可复用字段",
            )
        )
    database.flush()


def validate_result(
    request: TemplateDiffRequest,
    result: TemplateDiffResult,
) -> None:
    column_ids = {header.source_column_id for header in request.headers}
    evidence_ids = _request_evidence_ids(request)
    for sample in request.source_samples:
        for cell in sample.cells:
            if cell.source_column_id not in column_ids:
                raise RecognitionValidationError(
                    f"sample references unknown source column: {cell.source_column_id}"
                )
    region_by_id = {region.candidate_id: region for region in request.regions}
    header_region = {
        header.header_candidate_id: header.region_candidate_id for header in request.headers
    }
    candidate_codes_by_column = {
        header.source_column_id: {candidate.code for candidate in header.semantic_candidates}
        for header in request.headers
    }

    decisions = result.field_decisions
    zero_record_structure = bool(result.layout_decisions) and not any(
        layout.materialize for layout in result.layout_decisions
    )
    if (
        request.match_type == "none"
        and result.template_suggestion is None
        and not zero_record_structure
    ):
        raise RecognitionValidationError("template suggestion is required when no template matches")
    if result.template_suggestion is not None:
        unknown = set(result.template_suggestion.evidence_ids) - evidence_ids
        if unknown:
            raise RecognitionValidationError(
                f"unknown template suggestion evidence ids: {', '.join(sorted(unknown))}"
            )
    decision_column_ids = [decision.source_column_id for decision in decisions]
    if len(decision_column_ids) != len(set(decision_column_ids)):
        raise RecognitionValidationError("duplicate field decision for source column")
    changed_column_ids = (
        set(request.unresolved_source_column_ids)
        if request.unresolved_source_column_ids
        else {
            header.source_column_id
            for header in request.headers
            if " / ".join(header.header_path) in set(request.new_headers)
        }
    )
    if set(decision_column_ids) != changed_column_ids:
        missing = changed_column_ids - set(decision_column_ids)
        extra = set(decision_column_ids) - changed_column_ids
        raise RecognitionValidationError(
            "field decisions must cover changed columns exactly; "
            f"missing={len(missing)}, extra={len(extra)}"
        )
    catalog_codes = {field.code for field in request.semantic_catalog}
    proposed_code_list = [
        decision.proposed_field_code
        for decision in decisions
        if decision.action == "PROPOSE_NEW_FIELD" and decision.proposed_field_code is not None
    ]
    if len(proposed_code_list) != len(set(proposed_code_list)):
        raise RecognitionValidationError(
            "a semantic field code may be proposed only once; use ROLE_VARIANT to reuse it"
        )
    proposed_codes = set(proposed_code_list)
    for decision in decisions:
        if decision.source_column_id not in column_ids:
            raise RecognitionValidationError(
                f"unknown source_column_id: {decision.source_column_id}"
            )
        if decision.action in {"REUSE_FIELD", "ADD_ALIAS"}:
            if decision.semantic_field_code not in catalog_codes:
                raise RecognitionValidationError(
                    "reused semantic field is not in the published catalog"
                )
            if decision.semantic_field_code not in candidate_codes_by_column.get(
                decision.source_column_id,
                set(),
            ):
                raise RecognitionValidationError(
                    "reused semantic field was not supplied as a field candidate"
                )
        if decision.action == "ROLE_VARIANT":
            if decision.semantic_field_code not in catalog_codes | proposed_codes:
                raise RecognitionValidationError(
                    "role variant field is neither published nor proposed"
                )
            if (
                decision.semantic_field_code in catalog_codes
                and decision.semantic_field_code
                not in candidate_codes_by_column.get(
                    decision.source_column_id,
                    set(),
                )
            ):
                raise RecognitionValidationError(
                    "role variant field was not supplied as a field candidate"
                )
        if decision.action == "PROPOSE_NEW_FIELD" and decision.proposed_field_code in catalog_codes:
            raise RecognitionValidationError(
                "proposed semantic field already exists in the published catalog"
            )
        unknown = set(decision.evidence_ids) - evidence_ids
        if unknown:
            raise RecognitionValidationError(f"unknown evidence ids: {', '.join(sorted(unknown))}")
    if result.record_grain is not None:
        unknown = set(result.record_grain.evidence_ids) - evidence_ids
        if unknown:
            raise RecognitionValidationError(
                f"unknown record grain evidence ids: {', '.join(sorted(unknown))}"
            )
    layout_region_ids = [layout.region_candidate_id for layout in result.layout_decisions]
    expected_region_ids = {region.candidate_id for region in request.regions}
    if len(layout_region_ids) != len(set(layout_region_ids)):
        raise RecognitionValidationError("duplicate layout decision for region")
    if set(layout_region_ids) != expected_region_ids:
        raise RecognitionValidationError("layout decisions must cover supplied regions exactly")
    for layout in result.layout_decisions:
        selected_region = region_by_id.get(layout.region_candidate_id)
        if selected_region is None:
            raise RecognitionValidationError(f"unknown layout region: {layout.region_candidate_id}")
        if header_region.get(layout.header_candidate_id) != layout.region_candidate_id:
            raise RecognitionValidationError(
                "layout header candidate does not belong to its region"
            )
        if layout.data_end_row < layout.data_start_row:
            raise RecognitionValidationError("layout data range is reversed")
        _, min_row, _, max_row = range_boundaries(selected_region.range)
        if layout.data_start_row < min_row or layout.data_end_row > max_row:
            raise RecognitionValidationError("layout data range exceeds candidate region")
        if layout.materialize and layout.classification == "noise":
            raise RecognitionValidationError("noise layout cannot be materialized")
        if not layout.materialize and layout.classification != "noise":
            raise RecognitionValidationError("ignored layout must be classified as noise")
        unknown = set(layout.evidence_ids) - evidence_ids
        if unknown:
            raise RecognitionValidationError(
                f"unknown layout evidence ids: {', '.join(sorted(unknown))}"
            )
        for merge in layout.merge_decisions:
            if merge.merge_id not in request.merge_ids:
                raise RecognitionValidationError(f"unknown merge evidence: {merge.merge_id}")
            if set(merge.target_source_column_ids) - column_ids:
                raise RecognitionValidationError(
                    "merge propagation targets an unknown source column"
                )
    if len(result.evidence_requests) > MAX_RANGE_REQUESTS:
        raise RecognitionValidationError("Hermes requested too many sheet ranges")


def recognition_cache_key(
    request: TemplateDiffRequest,
    *,
    hermes_version: str,
) -> str:
    payload = {
        "request": request.model_dump(mode="json"),
        "hermes_version": hermes_version,
        "prompt_version": PROMPT_VERSION,
        "schema_version": SCHEMA_VERSION,
    }
    canonical = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical.encode()).hexdigest()


def installed_hermes_version() -> str:
    try:
        return version("hermes-agent")
    except PackageNotFoundError:
        return "unavailable"


def governance_reasons(result: TemplateDiffResult) -> tuple[list[str], float | None]:
    confidences = [decision.confidence for decision in result.field_decisions] + [
        decision.confidence for decision in result.layout_decisions
    ]
    if result.template_suggestion is not None:
        confidences.append(result.template_suggestion.confidence)
    if result.record_grain is not None:
        confidences.append(result.record_grain.confidence)
    minimum = min(confidences) if confidences else None
    if (
        result.layout_decisions
        and not any(decision.materialize for decision in result.layout_decisions)
        and not result.field_decisions
    ):
        return [], minimum
    reasons: list[str] = []
    if minimum is None or minimum < REVIEW_CONFIDENCE_THRESHOLD:
        reasons.append("HERMES_LOW_CONFIDENCE")
    if any(
        decision.action in {"SEMANTIC_CONFLICT", "AMBIGUOUS"} for decision in result.field_decisions
    ):
        reasons.append("HERMES_SEMANTIC_CONFLICT")
    return reasons, minimum


def normalize_field_catalog_references(
    request: TemplateDiffRequest,
    result: TemplateDiffResult,
) -> tuple[TemplateDiffResult, bool]:
    """Turn unknown reuse targets into reviewable new-field candidates."""
    catalog_codes = {field.code for field in request.semantic_catalog}
    candidate_codes_by_column = {
        header.source_column_id: {candidate.code for candidate in header.semantic_candidates}
        for header in request.headers
    }
    candidates: list[FieldDecision] = []
    changed = False
    for decision in result.field_decisions:
        if (
            decision.action in {"REUSE_FIELD", "ADD_ALIAS", "ROLE_VARIANT"}
            and decision.semantic_field_code not in catalog_codes
        ):
            changed = True
            if decision.semantic_field_code is None:
                candidates.append(
                    decision.model_copy(
                        update={
                            "action": "AMBIGUOUS",
                            "confidence": min(decision.confidence, 0.5),
                            "requires_review": True,
                        }
                    )
                )
                continue
            candidates.append(
                decision.model_copy(
                    update={
                        "action": "PROPOSE_NEW_FIELD",
                        "semantic_field_code": None,
                        "proposed_field_code": decision.semantic_field_code,
                        "layer": decision.layer or "domain",
                        "data_type": decision.data_type or "text",
                        "confidence": min(decision.confidence, 0.75),
                        "requires_review": True,
                    }
                )
            )
            continue
        if decision.action in {
            "REUSE_FIELD",
            "ADD_ALIAS",
            "ROLE_VARIANT",
        } and decision.semantic_field_code not in candidate_codes_by_column.get(
            decision.source_column_id, set()
        ):
            changed = True
            candidates.append(
                decision.model_copy(
                    update={
                        "action": "AMBIGUOUS",
                        "semantic_field_code": None,
                        "confidence": min(decision.confidence, 0.5),
                        "requires_review": True,
                    }
                )
            )
            continue
        if decision.action == "PROPOSE_NEW_FIELD" and decision.proposed_field_code in catalog_codes:
            changed = True
            candidate_codes = candidate_codes_by_column.get(
                decision.source_column_id,
                set(),
            )
            if decision.proposed_field_code in candidate_codes:
                candidates.append(
                    decision.model_copy(
                        update={
                            "action": "REUSE_FIELD",
                            "semantic_field_code": decision.proposed_field_code,
                            "proposed_field_code": None,
                            "confidence": min(decision.confidence, 0.75),
                            "requires_review": True,
                        }
                    )
                )
            else:
                candidates.append(
                    decision.model_copy(
                        update={
                            "action": "AMBIGUOUS",
                            "semantic_field_code": None,
                            "proposed_field_code": None,
                            "confidence": min(decision.confidence, 0.5),
                            "requires_review": True,
                        }
                    )
                )
            continue
        candidates.append(decision)

    normalized: list[FieldDecision] = []
    proposed_codes: set[str] = set()
    for decision in candidates:
        if decision.action != "PROPOSE_NEW_FIELD":
            normalized.append(decision)
            continue
        proposed_code = decision.proposed_field_code
        if proposed_code is None:
            changed = True
            normalized.append(
                decision.model_copy(
                    update={
                        "action": "AMBIGUOUS",
                        "confidence": min(decision.confidence, 0.5),
                        "requires_review": True,
                    }
                )
            )
            continue
        if proposed_code in proposed_codes:
            changed = True
            normalized.append(
                decision.model_copy(
                    update={
                        "action": "ROLE_VARIANT",
                        "semantic_field_code": proposed_code,
                        "proposed_field_code": None,
                        "confidence": min(decision.confidence, 0.75),
                        "requires_review": True,
                    }
                )
            )
            continue
        proposed_codes.add(proposed_code)
        normalized.append(decision)
    headers_by_column = {header.source_column_id: header for header in request.headers}
    role_normalized: list[FieldDecision] = []
    for decision in normalized:
        header = headers_by_column.get(decision.source_column_id)
        inferred_role = normalize_role_code(
            str(header.context.get("role") or "") if header is not None else ""
        )
        returned_role = str(decision.role or "")
        normalized_returned_role = normalize_role_code(returned_role)
        if inferred_role:
            if returned_role != inferred_role:
                changed = True
            role_normalized.append(decision.model_copy(update={"role": inferred_role}))
            continue
        if normalized_returned_role:
            if returned_role != normalized_returned_role:
                changed = True
            role_normalized.append(decision.model_copy(update={"role": normalized_returned_role}))
            continue
        if returned_role:
            changed = True
            role_normalized.append(
                decision.model_copy(
                    update={
                        "role": None,
                        "requires_review": True,
                    }
                )
            )
            continue
        role_normalized.append(decision)
    date_scoped_codes: set[str] = set()
    stable_date_roles: list[FieldDecision] = []
    for decision in role_normalized:
        role = normalize_role_code(decision.role)
        proposed_code = decision.proposed_field_code
        match = (
            re.fullmatch(r"(.+?)[._]\d{4}_\d{2}_\d{2}", proposed_code)
            if decision.action == "PROPOSE_NEW_FIELD" and proposed_code
            else None
        )
        if role is None or not role.startswith("date_") or match is None:
            stable_date_roles.append(decision)
            continue
        stable_code = match.group(1)
        changed = True
        if stable_code in date_scoped_codes:
            stable_date_roles.append(
                decision.model_copy(
                    update={
                        "action": "ROLE_VARIANT",
                        "semantic_field_code": stable_code,
                        "proposed_field_code": None,
                    }
                )
            )
            continue
        date_scoped_codes.add(stable_code)
        stable_date_roles.append(
            decision.model_copy(update={"proposed_field_code": stable_code})
        )
    return (
        result.model_copy(update={"field_decisions": stable_date_roles}),
        changed,
    )


def _chunk_recognition_request(
    request: TemplateDiffRequest,
) -> list[TemplateDiffRequest]:
    chunks: list[TemplateDiffRequest] = []
    changed_paths = set(request.new_headers)
    supplied_column_ids = {header.source_column_id for header in request.headers}
    changed_column_ids = set(request.unresolved_source_column_ids) & supplied_column_ids
    if not changed_paths and not changed_column_ids:
        return []
    for region in request.regions:
        region_headers = [
            header
            for header in request.headers
            if header.region_candidate_id == region.candidate_id
        ]
        changed_headers = [
            header
            for header in region_headers
            if (
                header.source_column_id in changed_column_ids
                if changed_column_ids
                else " / ".join(header.header_path) in changed_paths
            )
        ]
        field_groups = [
            changed_headers[index : index + MAX_FIELDS_PER_HERMES_CALL]
            for index in range(0, len(changed_headers), MAX_FIELDS_PER_HERMES_CALL)
        ] or [region_headers[:1]]
        for field_group in field_groups:
            supplied_headers = field_group
            supplied_column_ids = {header.source_column_id for header in supplied_headers}
            samples: list[SourceSampleRow] = []
            for sample in request.source_samples:
                if sample.region_candidate_id != region.candidate_id:
                    continue
                cells = [
                    cell for cell in sample.cells if cell.source_column_id in supplied_column_ids
                ]
                if cells:
                    samples.append(sample.model_copy(update={"cells": cells}))
            chunks.append(
                request.model_copy(
                    update={
                        "new_headers": sorted(
                            {" / ".join(header.header_path) for header in field_group}
                        ),
                        "unresolved_source_column_ids": [
                            header.source_column_id for header in field_group
                        ],
                        "headers": supplied_headers,
                        "regions": [region],
                        "merge_ids": [],
                        "source_samples": samples,
                        "range_evidence": [],
                    }
                )
            )
    return chunks


def _expand_chunk_result(
    request: TemplateDiffRequest,
    compact: TemplateDiffChunkResult,
) -> TemplateDiffResult:
    default_evidence = [request.regions[0].candidate_id] if request.regions else []
    return TemplateDiffResult(
        template_suggestion=(
            TemplateSuggestion(
                **compact.template_suggestion.model_dump(),
                evidence_ids=default_evidence,
            )
            if compact.template_suggestion is not None
            else None
        ),
        record_grain=(
            RecordGrainDecision(
                **compact.record_grain.model_dump(),
                evidence_ids=default_evidence,
            )
            if compact.record_grain is not None
            else None
        ),
        layout_decisions=[
            LayoutDecision(
                **decision.model_dump(),
                materialize=decision.classification != "noise",
                evidence_ids=[
                    decision.region_candidate_id,
                    decision.header_candidate_id,
                ],
                merge_decisions=[],
            )
            for decision in compact.layout_decisions
        ],
        field_decisions=[
            FieldDecision(
                **decision.model_dump(),
                evidence_ids=[decision.source_column_id],
            )
            for decision in compact.field_decisions
        ],
    )


def _attach_structure_layouts(
    request: TemplateDiffRequest,
    result: TemplateDiffResult,
    structure: WorkbookStructureDecision | None,
) -> TemplateDiffResult:
    if structure is None:
        return result
    chunk_region_ids = {region.candidate_id for region in request.regions}
    chunk_headers = {header.header_candidate_id for header in request.headers}
    return result.model_copy(
        update={
            "layout_decisions": [
                decision.model_copy(
                    update={
                        "evidence_ids": [
                            decision.region_candidate_id,
                            decision.header_candidate_id,
                        ],
                        "merge_decisions": [],
                    }
                )
                for decision in structure.layout_decisions
                if decision.region_candidate_id in chunk_region_ids
                and decision.header_candidate_id in chunk_headers
            ]
        }
    )


def _merge_chunk_results(
    results: list[TemplateDiffResult],
    request: TemplateDiffRequest,
) -> TemplateDiffResult:
    if not results:
        raise RecognitionValidationError("Hermes returned no recognition chunks")
    field_decisions: list[FieldDecision] = []
    proposed_codes: set[str] = set()
    expected_column_ids = {
        header.source_column_id
        for header in request.headers
        if " / ".join(header.header_path) in set(request.new_headers)
    }
    layout_by_region: dict[str, LayoutDecision] = {}
    suggestions: list[TemplateSuggestion] = []
    grains: list[RecordGrainDecision] = []
    for result in results:
        for decision in result.field_decisions:
            if decision.source_column_id not in expected_column_ids:
                continue
            proposed_code = decision.proposed_field_code
            if (
                decision.action == "PROPOSE_NEW_FIELD"
                and proposed_code is not None
                and proposed_code in proposed_codes
            ):
                decision = decision.model_copy(
                    update={
                        "action": "ROLE_VARIANT",
                        "semantic_field_code": proposed_code,
                        "proposed_field_code": None,
                    }
                )
            elif decision.action == "PROPOSE_NEW_FIELD" and proposed_code:
                proposed_codes.add(proposed_code)
            field_decisions.append(decision)
        for layout in result.layout_decisions:
            current = layout_by_region.get(layout.region_candidate_id)
            if current is None or layout.confidence > current.confidence:
                layout_by_region[layout.region_candidate_id] = layout
        if result.template_suggestion is not None:
            suggestions.append(result.template_suggestion)
        if result.record_grain is not None:
            grains.append(result.record_grain)
    return TemplateDiffResult(
        template_suggestion=(
            max(suggestions, key=lambda item: item.confidence) if suggestions else None
        ),
        record_grain=(max(grains, key=lambda item: item.confidence) if grains else None),
        layout_decisions=list(layout_by_region.values()),
        field_decisions=field_decisions,
    )


def _retain_requested_field_decisions(
    request: TemplateDiffRequest,
    result: TemplateDiffResult,
) -> TemplateDiffResult:
    """Prevent a model review from overwriting fields already resolved by templates."""
    requested_ids = (
        set(request.unresolved_source_column_ids)
        if request.unresolved_source_column_ids
        else {
            header.source_column_id
            for header in request.headers
            if " / ".join(header.header_path) in set(request.new_headers)
        }
    )
    return result.model_copy(
        update={
            "field_decisions": [
                decision
                for decision in result.field_decisions
                if decision.source_column_id in requested_ids
            ]
        }
    )


def _stabilize_field_only_result(
    request: TemplateDiffRequest,
    result: TemplateDiffResult,
) -> TemplateDiffResult:
    """Keep layout ownership in the backend when only exact-Region fields are unresolved."""
    if request.match_type != MatchType.EXACT or not request.unresolved_source_column_ids:
        return result
    header_by_region = {
        header.region_candidate_id: header.header_candidate_id for header in request.headers
    }
    layouts: list[LayoutDecision] = []
    for region in request.regions:
        header_candidate_id = header_by_region.get(region.candidate_id)
        if header_candidate_id is None:
            raise RecognitionValidationError(
                f"exact Region has no header candidate: {region.candidate_id}"
            )
        _, min_row, _, max_row = range_boundaries(region.range)
        classification = region.kind if region.kind in {"table", "form", "matrix"} else "table"
        layouts.append(
            LayoutDecision(
                region_candidate_id=region.candidate_id,
                header_candidate_id=header_candidate_id,
                data_start_row=min_row,
                data_end_row=max_row,
                excluded_rows=[],
                classification=classification,
                materialize=True,
                confidence=1.0,
                evidence_ids=[region.candidate_id, header_candidate_id],
                merge_decisions=[],
            )
        )
    return result.model_copy(
        update={
            "template_suggestion": None,
            "record_grain": None,
            "layout_decisions": layouts,
            "structure_decision": None,
        }
    )


def _request_for_sheet(
    request: TemplateDiffRequest,
    sheet_id: str,
) -> TemplateDiffRequest:
    region_ids = {region.candidate_id for region in request.regions if region.sheet_id == sheet_id}
    headers = [header for header in request.headers if header.region_candidate_id in region_ids]
    return request.model_copy(
        update={
            "new_headers": [],
            "missing_headers": [],
            "unresolved_source_column_ids": [],
            "headers": [
                header.model_copy(update={"evidence_cell_ids": header.evidence_cell_ids[:4]})
                for header in headers
            ],
            "regions": [region for region in request.regions if region.sheet_id == sheet_id],
            "merge_ids": [
                merge_id for merge_id in request.merge_ids if merge_id.startswith(f"{sheet_id}:")
            ],
            "source_samples": [],
            "sheets": [sheet for sheet in request.sheets if sheet.sheet_id == sheet_id],
            "range_evidence": [
                evidence for evidence in request.range_evidence if evidence.sheet_id == sheet_id
            ],
            "semantic_catalog": [],
        }
    )


async def _recognize_sheet_structure(
    *,
    profile: WorkbookProfile,
    request: TemplateDiffRequest,
    runtime: HermesRuntime,
    item_id: uuid.UUID,
    model: str,
    reasoning_model: str | None,
) -> tuple[WorkbookStructureDecision, TemplateDiffRequest, list[str], bool]:
    if len(request.sheets) != 1:
        raise RecognitionValidationError(
            "one structure interpretation request must contain exactly one physical sheet"
        )
    sheet_id = request.sheets[0].sheet_id
    task_suffix = hashlib.sha256(sheet_id.encode()).hexdigest()[:12]
    system_prompt = (
        "This is the structure interpretation for exactly one physical spreadsheet "
        "Sheet. The deterministic "
        "profile contains physical facts and high-recall candidates, not accepted "
        "business structure. Decide the real data regions for every supplied candidate. "
        "Each range-evidence row cell is the compact tuple "
        "[column, redacted_value, data_type, style_ref]. "
        "Set materialize=false and classification=noise for candidates that are ordinary "
        "data, notes, or false header detections. For retained regions, identify the "
        "header candidate and exact data row bounds. The layout bounds and excluded_rows "
        "are the authoritative compact row contract. Do not enumerate ordinary data rows. "
        "Use row_role_segments only for short exceptional ranges whose role materially "
        "explains a boundary, such as title, context, summary, note, footer, or separator; "
        "it may be empty. Use only supplied IDs and observed row bounds. Initial and body "
        "previews are analysis evidence only. If evidence is insufficient, request at "
        "most three additional ranges of at most 30 rows and 40 columns each, while still "
        "returning a complete conservative decision. Do not infer field semantics, write "
        "data, or treat rule candidates as authoritative."
    )
    performed_models: list[str] = []
    try:
        result = await runtime.run_json(
            system_prompt=system_prompt,
            user_prompt=request.model_dump_json(),
            output_model=WorkbookStructureDecision,
            policy=HermesCallPolicy(
                thinking_enabled=False,
                enabled_toolsets=(),
                repair_attempts=1,
                timeout_seconds=90,
                max_tokens=4096,
            ),
            task_id=f"sheet-structure-{item_id}-{task_suffix}",
        )
        performed_models.append(model)
    except HermesUnavailableError:
        result = await runtime.run_json(
            system_prompt=(
                system_prompt + " The fast structure pass failed. Produce the complete bounded "
                "structure decision directly from the supplied physical evidence."
            ),
            user_prompt=request.model_dump_json(),
            output_model=WorkbookStructureDecision,
            policy=HermesCallPolicy(
                thinking_enabled=True,
                reasoning_effort="high",
                enabled_toolsets=(),
                repair_attempts=1,
                timeout_seconds=120,
                max_tokens=4096,
            ),
            task_id=f"sheet-structure-fallback-{item_id}-{task_suffix}",
        )
        performed_models.append(reasoning_model or model)
    result, _ = normalize_structure_data_ranges(
        request,
        result,
    )
    result, _ = normalize_ignored_structure_ranges(
        request,
        result,
    )
    result, _ = normalize_structure_merge_references(
        profile,
        request,
        result,
    )
    try:
        validate_structure_decision(profile, request, result)
    except RecognitionValidationError as exc:
        repair_payload = request.model_dump(mode="json")
        repair_payload["previous_result"] = result.model_dump(mode="json")
        repair_payload["deterministic_validation_error"] = str(exc)
        result = await runtime.run_json(
            system_prompt=(
                system_prompt + " The backend rejected the previous structure decision with the "
                "supplied deterministic_validation_error. Return one complete corrected "
                "decision for this Sheet. Repair only the contract conflict; do not "
                "invent coordinates, evidence, headers, regions, or values."
            ),
            user_prompt=json.dumps(repair_payload, ensure_ascii=False),
            output_model=WorkbookStructureDecision,
            policy=HermesCallPolicy(
                thinking_enabled=True,
                reasoning_effort="high",
                enabled_toolsets=(),
                repair_attempts=1,
                timeout_seconds=120,
                max_tokens=4096,
            ),
            task_id=f"sheet-structure-contract-repair-{item_id}-{task_suffix}",
        )
        performed_models.append(reasoning_model or model)
        result, _ = normalize_structure_data_ranges(
            request,
            result,
        )
        result, _ = normalize_ignored_structure_ranges(
            request,
            result,
        )
        result, _ = normalize_structure_merge_references(
            profile,
            request,
            result,
        )
        validate_structure_decision(profile, request, result)

    unresolved_range_request = False
    if result.evidence_requests:
        additions = fulfill_range_requests(profile, result.evidence_requests)
        enriched = request.model_copy(
            update={
                "range_evidence": [*request.range_evidence, *additions],
            }
        )
        result = await runtime.run_json(
            system_prompt=(
                system_prompt
                + " The backend fulfilled your bounded range requests. Return the final "
                "complete decision. Do not request further evidence."
            ),
            user_prompt=enriched.model_dump_json(),
            output_model=WorkbookStructureDecision,
            policy=HermesCallPolicy(
                thinking_enabled=True,
                reasoning_effort="high",
                enabled_toolsets=(),
                repair_attempts=1,
                timeout_seconds=120,
                max_tokens=4096,
            ),
            task_id=f"sheet-structure-review-{item_id}-{task_suffix}",
        )
        performed_models.append(reasoning_model or model)
        result, _ = normalize_structure_data_ranges(
            enriched,
            result,
        )
        result, _ = normalize_ignored_structure_ranges(
            enriched,
            result,
        )
        result, _ = normalize_structure_merge_references(
            profile,
            enriched,
            result,
        )
        validate_structure_decision(profile, enriched, result)
        unresolved_range_request = bool(result.evidence_requests)
        request = enriched
    return (
        result,
        request,
        performed_models,
        unresolved_range_request,
    )


async def recognize_workbook_structure(
    *,
    profile: WorkbookProfile,
    request: TemplateDiffRequest,
    runtime: HermesRuntime,
    item_id: uuid.UUID,
    model: str,
    reasoning_model: str | None,
) -> tuple[WorkbookStructureDecision, TemplateDiffRequest, list[str], bool]:
    sheet_ids = list(dict.fromkeys(region.sheet_id for region in request.regions))
    if not sheet_ids:
        raise RecognitionValidationError("Hermes structure request has no Sheet")
    decisions: list[WorkbookStructureDecision] = []
    requested_evidence: list[SheetRangeEvidence] = []
    performed_models: list[str] = []
    requires_governance = False
    for sheet_id in sheet_ids:
        sheet_request = _request_for_sheet(request, sheet_id)
        decision, enriched, models, uncertain = await _recognize_sheet_structure(
            profile=profile,
            request=sheet_request,
            runtime=runtime,
            item_id=item_id,
            model=model,
            reasoning_model=reasoning_model,
        )
        decisions.append(decision)
        requested_evidence.extend(
            evidence for evidence in enriched.range_evidence if evidence.purpose == "requested"
        )
        performed_models.extend(models)
        requires_governance = requires_governance or uncertain
    merged = WorkbookStructureDecision(
        row_role_segments=[
            segment for decision in decisions for segment in decision.row_role_segments
        ],
        layout_decisions=[layout for decision in decisions for layout in decision.layout_decisions],
        evidence_requests=[
            request for decision in decisions for request in decision.evidence_requests
        ],
        confidence=min(decision.confidence for decision in decisions),
    )
    enriched_request = request.model_copy(
        update={
            "range_evidence": [*request.range_evidence, *requested_evidence],
        }
    )
    validate_structure_decision(profile, enriched_request, merged)
    return (
        merged,
        enriched_request,
        performed_models,
        requires_governance or bool(merged.evidence_requests),
    )


async def recognize_differences(
    database: Session,
    *,
    item_id: uuid.UUID,
    request: TemplateDiffRequest,
    profile: WorkbookProfile | None = None,
    runtime: HermesRuntime,
    provider: str,
    model: str,
    reasoning_model: str | None = None,
    hermes_version: str | None = None,
) -> TemplateProposal:
    runtime_version = hermes_version or installed_hermes_version()
    original_request = request
    field_only_reuse = (
        request.match_type == MatchType.EXACT
        and bool(request.unresolved_source_column_ids)
    )
    cache_key = recognition_cache_key(request, hermes_version=runtime_version)
    cached = database.get(HermesRecognitionCache, cache_key)
    call_performed = cached is None
    performed_models: list[str] = []
    if cached is None:
        fast_cache_key = hashlib.sha256(f"{cache_key}:fast".encode()).hexdigest()
        fast_cached = database.get(HermesRecognitionCache, fast_cache_key)
        structure_decision: WorkbookStructureDecision | None = None
        structure_requires_governance = False
        if fast_cached is None:
            if profile is not None and not field_only_reuse:
                (
                    structure_decision,
                    request,
                    structure_models,
                    structure_requires_governance,
                ) = await recognize_workbook_structure(
                    profile=profile,
                    request=request,
                    runtime=runtime,
                    item_id=item_id,
                    model=model,
                    reasoning_model=reasoning_model,
                )
                performed_models.extend(structure_models)
                request = apply_structure_decision(request, structure_decision)
            request_chunks = _chunk_recognition_request(request)
            chunk_results: list[TemplateDiffResult] = []
            for index, request_chunk in enumerate(request_chunks):
                compact_result = await runtime.run_json(
                    system_prompt=(
                        "You make only the minimal semantic judgment for the supplied "
                        "changed structured-document fields. Never invent coordinates, "
                        "fields, or source values. The backend, not you, attaches evidence, "
                        "handles merged cells, and applies governance. Use source_samples "
                        "only as redacted value-shape evidence. Return exactly one "
                        "field_decision for each supplied unresolved_source_column_id (or, "
                        "for older requests, each header path in new_headers), and "
                        "exactly one layout_decision per supplied region. REUSE_FIELD and "
                        "ADD_ALIAS may reference only codes in that header's "
                        "semantic_candidates. ROLE_VARIANT may reference a supplied "
                        "candidate code or a code proposed once in this response. If no "
                        "candidate fits, use PROPOSE_NEW_FIELD; if evidence cannot decide, "
                        "use AMBIGUOUS. Never invent a code and label it as reused. "
                        "The header context may contain a normalized role. Preserve that "
                        "role when reusing the base semantic field. Role values must be "
                        "short stable codes from context.role or a candidate's "
                        "compatible_roles; never copy a workbook title as a role. "
                        "A date, day, month, reporting period, or Excel date serial in a "
                        "header is a role/dimension, not part of semantic field identity. "
                        "For repeated daily or period measures, propose one stable code "
                        "without a date suffix and use ROLE_VARIANT with the supplied "
                        "context.role for the other periods. "
                        "When multiple supplied columns select the same semantic code, or "
                        "the header expresses a relationship such as household head, "
                        "spouse, mother, father, or child, set a short distinct role on "
                        "each non-canonical decision. "
                        "data_type must be "
                        "one of text, integer, decimal, boolean, date, or datetime. "
                        "When match_type is none, also propose a reusable "
                        "template_code, Chinese template_name, domain, and record_type "
                        "in template_suggestion. Do not return evidence IDs, merge "
                        "decisions, governance fields, candidates, or explanations."
                    ),
                    user_prompt=request_chunk.model_dump_json(),
                    output_model=TemplateDiffChunkResult,
                    policy=HermesCallPolicy(
                        thinking_enabled=False,
                        enabled_toolsets=(),
                        repair_attempts=1,
                        timeout_seconds=120,
                        max_tokens=4096,
                    ),
                    task_id=f"template-diff-{item_id}-part-{index + 1}",
                )
                chunk_result = _expand_chunk_result(request_chunk, compact_result)
                chunk_result, _ = normalize_field_catalog_references(
                    request_chunk,
                    chunk_result,
                )
                chunk_result = _retain_requested_field_decisions(
                    request_chunk,
                    chunk_result,
                )
                chunk_result = _stabilize_field_only_result(request_chunk, chunk_result)
                chunk_result = _attach_structure_layouts(
                    request_chunk,
                    chunk_result,
                    structure_decision,
                )
                try:
                    validate_result(request_chunk, chunk_result)
                except RecognitionValidationError as exc:
                    repair_payload = request_chunk.model_dump(mode="json")
                    repair_payload["previous_result"] = compact_result.model_dump(mode="json")
                    repair_payload["deterministic_validation_error"] = str(exc)
                    compact_result = await runtime.run_json(
                        system_prompt=(
                            "Repair one semantic field chunk rejected by the backend. "
                            "Return exactly one decision for every supplied "
                            "unresolved_source_column_id and use only supplied IDs and "
                            "semantic candidate codes. Correct only the reported contract "
                            "conflict. Do not invent coordinates, fields, values, evidence "
                            "IDs, or explanations."
                        ),
                        user_prompt=json.dumps(
                            repair_payload,
                            ensure_ascii=False,
                        ),
                        output_model=TemplateDiffChunkResult,
                        policy=HermesCallPolicy(
                            thinking_enabled=True,
                            reasoning_effort="high",
                            enabled_toolsets=(),
                            repair_attempts=1,
                            timeout_seconds=120,
                            max_tokens=4096,
                        ),
                        task_id=(f"template-diff-contract-repair-{item_id}-part-{index + 1}"),
                    )
                    chunk_result = _expand_chunk_result(
                        request_chunk,
                        compact_result,
                    )
                    chunk_result, _ = normalize_field_catalog_references(
                        request_chunk,
                        chunk_result,
                    )
                    chunk_result = _retain_requested_field_decisions(
                        request_chunk,
                        chunk_result,
                    )
                    chunk_result = _stabilize_field_only_result(request_chunk, chunk_result)
                    chunk_result = _attach_structure_layouts(
                        request_chunk,
                        chunk_result,
                        structure_decision,
                    )
                    validate_result(request_chunk, chunk_result)
                    performed_models.append(reasoning_model or model)
                chunk_results.append(chunk_result)
                performed_models.append(model)
            result = (
                _merge_chunk_results(chunk_results, request)
                if chunk_results
                else TemplateDiffResult(
                    layout_decisions=(
                        list(structure_decision.layout_decisions)
                        if structure_decision is not None
                        else []
                    ),
                    structure_decision=structure_decision,
                )
            )
            result = _stabilize_field_only_result(request, result)
            retained_region_ids = {region.candidate_id for region in request.regions}
            result = result.model_copy(
                update={
                    "structure_decision": (
                        structure_decision if not field_only_reuse else None
                    ),
                    "layout_decisions": (
                        [
                            decision
                            for decision in structure_decision.layout_decisions
                            if decision.region_candidate_id in retained_region_ids
                        ]
                        if structure_decision is not None and not field_only_reuse
                        else result.layout_decisions
                    ),
                }
            )
            validate_result(request, result)
            fast_cached = HermesRecognitionCache(
                cache_key=fast_cache_key,
                hermes_version=runtime_version,
                prompt_version=PROMPT_VERSION,
                schema_version=SCHEMA_VERSION,
                provider=provider,
                model=model,
                request_payload=original_request.model_dump(mode="json"),
                response_payload=result.model_dump(mode="json"),
            )
            database.add(fast_cached)
            database.commit()
        else:
            result = TemplateDiffResult.model_validate(fast_cached.response_payload)
            structure_decision = result.structure_decision
            result = _stabilize_field_only_result(request, result)
            if structure_decision is not None and not field_only_reuse:
                request = apply_structure_decision(request, structure_decision)
        chunked = len(_chunk_recognition_request(request)) > 1
        first_reasons, _ = governance_reasons(result)
        if (
            not field_only_reuse
            and any(decision.materialize for decision in result.layout_decisions)
            and (
            structure_requires_governance
            or (
                structure_decision is not None
                and structure_decision.confidence < REVIEW_CONFIDENCE_THRESHOLD
            )
            )
        ):
            first_reasons.append("HERMES_STRUCTURE_REVIEW_REQUIRED")
        fast_validation_error: str | None = None
        try:
            validate_result(request, result)
        except RecognitionValidationError as exc:
            fast_validation_error = str(exc)
            first_reasons.append("HERMES_CONTRACT_CONFLICT")
        call_count = 1
        if first_reasons and not chunked:
            reconsideration_payload = request.model_dump(mode="json")
            reconsideration_payload["previous_result"] = result.model_dump(mode="json")
            reconsideration_payload["deterministic_validation_error"] = fast_validation_error
            reconsideration_payload["review_instruction"] = (
                "Re-evaluate low-confidence, conflicting, or deterministically invalid "
                "decisions using the same headers and source_samples. Return the complete "
                "corrected result."
            )
            result = await runtime.run_json(
                system_prompt=(
                    "You are the second-pass reviewer for a structured spreadsheet plan. "
                    "Use the supplied headers, redacted representative source samples, and "
                    "the previous result to resolve conflicts. Do not invent evidence IDs, "
                    "coordinates, fields, or values. Return the complete result, covering "
                    "every supplied changed column and region exactly once."
                ),
                user_prompt=json.dumps(reconsideration_payload, ensure_ascii=False),
                output_model=TemplateDiffResult,
                policy=HermesCallPolicy(
                    thinking_enabled=True,
                    reasoning_effort="high",
                    enabled_toolsets=(),
                    repair_attempts=1,
                    timeout_seconds=300,
                ),
                task_id=f"template-diff-review-{item_id}",
            )
            result, _ = normalize_field_catalog_references(request, result)
            result = _retain_requested_field_decisions(request, result)
            result = _stabilize_field_only_result(request, result)
            validate_result(request, result)
            result = result.model_copy(
                update={
                    "structure_decision": (
                        structure_decision if not field_only_reuse else None
                    ),
                    "layout_decisions": (
                        [
                            decision
                            for decision in structure_decision.layout_decisions
                            if decision.region_candidate_id
                            in {region.candidate_id for region in request.regions}
                        ]
                        if structure_decision is not None and not field_only_reuse
                        else result.layout_decisions
                    ),
                }
            )
            call_count = 2
            performed_models.append(reasoning_model or model)
        final_reasons, minimum_confidence = governance_reasons(result)
        if (
            not field_only_reuse
            and any(decision.materialize for decision in result.layout_decisions)
            and (
                structure_requires_governance
                or (
                    structure_decision is not None
                    and structure_decision.confidence < REVIEW_CONFIDENCE_THRESHOLD
                )
            )
            and "HERMES_STRUCTURE_REVIEW_REQUIRED" not in final_reasons
        ):
            final_reasons.append("HERMES_STRUCTURE_REVIEW_REQUIRED")
        result = result.model_copy(
            update={
                "recognition_passes": call_count,
                "requires_governance": bool(final_reasons),
                "governance_reason_codes": final_reasons,
                "minimum_confidence": minimum_confidence,
            }
        )
        cached = HermesRecognitionCache(
            cache_key=cache_key,
            hermes_version=runtime_version,
            prompt_version=PROMPT_VERSION,
            schema_version=SCHEMA_VERSION,
            provider=provider,
            model=reasoning_model if call_count == 2 and reasoning_model else model,
            request_payload=original_request.model_dump(mode="json"),
            response_payload=result.model_dump(mode="json"),
        )
        database.add(cached)
        database.flush()
    else:
        result = TemplateDiffResult.model_validate(cached.response_payload)
        if result.structure_decision is not None and not field_only_reuse:
            request = apply_structure_decision(request, result.structure_decision)
        result, _ = normalize_field_catalog_references(request, result)
        result = _retain_requested_field_decisions(request, result)
        result = _stabilize_field_only_result(request, result)
        validate_result(request, result)

    call_performed = bool(performed_models)
    input_field_count = len(request.new_headers) + len(request.missing_headers)
    models = performed_models if call_performed else [model]
    for used_model in models:
        database.add(
            HermesRecognitionRecord(
                item_id=item_id,
                cache_key=cache_key,
                call_performed=call_performed,
                input_field_count=input_field_count,
                provider=provider,
                model=used_model,
            )
        )
    proposal_key = f"template-diff:{item_id}:{cache_key}"
    proposal = database.scalar(
        select(TemplateProposal).where(TemplateProposal.idempotency_key == proposal_key)
    )
    if proposal is None:
        item = database.get(IngestionItem, item_id)
        if item is None:
            raise RecognitionValidationError("proposal source item is unavailable")
        proposal = TemplateProposal(
            tenant_id=item.tenant_id,
            administrative_unit_id=item.administrative_unit_id,
            created_by_user_id=item.created_by_user_id,
            idempotency_key=proposal_key,
            source="hermes" if call_performed else "cache",
            source_item_id=item_id,
            model_name=(
                reasoning_model if result.recognition_passes == 2 and reasoning_model else model
            ),
            prompt_version=PROMPT_VERSION,
            proposal=result.model_dump(mode="json"),
            status=ProposalStatus.PENDING,
        )
        database.add(proposal)
    database.flush()
    return proposal


def _provisional_record_type(
    *,
    result: TemplateDiffResult,
    base_definition: TemplateDefinition | None,
) -> str:
    if base_definition is not None:
        return base_definition.record_type
    if result.template_suggestion is not None:
        return result.template_suggestion.record_type
    if result.record_grain is None:
        return "unclassified_record"
    grain = result.record_grain.value.strip().lower()
    if not re.fullmatch(r"[a-z][a-z0-9_]{1,119}", grain):
        return "unclassified_record"
    for prefix in ("one_row_per_", "one_"):
        if grain.startswith(prefix):
            grain = grain[len(prefix) :]
            break
    if grain in {"record", "row", "item"}:
        return "unclassified_record"
    return grain


def create_provisional_template(
    database: Session,
    *,
    proposal: TemplateProposal,
) -> tuple[DocumentTemplate, TemplateVersion]:
    if proposal.source_item_id is None:
        raise ProposalResolutionError("proposal is not bound to a source item")
    result = TemplateDiffResult.model_validate(proposal.proposal)
    existing_version = next(
        (
            version
            for version in database.scalars(
                select(TemplateVersion).where(
                    TemplateVersion.status == TemplateStatus.ADMIN_REVIEW,
                    TemplateVersion.source == "hermes_provisional",
                )
            )
            if str(version.source_metadata.get("proposal_id")) == str(proposal.id)
        ),
        None,
    )
    if existing_version is not None:
        definition = TemplateDefinition.model_validate(existing_version.definition)
        inferred_record_type = _provisional_record_type(
            result=result,
            base_definition=None,
        )
        if (
            definition.record_type == "unclassified_record"
            and inferred_record_type != definition.record_type
        ):
            existing_version.definition = definition.model_copy(
                update={"record_type": inferred_record_type}
            ).model_dump(mode="json")
            database.flush()
        return existing_version.template, existing_version

    match = database.get(TemplateMatch, proposal.source_item_id)
    if match is None:
        raise ProposalResolutionError("proposal template match is unavailable")
    suggestion = result.template_suggestion
    base_definition: TemplateDefinition | None = None
    candidate_template_id, candidate_template_version = _unmatched_template_candidate(match)
    if candidate_template_id is not None and candidate_template_version is not None:
        template = database.get(DocumentTemplate, candidate_template_id)
        base_version = database.scalar(
            select(TemplateVersion).where(
                TemplateVersion.template_id == candidate_template_id,
                TemplateVersion.version == candidate_template_version,
                TemplateVersion.status == TemplateStatus.PUBLISHED,
            )
        )
        if template is None or base_version is None:
            raise ProposalResolutionError("matched published template is unavailable")
        base_definition = TemplateDefinition.model_validate(base_version.definition)
        template_name = base_version.name
    else:
        if suggestion is None:
            source_item = database.get(IngestionItem, proposal.source_item_id)
            requested_code = (
                f"hermes.provisional.{str(proposal.source_item_id).replace('-', '')[:20]}"
            )
            template_name = (
                f"{source_item.original_name} 临时结构"
                if source_item is not None
                else f"待治理结构 {str(proposal.source_item_id)[:8]}"
            )
        else:
            requested_code = suggestion.template_code
            template_name = suggestion.template_name
        source_suffix = str(proposal.source_item_id).replace("-", "")[:8]
        proposal_suffix = str(proposal.id).replace("-", "")[:8]
        template = None
        for code in (
            requested_code,
            f"{requested_code}.{source_suffix}",
            f"{requested_code}.{source_suffix}.{proposal_suffix}",
        ):
            existing = database.scalar(
                select(DocumentTemplate).where(DocumentTemplate.code == code)
            )
            if existing is None:
                template = DocumentTemplate(code=code)
                database.add(template)
                break
            if any(
                version.source == "hermes_provisional"
                and str(version.source_metadata.get("source_item_id"))
                == str(proposal.source_item_id)
                for version in existing.versions
            ):
                template = existing
                break
        if template is None:
            raise ProposalResolutionError(
                "cannot allocate an idempotent provisional template code"
            )

    definition = TemplateDefinition(
        domain=(
            base_definition.domain
            if base_definition is not None
            else suggestion.domain
            if suggestion is not None
            else "unclassified"
        ),
        region_kind=base_definition.region_kind if base_definition is not None else "table",
        record_type=(
            _provisional_record_type(
                result=result,
                base_definition=base_definition,
            )
        ),
        record_grain=(
            result.record_grain.value
            if result.record_grain is not None
            else base_definition.record_grain
            if base_definition is not None
            else "one_row_per_record"
        ),
        field_bindings=(
            list(base_definition.field_bindings) if base_definition is not None else []
        ),
        data_row_rules=base_definition.data_row_rules if base_definition is not None else [],
        exclusion_rules=base_definition.exclusion_rules if base_definition is not None else [],
        metric_codes=base_definition.metric_codes if base_definition is not None else [],
        identity_field_codes=(
            base_definition.identity_field_codes if base_definition is not None else []
        ),
    )
    next_version = max((item.version for item in template.versions), default=0) + 1
    version = TemplateVersion(
        version=next_version,
        name=template_name,
        description="Hermes 自动生成的临时模板，仅用于原始 JSONB 部分入库",
        status=TemplateStatus.ADMIN_REVIEW,
        layout_fingerprint=match.layout_fingerprint,
        definition=definition.model_dump(mode="json"),
        source="hermes_provisional",
        source_metadata={
            "proposal_id": str(proposal.id),
            "source_item_id": str(proposal.source_item_id),
            "requires_governance": result.requires_governance,
            "governance_reason_codes": result.governance_reason_codes,
            "approved_layout_plan": [
                decision.model_dump(mode="json") for decision in result.layout_decisions
            ],
        },
    )
    template.versions.append(version)
    database.flush()
    return template, version


def _unmatched_template_candidate(
    match: TemplateMatch,
) -> tuple[uuid.UUID | None, int | None]:
    entries = match.differences.get("unmatched_regions", [])
    if isinstance(entries, list):
        candidates = {
            (entry.get("template_id"), entry.get("template_version"))
            for entry in entries
            if isinstance(entry, dict)
            and entry.get("template_id")
            and entry.get("template_version") is not None
        }
        if len(candidates) == 1:
            template_id, template_version = next(iter(candidates))
            if template_id is not None and template_version is not None:
                return uuid.UUID(str(template_id)), int(template_version)
        if entries:
            return None, None
    if match.template_id is not None and match.match_type != MatchType.NONE:
        return match.template_id, match.template_version
    return None, None


def accept_recognition_proposal(
    database: Session,
    *,
    proposal: TemplateProposal,
    actor: str,
    comment: str,
    template_code: str | None,
    template_name: str,
    domain: str,
    record_type: str,
    record_grain: str | None,
    field_decisions: list[dict[str, Any]] | None = None,
    actor_user_id: uuid.UUID | None = None,
) -> DocumentTemplate:
    if proposal.status != ProposalStatus.PENDING:
        raise ProposalResolutionError("proposal has already been resolved")
    if proposal.source_item_id is None:
        raise ProposalResolutionError("proposal is not bound to a source item")
    match = database.get(TemplateMatch, proposal.source_item_id)
    profile_record = database.get(DocumentProfile, proposal.source_item_id)
    if match is None or profile_record is None:
        raise ProposalResolutionError("proposal evidence is unavailable")
    result = TemplateDiffResult.model_validate(proposal.proposal)
    profile = load_workbook_profile(profile_record)
    header_by_column = {
        column.source_column_id: column.header_path
        for sheet in profile.sheets
        for candidate in select_header_candidates(sheet.header_candidates)
        for column in candidate.columns
    }

    template: DocumentTemplate | None = None
    base_definition: TemplateDefinition | None = None
    provisional_version = next(
        (
            version
            for version in database.scalars(
                select(TemplateVersion).where(
                    TemplateVersion.status == TemplateStatus.ADMIN_REVIEW,
                    TemplateVersion.source == "hermes_provisional",
                )
            )
            if str(version.source_metadata.get("proposal_id")) == str(proposal.id)
        ),
        None,
    )
    candidate_template_id, candidate_template_version = _unmatched_template_candidate(match)
    if candidate_template_id is not None and candidate_template_version is not None:
        template = database.get(DocumentTemplate, candidate_template_id)
        base_version = database.scalar(
            select(TemplateVersion).where(
                TemplateVersion.template_id == candidate_template_id,
                TemplateVersion.version == candidate_template_version,
                TemplateVersion.status == TemplateStatus.PUBLISHED,
            )
        )
        if template is None or base_version is None:
            raise ProposalResolutionError("matched published template is unavailable")
        if any(
            item.status not in {TemplateStatus.PUBLISHED, TemplateStatus.DEPRECATED}
            and item.id != getattr(provisional_version, "id", None)
            for item in template.versions
        ):
            raise ProposalResolutionError("an editable template version already exists")
        base_definition = TemplateDefinition.model_validate(base_version.definition)
    else:
        if provisional_version is not None:
            template = provisional_version.template
        else:
            if template_code is None:
                raise ProposalResolutionError("template_code is required for a new template")
            existing = database.scalar(
                select(DocumentTemplate).where(DocumentTemplate.code == template_code)
            )
            if existing is not None:
                raise ProposalResolutionError("template code already exists")
            template = DocumentTemplate(code=template_code)
            database.add(template)

    bindings = list(base_definition.field_bindings) if base_definition else []
    refreshed_bindings: list[TemplateFieldBinding] = []
    for binding in bindings:
        field = database.scalar(
            select(SemanticField).where(SemanticField.code == binding.semantic_field_code)
        )
        refreshed_bindings.append(
            binding.model_copy(
                update={
                    "semantic_field_version": (
                        field.published_version
                        if field is not None and field.published_version is not None
                        else binding.semantic_field_version
                    )
                }
            )
        )
    bindings = refreshed_bindings
    binding_by_column = {binding.source_column_id: binding for binding in bindings}
    final_field_decisions = (
        [FieldDecision.model_validate(decision) for decision in field_decisions]
        if field_decisions is not None
        else result.field_decisions
    )
    for decision in final_field_decisions:
        if decision.action in {
            "IGNORE_COLUMN",
            "SEMANTIC_CONFLICT",
            "AMBIGUOUS",
        }:
            continue
        field_code = decision.semantic_field_code or decision.proposed_field_code
        if not field_code:
            raise ProposalResolutionError(f"{decision.action} requires a semantic field code")
        field = database.scalar(select(SemanticField).where(SemanticField.code == field_code))
        if field is None:
            if decision.action != "PROPOSE_NEW_FIELD":
                raise ProposalResolutionError(f"semantic field does not exist: {field_code}")
            if decision.layer is None or decision.data_type is None:
                raise ProposalResolutionError(f"new field metadata is incomplete: {field_code}")
            field = SemanticField(code=field_code)
            field.versions.append(
                SemanticFieldVersion(
                    version=1,
                    name=header_by_column.get(
                        decision.source_column_id,
                        [field_code],
                    )[-1],
                    description="由用户确认的 Hermes 建议生成",
                    layer=decision.layer,
                    data_type=decision.data_type,
                    unit_dimension=decision.unit,
                )
            )
            database.add(field)
            field_version = 1
        else:
            field_version = field.published_version or max(item.version for item in field.versions)
        binding_by_column[decision.source_column_id] = TemplateFieldBinding(
            source_column_id=decision.source_column_id,
            header_path=header_by_column.get(decision.source_column_id, [field_code]),
            semantic_field_code=field_code,
            semantic_field_version=field_version,
            role=decision.role,
            unit=decision.unit,
        )

    definition = TemplateDefinition(
        domain=domain,
        region_kind=base_definition.region_kind if base_definition else "table",
        record_type=record_type,
        record_grain=(
            record_grain
            or (result.record_grain.value if result.record_grain else None)
            or (base_definition.record_grain if base_definition else "one_row_per_record")
        ),
        field_bindings=list(binding_by_column.values()),
        data_row_rules=base_definition.data_row_rules if base_definition else [],
        exclusion_rules=base_definition.exclusion_rules if base_definition else [],
        metric_codes=base_definition.metric_codes if base_definition else [],
        identity_field_codes=(base_definition.identity_field_codes if base_definition else []),
    )
    next_version = max((item.version for item in template.versions), default=0) + 1
    template.versions.append(
        TemplateVersion(
            version=next_version,
            name=template_name,
            description="由用户确认的 Hermes 差异建议生成",
            status=TemplateStatus.USER_CONFIRMED,
            layout_fingerprint=match.layout_fingerprint,
            definition=definition.model_dump(mode="json"),
            source="hermes",
            source_metadata={
                "proposal_id": str(proposal.id),
                "confirmed_by": actor,
                "source_item_id": str(proposal.source_item_id),
                "approved_layout_plan": [
                    decision.model_dump(mode="json") for decision in result.layout_decisions
                ],
            },
        )
    )
    proposal.status = ProposalStatus.ACCEPTED
    proposal.resolution_comment = comment
    proposal.resolved_by_user_id = actor_user_id
    proposal.resolved_at = utcnow()
    if provisional_version is not None:
        provisional_version.status = TemplateStatus.DEPRECATED
    database.flush()
    return template
