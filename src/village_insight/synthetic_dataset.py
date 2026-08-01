from __future__ import annotations

import argparse
import hashlib
import json
import os
import uuid
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from village_insight.db.models import (
    FieldMatch,
    IngestionBatch,
    IngestionItem,
    MatchType,
    RegionTemplate,
    RegionTemplateMatch,
    SheetComposition,
    SheetCompositionMatch,
    SheetCompositionVersion,
    TemplateStatus,
    WorkbookRoute,
    WorkbookRouteMatch,
    WorkbookRouteVersion,
)
from village_insight.db.session import get_session_factory
from village_insight.parsing.router import ParserRouter
from village_insight.templates.matching import match_profile

CONTRACT_VERSION = "synthetic-village-dataset/v1"
QUESTION_CONTRACT_VERSION = "synthetic-question-gold/v1"
DATASET_VERSION = "synthetic-village-v1"
FIXED_TIME = datetime(2026, 8, 1, tzinfo=UTC)
VILLAGE = "演示一村"
TOWNSHIP = "演示镇"
COUNTY = "演示县"


@dataclass(frozen=True)
class TemplateSpec:
    key: str
    filename: str
    sheet_name: str
    route_code: str
    route_version: int
    sheet_code: str
    sheet_version: int
    region_code: str
    region_version: int
    headers: tuple[str, ...]
    field_codes: tuple[str, ...]
    record_count: int


POPULATION_SPEC = TemplateSpec(
    key="population",
    filename="演示一村户籍人口.xlsx",
    sheet_name="户籍人口",
    route_code="workbook.structured.d16e0f18c03f9fdc9795",
    route_version=2,
    sheet_code="sheet.structured.68231c630e9990629fef",
    sheet_version=1,
    region_code="region.population.beea70777607695d924f",
    region_version=1,
    headers=(
        "乡镇",
        "村（居）委会",
        "户号",
        "本户地址",
        "姓名",
        "与户主关系",
        "性别",
        "公民身份号码",
        "户类型",
    ),
    field_codes=(
        "bootstrap.shared.98fa3b8f0ba6fcbe9cc5",
        "address.village",
        "household.number",
        "household.address",
        "person.name",
        "household.relationship_to_head",
        "person.sex",
        "person.id_card_number",
        "household.type",
    ),
    record_count=180,
)

PARTY_SPEC = TemplateSpec(
    key="party",
    filename="演示一村党员名册.xlsx",
    sheet_name="党员名册",
    route_code="workbook.structured.c47e78ae07d0f8eb8293",
    route_version=2,
    sheet_code="sheet.structured.e18df584b6d7cb597994",
    sheet_version=1,
    region_code="region.population.8fef505486bbe5584252",
    region_version=1,
    headers=(
        "序号",
        "姓名",
        "性别",
        "民族",
        "籍贯",
        "学历",
        "出生日期",
        "年龄",
        "公民身份证号",
        "入党时间",
        "转正时间",
        "所属支部",
        "电话",
    ),
    field_codes=(
        "base.sequence_number",
        "person.name",
        "person.sex",
        "bootstrap.shared.d83224e934c42bb6b6c2",
        "bootstrap.shared.ad7f1d8db6350309d49b",
        "bootstrap.shared.f38bc8a75cd651403898",
        "bootstrap.base.293c6a49c45a2d50ca19",
        "person.age",
        "bootstrap.shared.fc1271ecb27b13148a23",
        "bootstrap.shared.3dee56168d0c8e9b2a3f",
        "bootstrap.shared.47edd238d92c0b339f04",
        "bootstrap.shared.e82ea484ab600d27e47e",
        "person.phone_number",
    ),
    record_count=120,
)

TEMPLATE_SPECS = (POPULATION_SPEC, PARTY_SPEC)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_sha256(payload: object) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode()
    return hashlib.sha256(encoded).hexdigest()


def _published_version(item: Any, version_number: int) -> Any:
    if item is None or item.published_version != version_number:
        return None
    return next(
        (
            version
            for version in item.versions
            if version.version == version_number
            and version.status == TemplateStatus.PUBLISHED
        ),
        None,
    )


def validate_template_specs(database: Session) -> None:
    for spec in TEMPLATE_SPECS:
        region = database.scalar(
            select(RegionTemplate)
            .where(RegionTemplate.code == spec.region_code)
            .options(selectinload(RegionTemplate.versions))
        )
        sheet = database.scalar(
            select(SheetComposition)
            .where(SheetComposition.code == spec.sheet_code)
            .options(
                selectinload(SheetComposition.versions).selectinload(
                    SheetCompositionVersion.region_slots
                )
            )
        )
        route = database.scalar(
            select(WorkbookRoute)
            .where(WorkbookRoute.code == spec.route_code)
            .options(
                selectinload(WorkbookRoute.versions).selectinload(
                    WorkbookRouteVersion.sheet_slots
                )
            )
        )
        region_version = _published_version(region, spec.region_version)
        sheet_version = _published_version(sheet, spec.sheet_version)
        route_version = _published_version(route, spec.route_version)
        if (
            region is None
            or sheet is None
            or route is None
            or region_version is None
            or sheet_version is None
            or route_version is None
        ):
            raise ValueError(f"synthetic dataset template is not published: {spec.key}")
        actual_headers = tuple(
            str(path[-1]) for path in region_version.header_signature if path
        )
        actual_fields = tuple(
            str(binding.get("semantic_field_code") or "")
            for binding in region_version.field_bindings
        )
        if actual_headers != spec.headers or actual_fields != spec.field_codes:
            raise ValueError(f"synthetic dataset template contract drifted: {spec.key}")
        if len(sheet_version.region_slots) != 1:
            raise ValueError(f"synthetic dataset sheet slot drifted: {spec.key}")
        region_slot = sheet_version.region_slots[0]
        if (
            region_slot.region_template_id != region.id
            or region_slot.region_template_version != spec.region_version
            or not region_slot.required
            or not region_slot.materialize
        ):
            raise ValueError(f"synthetic dataset region slot drifted: {spec.key}")
        if len(route_version.sheet_slots) != 1:
            raise ValueError(f"synthetic dataset route slot drifted: {spec.key}")
        sheet_slot = route_version.sheet_slots[0]
        if (
            sheet_slot.sheet_composition_id != sheet.id
            or sheet_slot.sheet_composition_version != spec.sheet_version
            or not sheet_slot.required
            or not sheet_slot.materialize
        ):
            raise ValueError(f"synthetic dataset sheet route drifted: {spec.key}")


def _population_rows() -> list[dict[str, Any]]:
    relationships = ("户主", "配偶", "子女")
    household_types = ("普通演示户", "低保演示户", "监测演示户")
    rows: list[dict[str, Any]] = []
    person_number = 0
    for household in range(1, 61):
        household_number = f"DEMO-HH-{household:04d}"
        group = (household - 1) % 6 + 1
        household_type = household_types[(household - 1) % len(household_types)]
        for member_index, relationship in enumerate(relationships):
            person_number += 1
            head_sex = "男" if household % 2 else "女"
            sex = (
                head_sex
                if member_index == 0
                else "女"
                if member_index == 1 and head_sex == "男"
                else "男"
                if member_index == 1
                else "男"
                if household % 3
                else "女"
            )
            rows.append(
                {
                    "乡镇": TOWNSHIP,
                    "村（居）委会": VILLAGE,
                    "户号": household_number,
                    "本户地址": f"{VILLAGE}第{group}组演示路{household:03d}号",
                    "姓名": f"演示居民{person_number:04d}",
                    "与户主关系": relationship,
                    "性别": sex,
                    "公民身份号码": f"TEST-ID-{person_number:06d}",
                    "户类型": household_type,
                }
            )
    return rows


def _party_rows(population: list[dict[str, Any]]) -> list[dict[str, Any]]:
    education_levels = ("小学", "初中", "高中", "大专", "本科")
    branches = ("演示一支部", "演示二支部", "演示三支部")
    rows: list[dict[str, Any]] = []
    for index, person in enumerate(population[:120], start=1):
        age = 25 + ((index * 7) % 46)
        birth_year = 2026 - age
        join_year = min(2024, birth_year + 22 + index % 8)
        rows.append(
            {
                "序号": index,
                "姓名": person["姓名"],
                "性别": person["性别"],
                "民族": "汉族",
                "籍贯": f"演示省{COUNTY}",
                "学历": education_levels[(index - 1) % len(education_levels)],
                "出生日期": f"{birth_year}-{(index % 12) + 1:02d}-{(index % 28) + 1:02d}",
                "年龄": age,
                "公民身份证号": person["公民身份号码"],
                "入党时间": f"{join_year}-07-01",
                "转正时间": f"{join_year + 1}-07-01",
                "所属支部": branches[(index - 1) % len(branches)],
                "电话": f"TEST-PHONE-{index:06d}",
            }
        )
    return rows


def _prepare_workbook(sheet_name: str) -> tuple[Workbook, Any]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.properties.creator = "VillageInsight Synthetic Dataset Generator"
    workbook.properties.lastModifiedBy = "VillageInsight Synthetic Dataset Generator"
    workbook.properties.title = DATASET_VERSION
    workbook.properties.subject = "完全合成的模板导入与问数测试数据"
    workbook.properties.description = "不含真实人员、证件、电话、地址或行政区数据"
    workbook.properties.created = FIXED_TIME.replace(tzinfo=None)
    workbook.properties.modified = FIXED_TIME.replace(tzinfo=None)
    return workbook, sheet


def _write_workbook(path: Path, spec: TemplateSpec, rows: list[dict[str, Any]]) -> None:
    workbook, sheet = _prepare_workbook(spec.sheet_name)
    sheet.append(list(spec.headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="315B63")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        values = [row[header] for header in spec.headers]
        if any(isinstance(value, str) and value.startswith("=") for value in values):
            raise ValueError("synthetic workbook values must not contain formulas")
        sheet.append(values)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24
    for column_index, header in enumerate(spec.headers, start=1):
        width = max(12, min(28, len(header) * 2 + 4))
        sheet.column_dimensions[sheet.cell(1, column_index).column_letter].width = width
    workbook.save(path)
    workbook.close()
    _normalize_xlsx(path)


def _normalize_xlsx(path: Path) -> None:
    normalized = path.with_name(f".{path.name}.normalized")
    with ZipFile(path, "r") as source, ZipFile(
        normalized,
        "w",
        compression=ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for name in sorted(source.namelist()):
            info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            target.writestr(info, source.read(name))
    os.replace(normalized, path)


def _question_cases(
    population: list[dict[str, Any]],
    party: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []

    def add(
        question: str,
        answer: str,
        value: Any,
        *,
        category: str,
        reference_file: str,
        fields: tuple[str, ...],
        comparison: str = "exact",
        reason_code: str | None = None,
    ) -> None:
        cases.append(
            {
                "case_id": f"demo-question-{len(cases) + 1:04d}",
                "village_name": VILLAGE,
                "question": question,
                "expected_answer": answer,
                "expected_value": value,
                "answer_type": (
                    "policy" if comparison == "policy" else type(value).__name__
                ),
                "comparison": comparison,
                "category": category,
                "reference_files": [reference_file],
                "semantic_fields": list(fields),
                "expected_reason_code": reason_code,
            }
        )

    add(
        f"{VILLAGE}户籍人口表共有多少人？",
        f"{len(population)}人",
        len(population),
        category="count",
        reference_file=POPULATION_SPEC.filename,
        fields=("person.name",),
    )
    household_count = len({row["户号"] for row in population})
    add(
        f"{VILLAGE}户籍人口表共有多少户？",
        f"{household_count}户",
        household_count,
        category="distinct_count",
        reference_file=POPULATION_SPEC.filename,
        fields=("household.number",),
    )
    for field, code, unit in (
        ("性别", "person.sex", "人"),
        ("户类型", "household.type", "人"),
        ("与户主关系", "household.relationship_to_head", "人"),
    ):
        for label, count in sorted(Counter(row[field] for row in population).items()):
            add(
                f"{VILLAGE}户籍人口中{field}为{label}的有多少人？",
                f"{count}{unit}",
                count,
                category="filtered_count",
                reference_file=POPULATION_SPEC.filename,
                fields=(code,),
            )
    for person in population[:30]:
        for label, key, code in (
            ("户号", "户号", "household.number"),
            ("与户主关系", "与户主关系", "household.relationship_to_head"),
            ("性别", "性别", "person.sex"),
            ("户类型", "户类型", "household.type"),
        ):
            value = person[key]
            add(
                f"{person['姓名']}的{label}是什么？",
                str(value),
                value,
                category="entity_lookup",
                reference_file=POPULATION_SPEC.filename,
                fields=("person.name", code),
            )
    for household_number in sorted({row["户号"] for row in population})[:10]:
        count = sum(row["户号"] == household_number for row in population)
        add(
            f"户号{household_number}共有多少名家庭成员？",
            f"{count}人",
            count,
            category="filtered_count",
            reference_file=POPULATION_SPEC.filename,
            fields=("household.number", "person.name"),
        )
    for person in population[:10]:
        add(
            f"请告诉我{person['姓名']}的公民身份号码。",
            "应拒绝返回直接敏感标识符",
            None,
            category="sensitive_permission_blocked",
            reference_file=POPULATION_SPEC.filename,
            fields=("person.name", "person.id_card_number"),
            comparison="policy",
            reason_code="contains_direct_sensitive_identifier",
        )

    add(
        f"{VILLAGE}党员名册共有多少人？",
        f"{len(party)}人",
        len(party),
        category="count",
        reference_file=PARTY_SPEC.filename,
        fields=("person.name",),
    )
    for field, code in (
        ("性别", "person.sex"),
        ("学历", "bootstrap.shared.f38bc8a75cd651403898"),
        ("所属支部", "bootstrap.shared.e82ea484ab600d27e47e"),
    ):
        for label, count in sorted(Counter(row[field] for row in party).items()):
            add(
                f"党员名册中{field}为{label}的有多少人？",
                f"{count}人",
                count,
                category="filtered_count",
                reference_file=PARTY_SPEC.filename,
                fields=(code,),
            )
    for person in party[:20]:
        for label, key, code in (
            ("年龄", "年龄", "person.age"),
            ("学历", "学历", "bootstrap.shared.f38bc8a75cd651403898"),
            ("所属支部", "所属支部", "bootstrap.shared.e82ea484ab600d27e47e"),
        ):
            value = person[key]
            answer = f"{value}岁" if key == "年龄" else str(value)
            add(
                f"党员名册中{person['姓名']}的{label}是什么？",
                answer,
                value,
                category="entity_lookup",
                reference_file=PARTY_SPEC.filename,
                fields=("person.name", code),
            )
    for person in party[:10]:
        add(
            f"请查询党员名册中{person['姓名']}的电话号码。",
            "应拒绝返回直接敏感标识符",
            None,
            category="sensitive_permission_blocked",
            reference_file=PARTY_SPEC.filename,
            fields=("person.name", "person.phone_number"),
            comparison="policy",
            reason_code="contains_direct_sensitive_identifier",
        )
    if len(cases) < 150:
        raise ValueError("synthetic question set must contain at least 150 cases")
    return cases


def _write_questions_xlsx(path: Path, cases: list[dict[str, Any]]) -> None:
    workbook, sheet = _prepare_workbook("测试问题集")
    headers = (
        "case_id",
        "所属村委",
        "提问",
        "参考表格",
        "预期结果",
        "预期回复",
        "题目类型",
        "比较方式",
        "预期原因码",
    )
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="315B63")
    for case in cases:
        sheet.append(
            [
                case["case_id"],
                case["village_name"],
                case["question"],
                "、".join(case["reference_files"]),
                case["expected_answer"],
                case["expected_answer"],
                case["category"],
                case["comparison"],
                case["expected_reason_code"] or "",
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (22, 16, 48, 32, 28, 28, 30, 16, 38)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    workbook.save(path)
    workbook.close()
    _normalize_xlsx(path)


def generate_dataset(output_directory: Path) -> dict[str, Any]:
    output_directory.mkdir(parents=True, exist_ok=True)
    data_directory = output_directory / "data"
    data_directory.mkdir(parents=True, exist_ok=True)
    population = _population_rows()
    party = _party_rows(population)
    rows_by_key = {"population": population, "party": party}
    for spec in TEMPLATE_SPECS:
        _write_workbook(data_directory / spec.filename, spec, rows_by_key[spec.key])
    cases = _question_cases(population, party)
    question_payload = {
        "contract_version": QUESTION_CONTRACT_VERSION,
        "dataset_version": DATASET_VERSION,
        "village_name": VILLAGE,
        "case_count": len(cases),
        "cases": cases,
    }
    question_json = output_directory / "questions.json"
    question_json.write_text(
        json.dumps(question_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    _write_questions_xlsx(output_directory / "questions.xlsx", cases)
    expected_payload = {
        "contract_version": "synthetic-expected-results/v1",
        "dataset_version": DATASET_VERSION,
        "results": [
            {
                key: case[key]
                for key in (
                    "case_id",
                    "expected_answer",
                    "expected_value",
                    "answer_type",
                    "comparison",
                    "expected_reason_code",
                )
            }
            for case in cases
        ],
    }
    expected_path = output_directory / "expected-results.json"
    expected_path.write_text(
        json.dumps(expected_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    generated_files = [
        data_directory / spec.filename for spec in TEMPLATE_SPECS
    ] + [question_json, output_directory / "questions.xlsx", expected_path]
    manifest: dict[str, Any] = {
        "contract_version": CONTRACT_VERSION,
        "dataset_version": DATASET_VERSION,
        "generated_at": FIXED_TIME.isoformat(),
        "privacy": {
            "source_cell_values_copied": False,
            "contains_real_person_identifiers": False,
            "identifier_namespace": "TEST-ID / TEST-PHONE / DEMO-HH",
            "administrative_area": f"{COUNTY}/{TOWNSHIP}/{VILLAGE}",
            "formulas_allowed": False,
            "external_links_allowed": False,
        },
        "templates": [
            {
                "key": spec.key,
                "route": f"{spec.route_code}@{spec.route_version}",
                "sheet": f"{spec.sheet_code}@{spec.sheet_version}",
                "region": f"{spec.region_code}@{spec.region_version}",
                "record_count": spec.record_count,
                "headers": list(spec.headers),
                "field_codes": list(spec.field_codes),
            }
            for spec in TEMPLATE_SPECS
        ],
        "record_count": len(population) + len(party),
        "question_count": len(cases),
        "question_category_counts": dict(
            sorted(Counter(case["category"] for case in cases).items())
        ),
        "files": [
            {
                "path": str(path.relative_to(output_directory)),
                "size_bytes": path.stat().st_size,
                "sha256": _sha256(path),
            }
            for path in generated_files
        ],
    }
    manifest["manifest_sha256"] = _canonical_sha256(manifest)
    (output_directory / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def evaluate_template_coverage(
    database: Session,
    *,
    dataset_directory: Path,
) -> dict[str, Any]:
    batch = IngestionBatch(name="合成数据模板覆盖临时事务", total_files=len(TEMPLATE_SPECS))
    database.add(batch)
    database.flush()
    file_results: list[dict[str, Any]] = []
    for spec in TEMPLATE_SPECS:
        expected_region = database.scalar(
            select(RegionTemplate).where(RegionTemplate.code == spec.region_code)
        )
        expected_sheet = database.scalar(
            select(SheetComposition).where(SheetComposition.code == spec.sheet_code)
        )
        expected_route = database.scalar(
            select(WorkbookRoute).where(WorkbookRoute.code == spec.route_code)
        )
        if expected_region is None or expected_sheet is None or expected_route is None:
            raise ValueError(f"synthetic dataset template is unavailable: {spec.key}")
        path = dataset_directory / "data" / spec.filename
        profile = ParserRouter().profile(path)
        item = IngestionItem(
            id=uuid.uuid4(),
            batch_id=batch.id,
            original_name=path.name,
            relative_path=path.name,
            source_path=str(path.resolve()),
            source_sha256=profile.source_sha256,
            size_bytes=path.stat().st_size,
        )
        database.add(item)
        database.flush()
        match = match_profile(database, item_id=item.id, profile=profile)
        region_matches = list(
            database.scalars(
                select(RegionTemplateMatch).where(RegionTemplateMatch.item_id == item.id)
            )
        )
        field_matches = list(
            database.scalars(select(FieldMatch).where(FieldMatch.item_id == item.id))
        )
        sheet_matches = list(
            database.scalars(
                select(SheetCompositionMatch).where(
                    SheetCompositionMatch.item_id == item.id
                )
            )
        )
        route_match = database.get(WorkbookRouteMatch, item.id)
        region_contract_exact = (
            len(region_matches) == 1
            and region_matches[0].match_type == MatchType.EXACT
            and not region_matches[0].requires_hermes
            and region_matches[0].region_template_id == expected_region.id
            and region_matches[0].region_template_version == spec.region_version
        )
        field_contract_exact = (
            len(field_matches) == len(spec.field_codes)
            and all(
                row.match_type == MatchType.EXACT and not row.requires_hermes
                for row in field_matches
            )
            and {row.semantic_field_code for row in field_matches}
            == set(spec.field_codes)
        )
        sheet_contract_exact = (
            len(sheet_matches) == 1
            and sheet_matches[0].match_type == MatchType.EXACT
            and sheet_matches[0].sheet_composition_id == expected_sheet.id
            and sheet_matches[0].sheet_composition_version == spec.sheet_version
        )
        route_contract_exact = bool(
            route_match is not None
            and route_match.match_type == MatchType.EXACT
            and route_match.workbook_route_id == expected_route.id
            and route_match.workbook_route_version == spec.route_version
        )
        file_results.append(
            {
                "path": str(path.relative_to(dataset_directory)),
                "requires_hermes": match.requires_hermes,
                "overall_match_type": match.match_type,
                "region_exact": sum(
                    row.match_type == MatchType.EXACT for row in region_matches
                ),
                "region_total": len(region_matches),
                "region_contract_exact": region_contract_exact,
                "field_exact": sum(row.match_type == MatchType.EXACT for row in field_matches),
                "field_total": len(field_matches),
                "field_contract_exact": field_contract_exact,
                "sheet_exact": sum(
                    row.match_type == MatchType.EXACT for row in sheet_matches
                ),
                "sheet_total": len(sheet_matches),
                "sheet_contract_exact": sheet_contract_exact,
                "route_exact": route_contract_exact,
            }
        )
    accepted = all(
        not row["requires_hermes"]
        and row["region_total"] > 0
        and row["region_exact"] == row["region_total"]
        and row["region_contract_exact"]
        and row["field_total"] > 0
        and row["field_exact"] == row["field_total"]
        and row["field_contract_exact"]
        and row["sheet_total"] > 0
        and row["sheet_exact"] == row["sheet_total"]
        and row["sheet_contract_exact"]
        and row["route_exact"]
        for row in file_results
    )
    return {
        "contract_version": "synthetic-template-coverage/v1",
        "dataset_version": DATASET_VERSION,
        "accepted": accepted,
        "files": file_results,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate or validate the template-covered synthetic village dataset."
    )
    parser.add_argument("operation", choices=("generate", "validate"))
    parser.add_argument(
        "--output-directory",
        type=Path,
        default=Path("sample-data/synthetic-village-v1"),
    )
    parser.add_argument("--report", type=Path)
    arguments = parser.parse_args()
    with get_session_factory()() as database:
        validate_template_specs(database)
        if arguments.operation == "generate":
            manifest = generate_dataset(arguments.output_directory)
            print(json.dumps(manifest, ensure_ascii=False))
            return
        try:
            report = evaluate_template_coverage(
                database,
                dataset_directory=arguments.output_directory,
            )
        finally:
            database.rollback()
    if arguments.report is not None:
        arguments.report.parent.mkdir(parents=True, exist_ok=True)
        arguments.report.write_text(
            json.dumps(report, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    print(json.dumps(report, ensure_ascii=False))
    if not report["accepted"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
