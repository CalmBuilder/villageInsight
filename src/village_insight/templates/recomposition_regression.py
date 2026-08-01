from __future__ import annotations

import argparse
import copy
import hashlib
import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session

from village_insight.db.models import (
    FieldMatch,
    IngestionBatch,
    IngestionItem,
    RegionTemplateMatch,
    SheetCompositionMatch,
    WorkbookRouteMatch,
)
from village_insight.db.session import get_session_factory
from village_insight.parsing.candidates import select_header_candidates
from village_insight.parsing.router import ParserRouter
from village_insight.templates.four_layer_seeds import _looks_like_observed_value
from village_insight.templates.matching import (
    layout_fingerprint,
    match_profile,
)

CONTRACT_VERSION = "recomposed-template-regression/v1"
DEFAULT_POPULATION_SOURCE = Path(
    "docs/datafiles/所有村/官庄村村民委员会/2025年官庄村人口明细表8.20日 - 区分各小组.xlsx"
)
DEFAULT_CROP_SOURCE = Path("docs/datafiles/所有村/群慧村/2024年农作物登记.xlsx")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _synthetic_value(value: object, *, case_number: int, row: int, column: int) -> object:
    if value in (None, ""):
        return value
    if isinstance(value, str) and value.startswith("="):
        return value
    if isinstance(value, bool):
        return (case_number + row + column) % 2 == 0
    if isinstance(value, datetime):
        return value + timedelta(days=case_number)
    if isinstance(value, date):
        return value + timedelta(days=case_number)
    if isinstance(value, int):
        return case_number * 100_000 + row * 100 + column
    if isinstance(value, float):
        return round(case_number * 1000 + row + column / 100, 2)
    if isinstance(value, str):
        compact = "".join(value.split())
        numeric = compact.replace(",", "")
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", numeric):
            return str(case_number * 100_000 + row * 100 + column)
        if re.fullmatch(r"\d{4}[-/.年]\d{1,2}(?:[-/.月]\d{1,2}日?)?", compact):
            return f"2099-01-{(row % 28) + 1:02d}"
    return f"测试同名人员-{row % 5}" if column == 1 else f"样例{case_number}-{row}-{column}"


def _safe_header_value(value: object, *, column: int) -> object:
    """Keep semantic headers while refusing to duplicate source data as headers."""
    if value in (None, ""):
        return value
    if isinstance(value, str) and value.startswith("="):
        return str(900_000_000_000_000_000 + column)
    if _looks_like_observed_value([str(value)]):
        # Keep the structural signal that this is an observed value, without
        # copying the source value into a regression artifact.
        return str(900_000_000_000_000_000 + column)
    return value


def _copy_cell(
    source: Cell,
    target: Cell,
    *,
    preserve_value: bool,
    case_number: int,
) -> None:
    target.value = (
        _safe_header_value(source.value, column=source.column)
        if preserve_value
        else _synthetic_value(
            source.value,
            case_number=case_number,
            row=source.row,
            column=source.column,
        )
    )
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy.copy(source.alignment)
    if source.protection:
        target.protection = copy.copy(source.protection)


def _copy_sheet(
    *,
    source: Worksheet,
    target: Worksheet,
    header_end: int,
    case_number: int,
    max_body_rows: int | None = None,
) -> None:
    max_row = (
        source.max_row if max_body_rows is None else min(source.max_row, header_end + max_body_rows)
    )
    for row in source.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=source.max_column,
    ):
        for cell in row:
            _copy_cell(
                cell,
                target.cell(row=cell.row, column=cell.column),
                preserve_value=cell.row <= header_end,
                case_number=case_number,
            )
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden
    for index, dimension in source.row_dimensions.items():
        if index <= max_row:
            target.row_dimensions[index].height = dimension.height
            target.row_dimensions[index].hidden = dimension.hidden
    for merged_range in source.merged_cells.ranges:
        if merged_range.max_row <= max_row:
            target.merge_cells(str(merged_range))
    target.freeze_panes = source.freeze_panes
    target.sheet_view.showGridLines = source.sheet_view.showGridLines


def _header_end_by_sheet(path: Path) -> dict[int, int]:
    profile = ParserRouter().profile(path)
    detected = {
        sheet.index: max(
            (
                max(candidate.header_rows)
                for candidate in select_header_candidates(sheet.header_candidates)
            ),
            default=1,
        )
        for sheet in profile.sheets
    }
    with get_session_factory()() as database:
        row = (
            database.execute(
                text(
                    """
                SELECT layout_plan
                FROM approved_import_plans
                WHERE source_sha256 = :source_sha256
                ORDER BY revision DESC
                LIMIT 1
                """
                ),
                {"source_sha256": profile.source_sha256},
            )
            .mappings()
            .first()
        )
    if row is None:
        return detected
    approved: dict[int, int] = {}
    for decision in (row["layout_plan"] or {}).get("decisions", []):
        if not decision.get("materialize", True):
            continue
        match = re.search(r":sheet:(\d+):", str(decision.get("region_candidate_id") or ""))
        data_start = decision.get("data_start_row")
        if match is None or not isinstance(data_start, int) or data_start <= 1:
            continue
        sheet_index = int(match.group(1))
        approved[sheet_index] = min(
            approved.get(sheet_index, data_start - 1),
            data_start - 1,
        )
    return {**detected, **approved}


def _approved_mapped_sheet_indexes(path: Path) -> set[int]:
    source_sha256 = _sha256(path)
    with get_session_factory()() as database:
        row = (
            database.execute(
                text(
                    """
                SELECT layout_plan, field_mappings
                FROM approved_import_plans
                WHERE source_sha256 = :source_sha256
                ORDER BY revision DESC
                LIMIT 1
                """
                ),
                {"source_sha256": source_sha256},
            )
            .mappings()
            .first()
        )
    if row is None:
        return set()
    indexes: set[int] = set()
    mappings = list(row["field_mappings"] or [])
    for decision in (row["layout_plan"] or {}).get("decisions", []):
        mappings.extend(decision.get("field_mappings", []))
    for mapping in mappings:
        match = re.search(r":sheet:(\d+):", str(mapping.get("region_id") or ""))
        if match is not None:
            indexes.add(int(match.group(1)))
    return indexes


def generate_recomposed_workbooks(
    *,
    population_source: Path,
    crop_source: Path,
    output_directory: Path,
) -> dict[str, Any]:
    output_directory.mkdir(parents=True, exist_ok=True)
    population_headers = _header_end_by_sheet(population_source)
    crop_headers = _header_end_by_sheet(crop_source)
    population = load_workbook(population_source, data_only=False)
    crop = load_workbook(crop_source, data_only=False)
    rows: list[dict[str, Any]] = []
    try:
        approved_population_indexes = _approved_mapped_sheet_indexes(population_source)
        population_indexes = [
            index
            for index in range(1, min(len(population.worksheets), 13))
            if not approved_population_indexes or index in approved_population_indexes
        ]
        recipes: list[list[tuple[str, int]]] = []
        for case_index in range(16):
            indexes = [
                population_indexes[case_index % len(population_indexes)],
                population_indexes[(case_index * 3 + 4) % len(population_indexes)],
                population_indexes[(case_index * 5 + 7) % len(population_indexes)],
            ]
            recipes.append([("population", index) for index in dict.fromkeys(indexes)])
        for case_index in range(5):
            recipes.append(
                [
                    ("crop", 0),
                    (
                        "population",
                        population_indexes[(case_index * 2 + 1) % len(population_indexes)],
                    ),
                    (
                        "population",
                        population_indexes[(case_index * 4 + 6) % len(population_indexes)],
                    ),
                ]
            )

        for case_number, recipe in enumerate(recipes, start=1):
            workbook = Workbook()
            workbook.remove(workbook.active)
            sources: list[dict[str, Any]] = []
            for ordinal, (source_kind, sheet_index) in enumerate(recipe):
                source_workbook = population if source_kind == "population" else crop
                header_map = population_headers if source_kind == "population" else crop_headers
                source_sheet = source_workbook.worksheets[sheet_index]
                target = workbook.create_sheet(f"重组{case_number:02d}-{ordinal + 1}")
                _copy_sheet(
                    source=source_sheet,
                    target=target,
                    header_end=header_map[sheet_index],
                    case_number=case_number,
                )
                sources.append(
                    {
                        "source_kind": source_kind,
                        "source_path": str(
                            population_source if source_kind == "population" else crop_source
                        ),
                        "source_sheet_index": sheet_index,
                        "source_sheet_name": source_sheet.title,
                        "header_end": header_map[sheet_index],
                    }
                )
            path = output_directory / f"重组模板回归-{case_number:02d}.xlsx"
            workbook.save(path)
            workbook.close()
            rows.append(
                {
                    "path": str(path.resolve()),
                    "sha256": _sha256(path),
                    "challenge": False,
                    "sources": sources,
                }
            )

        challenge_number = len(recipes) + 1
        workbook = Workbook()
        workbook.remove(workbook.active)
        challenge_indexes = (
            population_indexes[0],
            population_indexes[len(population_indexes) // 2],
            population_indexes[-1],
        )
        for ordinal, sheet_index in enumerate(challenge_indexes):
            source_sheet = population.worksheets[sheet_index]
            target = workbook.create_sheet(f"挑战-{ordinal + 1}")
            _copy_sheet(
                source=source_sheet,
                target=target,
                header_end=population_headers[sheet_index],
                case_number=challenge_number,
            )
            if ordinal == 0:
                new_column = source_sheet.max_column + 1
                target.cell(
                    row=population_headers[sheet_index],
                    column=new_column,
                    value="新增回归字段",
                )
                for row in range(population_headers[sheet_index] + 1, target.max_row + 1):
                    target.cell(row=row, column=new_column, value=f"新增值-{row}")
        challenge_path = output_directory / f"重组模板回归-{challenge_number:02d}-新增字段.xlsx"
        workbook.save(challenge_path)
        workbook.close()
        rows.append(
            {
                "path": str(challenge_path.resolve()),
                "sha256": _sha256(challenge_path),
                "challenge": True,
                "sources": [
                    {
                        "source_kind": "population",
                        "source_path": str(population_source),
                        "source_sheet_index": sheet_index,
                        "source_sheet_name": population.worksheets[sheet_index].title,
                        "header_end": population_headers[sheet_index],
                    }
                    for sheet_index in challenge_indexes
                ],
            }
        )
    finally:
        population.close()
        crop.close()

    source_hashes = {_sha256(population_source), _sha256(crop_source)}
    return {
        "contract_version": CONTRACT_VERSION,
        "population_source": str(population_source.resolve()),
        "crop_source": str(crop_source.resolve()),
        "file_count": len(rows),
        "source_hash_collision_count": sum(row["sha256"] in source_hashes for row in rows),
        "files": rows,
    }


def evaluate_recomposed_workbooks(
    database: Session,
    *,
    manifest: dict[str, Any],
    known_route_fingerprints: set[str],
) -> dict[str, Any]:
    batch = IngestionBatch(name="重组模板回归临时事务", total_files=len(manifest["files"]))
    database.add(batch)
    database.flush()
    file_rows: list[dict[str, Any]] = []
    for manifest_row in manifest["files"]:
        path = Path(manifest_row["path"])
        profile = ParserRouter().profile(path)
        fingerprint = layout_fingerprint(profile)
        item = IngestionItem(
            batch_id=batch.id,
            original_name=path.name,
            relative_path=path.name,
            source_path=str(path),
            source_sha256=profile.source_sha256,
            size_bytes=path.stat().st_size,
        )
        database.add(item)
        database.flush()
        match = match_profile(database, item_id=item.id, profile=profile)
        region_counts: dict[str, int] = {
            str(match_type): int(count)
            for match_type, count in database.execute(
                select(RegionTemplateMatch.match_type, func.count())
                .where(RegionTemplateMatch.item_id == item.id)
                .group_by(RegionTemplateMatch.match_type)
            ).tuples()
        }
        field_counts: dict[str, int] = {
            str(match_type): int(count)
            for match_type, count in database.execute(
                select(FieldMatch.match_type, func.count())
                .where(FieldMatch.item_id == item.id)
                .group_by(FieldMatch.match_type)
            ).tuples()
        }
        sheet_counts: dict[str, int] = {
            str(match_type): int(count)
            for match_type, count in database.execute(
                select(SheetCompositionMatch.match_type, func.count())
                .where(SheetCompositionMatch.item_id == item.id)
                .group_by(SheetCompositionMatch.match_type)
            ).tuples()
        }
        workbook_route = database.scalar(
            select(WorkbookRouteMatch).where(WorkbookRouteMatch.item_id == item.id)
        )
        file_rows.append(
            {
                "path": str(path),
                "sha256": profile.source_sha256,
                "challenge": bool(manifest_row["challenge"]),
                "layout_fingerprint": fingerprint,
                "known_workbook_route_fingerprint": fingerprint in known_route_fingerprints,
                "match_type": match.match_type,
                "requires_hermes": match.requires_hermes,
                "region_counts": region_counts,
                "field_counts": field_counts,
                "sheet_counts": sheet_counts,
                "workbook_route_match_type": (
                    workbook_route.match_type if workbook_route is not None else None
                ),
            }
        )

    total_files = len(file_rows)
    no_hermes_files = sum(not row["requires_hermes"] for row in file_rows)
    total_regions = sum(sum(row["region_counts"].values()) for row in file_rows)
    exact_regions = sum(row["region_counts"].get("exact", 0) for row in file_rows)
    total_fields = sum(sum(row["field_counts"].values()) for row in file_rows)
    exact_fields = sum(row["field_counts"].get("exact", 0) for row in file_rows)
    challenge_rows = [row for row in file_rows if row["challenge"]]
    return {
        "contract_version": CONTRACT_VERSION,
        "file_count": total_files,
        "metrics": {
            "source_hash_collision_count": manifest["source_hash_collision_count"],
            "known_workbook_route_fingerprint_count": sum(
                row["known_workbook_route_fingerprint"] for row in file_rows
            ),
            "no_hermes_file_count": no_hermes_files,
            "no_hermes_file_basis_points": (
                round(10_000 * no_hermes_files / total_files) if total_files else 0
            ),
            "exact_region_count": exact_regions,
            "total_region_count": total_regions,
            "exact_region_basis_points": (
                round(10_000 * exact_regions / total_regions) if total_regions else 0
            ),
            "exact_field_count": exact_fields,
            "total_field_count": total_fields,
            "exact_field_basis_points": (
                round(10_000 * exact_fields / total_fields) if total_fields else 0
            ),
        },
        "acceptance": {
            "no_source_hash_reuse": manifest["source_hash_collision_count"] == 0,
            "no_known_workbook_route_reuse": not any(
                row["known_workbook_route_fingerprint"] for row in file_rows
            ),
            "no_hermes_file_rate_at_least_95_percent": (
                total_files > 0 and no_hermes_files / total_files >= 0.95
            ),
            "exact_region_rate_at_least_95_percent": (
                total_regions > 0 and exact_regions / total_regions >= 0.95
            ),
            "exact_field_rate_at_least_95_percent": (
                total_fields > 0 and exact_fields / total_fields >= 0.95
            ),
            "challenge_files_require_hermes": bool(challenge_rows)
            and all(row["requires_hermes"] for row in challenge_rows),
        },
        "files": file_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate and evaluate recomposed real-workbook template fixtures."
    )
    parser.add_argument(
        "operation",
        choices=("generate", "evaluate"),
    )
    parser.add_argument("--population-source", type=Path, default=DEFAULT_POPULATION_SOURCE)
    parser.add_argument("--crop-source", type=Path, default=DEFAULT_CROP_SOURCE)
    parser.add_argument("--output-directory", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--known-routes", type=Path)
    parser.add_argument("--report", type=Path)
    arguments = parser.parse_args()
    if arguments.operation == "generate":
        manifest = generate_recomposed_workbooks(
            population_source=arguments.population_source,
            crop_source=arguments.crop_source,
            output_directory=arguments.output_directory,
        )
        arguments.manifest.parent.mkdir(parents=True, exist_ok=True)
        arguments.manifest.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(
            json.dumps(
                {key: manifest[key] for key in ("file_count", "source_hash_collision_count")}
            )
        )
        return
    if arguments.known_routes is None or arguments.report is None:
        parser.error("--known-routes and --report are required for evaluate")
    manifest = json.loads(arguments.manifest.read_text(encoding="utf-8"))
    routes = json.loads(arguments.known_routes.read_text(encoding="utf-8"))
    session = get_session_factory()()
    try:
        report = evaluate_recomposed_workbooks(
            session,
            manifest=manifest,
            known_route_fingerprints={str(route["route_fingerprint"]) for route in routes},
        )
    finally:
        session.rollback()
        session.close()
    arguments.report.parent.mkdir(parents=True, exist_ok=True)
    arguments.report.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report["metrics"], ensure_ascii=False))


if __name__ == "__main__":
    main()
